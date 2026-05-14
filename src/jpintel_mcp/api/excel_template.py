"""Excel application-estimate template generator.

Wave 35 Axis 6b (2026-05-12). Surfaces a single
``POST /v1/excel/application_estimate`` endpoint that renders a 5-sheet
XLSX workbook from the live ``programs`` row + adoption stats using the
pure-Python ``openpyxl`` library.

Sheets
------
1. 制度サマリ  — overview + 採択率 + 上限額 + 出典 URL
2. 必要書類 chk  — checklist row per likely required document
3. 金額目安 — historical adoption × 上限額から作る確認用 estimate
4. 期限 calendar — 12 ヶ月 deadline calendar with the program's
   ``application_round`` rows
5. 補足 / 連絡先 — fence summary + Bookyou株式会社 contact

Design constraints (memory)
---------------------------
* No LLM API import — openpyxl only.
* Per-call price = 5 billable units (= ¥15).
* Anonymous callers rejected; per-key floor 60s.
* Brand: jpcite.
"""

from __future__ import annotations

import hashlib
import io
import logging
import os
import secrets
import sqlite3  # noqa: TC003
import time
from datetime import UTC, datetime, timedelta
from typing import Any

from fastapi import APIRouter, HTTPException, Request, status
from pydantic import AliasChoices, BaseModel, ConfigDict, Field

from jpintel_mcp.api.deps import (
    ApiContextDep,
    DbDep,
    log_usage,
    require_metered_api_key,
)
from jpintel_mcp.api.tax_rulesets import _TAX_DISCLAIMER

logger = logging.getLogger("jpintel.excel_template")

router = APIRouter(prefix="/v1/excel", tags=["excel"])

EXCEL_UNIT_COUNT = 5
EXCEL_URL_TTL_S = 7 * 24 * 3600
EXCEL_MIN_INTERVAL_S = 60

_excel_rate_state: dict[str, float] = {}


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------


class ExcelEstimateRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    include_value_estimate: bool = Field(
        True,
        description="Include value-estimate sheet",
        validation_alias=AliasChoices("include_value_estimate", "include_roi"),
    )
    include_calendar: bool = Field(True, description="Include 12-month deadline calendar")
    note: str | None = Field(None, max_length=500)


class ExcelEstimateResponse(BaseModel):
    excel_id: str
    program_id: str
    download_url: str
    expires_at: str
    byte_size: int
    sheet_count: int
    sha256: str
    disclaimer: str = Field(alias="_disclaimer")
    model_config = ConfigDict(populate_by_name=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _rate_floor_check(key_hash: str) -> None:
    now = time.monotonic()
    last = _excel_rate_state.get(key_hash)
    if last is not None and (now - last) < EXCEL_MIN_INTERVAL_S:
        retry_after = int(EXCEL_MIN_INTERVAL_S - (now - last)) + 1
        raise HTTPException(
            status.HTTP_429_TOO_MANY_REQUESTS,
            detail={
                "detail": "excel rate limit (1/minute per key)",
                "retry_after_s": retry_after,
            },
            headers={"Retry-After": str(retry_after)},
        )
    _excel_rate_state[key_hash] = now


def _load_program_substrate(conn: sqlite3.Connection, program_id: str) -> dict[str, Any]:
    """Pull the program row + adoption stats + application rounds."""
    out: dict[str, Any] = {
        "program": None,
        "adoption_stats": None,
        "application_rounds": [],
    }
    cur = conn.cursor()

    row = cur.execute(
        "SELECT * FROM programs WHERE program_id=? LIMIT 1",
        (program_id,),
    ).fetchone()
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"program_id={program_id} not found")
    out["program"] = dict(row)

    for tbl in ("adoption_records", "case_studies"):
        try:
            stat_row = cur.execute(
                f"SELECT COUNT(*) AS cnt FROM {tbl} WHERE program_id=?",  # noqa: S608
                (program_id,),
            ).fetchone()
            if stat_row and stat_row["cnt"]:
                out["adoption_stats"] = {
                    "source": tbl,
                    "count": int(stat_row["cnt"]),
                }
                break
        except sqlite3.Error:
            continue

    try:
        rows = cur.execute(
            "SELECT round_id, application_open_at, application_close_at, status "
            "FROM application_rounds WHERE program_id=? "
            "ORDER BY application_open_at LIMIT 12",
            (program_id,),
        ).fetchall()
        out["application_rounds"] = [dict(r) for r in rows]
    except sqlite3.Error:
        pass

    return out


def _render_xlsx(
    program_id: str,
    substrate: dict[str, Any],
    *,
    include_value_estimate: bool,
    include_calendar: bool,
    note: str | None,
) -> tuple[bytes, int]:
    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
        from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(
            status.HTTP_503_SERVICE_UNAVAILABLE,
            detail={"detail": "openpyxl not installed on this Fly machine"},
        ) from exc

    wb = Workbook()
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)

    s1 = wb.active
    s1.title = "制度サマリ"
    s1.append(["項目", "内容"])
    p = substrate["program"]
    for k in (
        "program_id",
        "title",
        "authority",
        "tier",
        "source_url",
        "summary",
        "max_amount_yen",
        "source_fetched_at",
    ):
        if k in p:
            s1.append([k, str(p.get(k, ""))])
    for cell in s1[1]:
        cell.fill = header_fill
        cell.font = header_font
    s1.column_dimensions["A"].width = 22
    s1.column_dimensions["B"].width = 88

    s2 = wb.create_sheet("必要書類chk")
    s2.append(["書類", "提出要否", "備考"])
    docs = [
        ("登記事項証明書 (履歴事項全部証明書)", "必要", "原則 3 ヶ月以内"),
        ("決算書 (直近 2 期)", "多くの制度で必要", "貸借対照表・損益計算書"),
        ("法人税確定申告書 (別表一・別表四)", "多くの制度で必要", "受付印付き"),
        ("会社案内・パンフレット", "推奨", "事業内容の説明"),
        ("見積書・カタログ", "設備投資型で必要", "複数業者で相見積"),
        ("事業計画書", "ほぼ全制度で必要", "様式は制度毎"),
        ("収支計画書 / 資金繰り表", "ほぼ全制度で必要", "3-5 年"),
        ("市町村税の納税証明書", "ほぼ全制度で必要", "未納がないこと"),
        ("適格請求書発行事業者登録通知 (T 番号)", "推奨", "事業者証明"),
        ("社会保険・労働保険 加入証明", "雇用関連制度で必要", "社労士確認推奨"),
    ]
    for d in docs:
        s2.append(list(d))
    for cell in s2[1]:
        cell.fill = header_fill
        cell.font = header_font
    s2.column_dimensions["A"].width = 38
    s2.column_dimensions["B"].width = 18
    s2.column_dimensions["C"].width = 32

    if include_value_estimate:
        s3 = wb.create_sheet("金額目安")
        s3.append(["指標", "値"])
        max_amt = p.get("max_amount_yen") or 0
        adoption_n = (substrate.get("adoption_stats") or {}).get("count", 0)
        s3.append(["上限額 (円)", max_amt])
        s3.append(["過去採択件数", adoption_n])
        if isinstance(max_amt, (int, float)) and max_amt > 0:
            est_grant = int(max_amt * 0.5)
            est_own = int(max_amt * 0.5)
            s3.append(["推定 補助額 (50% 補助率)", est_grant])
            s3.append(["推定 自己負担額", est_own])
            s3.append(["補助額 / 自己負担額", f"{round(est_grant / max(est_own, 1), 2)}x"])
        else:
            s3.append(["注記", "上限額が未確定のため金額目安なし"])
        for cell in s3[1]:
            cell.fill = header_fill
            cell.font = header_font
        s3.column_dimensions["A"].width = 32
        s3.column_dimensions["B"].width = 22

    if include_calendar:
        s4 = wb.create_sheet("期限calendar")
        s4.append(["round_id", "受付開始", "受付終了", "状態"])
        rounds = substrate.get("application_rounds", [])
        if rounds:
            for r in rounds:
                s4.append(
                    [
                        str(r.get("round_id", "")),
                        str(r.get("application_open_at", "")),
                        str(r.get("application_close_at", "")),
                        str(r.get("status", "")),
                    ]
                )
        else:
            today = datetime.now(UTC).date()
            for i in range(12):
                month_start = (today.replace(day=1) + timedelta(days=32 * i)).replace(day=1)
                s4.append(
                    [
                        f"placeholder-{i + 1}",
                        month_start.isoformat(),
                        (month_start + timedelta(days=27)).isoformat(),
                        "予定未公開 — 制度ページを直接確認のこと",
                    ]
                )
        for cell in s4[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col in ("A", "B", "C", "D"):
            s4.column_dimensions[col].width = 24

    s5 = wb.create_sheet("補足連絡先")
    s5.append(["項目", "内容"])
    s5.append(["発行", "Bookyou株式会社 (T8010001213708)"])
    s5.append(["連絡先", "info@bookyou.net"])
    s5.append(["商号", "jpcite"])
    s5.append(["免責", _TAX_DISCLAIMER])
    if note:
        s5.append(["備考", note[:500]])
    s5.append(["", ""])
    s5.append(["8 業法フェンス", ""])
    s5.append(["税理士法 §52", "税務代理・税務書類作成・税務相談は税理士のみ"])
    s5.append(["行政書士法 §1の2", "申請書類の代理作成は行政書士業務"])
    s5.append(["弁護士法 §72", "法律事件の代理・鑑定は弁護士業務"])
    s5.append(["社労士法 §27", "労務代理・36協定は社労士業務"])
    s5.append(["公認会計士法 §47", "監査証明は公認会計士業務"])
    for cell in s5[1]:
        cell.fill = header_fill
        cell.font = header_font
    s5.column_dimensions["A"].width = 22
    s5.column_dimensions["B"].width = 88

    buf = io.BytesIO()
    wb.save(buf)
    blob = buf.getvalue()
    sheet_count = len(wb.sheetnames)
    return blob, sheet_count


def _upload_to_r2(program_id: str, blob: bytes) -> tuple[str, str, str]:
    yyyymm = datetime.now(UTC).strftime("%Y%m")
    excel_id = f"xl_{secrets.token_hex(8)}"
    key = f"excel_templates/{program_id}/{yyyymm}/{excel_id}.xlsx"
    expires_at = (datetime.now(UTC) + timedelta(seconds=EXCEL_URL_TTL_S)).isoformat()
    if not os.environ.get("R2_ENDPOINT"):
        local = f"/tmp/excel_template/{excel_id}.xlsx"  # noqa: S108
        os.makedirs(os.path.dirname(local), exist_ok=True)
        with open(local, "wb") as fh:
            fh.write(blob)
        return key, f"/v1/excel/inline/{excel_id}", expires_at
    try:
        from pathlib import Path

        from scripts.cron._r2_client import upload

        tmp = Path(f"/tmp/{excel_id}.xlsx")  # noqa: S108
        tmp.write_bytes(blob)
        upload(tmp, key)
        tmp.unlink(missing_ok=True)
    except (ImportError, RuntimeError) as exc:
        logger.warning("r2 upload fell back to inline: %s", exc)
        local = f"/tmp/excel_template/{excel_id}.xlsx"  # noqa: S108
        os.makedirs(os.path.dirname(local), exist_ok=True)
        with open(local, "wb") as fh:
            fh.write(blob)
        return key, f"/v1/excel/inline/{excel_id}", expires_at
    base = os.environ.get(
        "R2_PUBLIC_BASE",
        f"https://{os.environ.get('R2_BUCKET', 'autonomath-backup')}.r2.cloudflarestorage.com",
    )
    return key, f"{base.rstrip('/')}/{key}", expires_at


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.post(
    "/application_estimate", response_model=ExcelEstimateResponse, response_model_by_alias=True
)
async def application_estimate(
    request: Request,
    program_id: str,
    body: ExcelEstimateRequest | None = None,
    ctx: ApiContextDep = None,  # type: ignore[assignment]
    db: DbDep = None,  # type: ignore[assignment]
) -> ExcelEstimateResponse:
    """Render a 5-sheet Excel application-estimate workbook."""
    if not program_id or len(program_id) > 128:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "invalid program_id")

    require_metered_api_key(ctx, "excel.application_estimate")
    _rate_floor_check(ctx.key_hash or "")

    body = body or ExcelEstimateRequest()
    substrate = _load_program_substrate(db, program_id)
    blob, sheet_count = _render_xlsx(
        program_id,
        substrate,
        include_value_estimate=body.include_value_estimate,
        include_calendar=body.include_calendar,
        note=body.note,
    )
    sha256 = hashlib.sha256(blob).hexdigest()
    excel_id = f"xl_{secrets.token_hex(8)}"
    _r2_key, download_url, expires_at = _upload_to_r2(program_id, blob)

    log_usage(
        db,
        ctx,
        endpoint="excel.application_estimate",
        quantity=EXCEL_UNIT_COUNT,
        request_id=getattr(request.state, "request_id", None),
        strict_metering=True,
    )
    logger.info(
        "excel.application_estimate program_id=%s sheets=%d bytes=%d",
        program_id,
        sheet_count,
        len(blob),
    )
    return ExcelEstimateResponse(
        excel_id=excel_id,
        program_id=program_id,
        download_url=download_url,
        expires_at=expires_at,
        byte_size=len(blob),
        sheet_count=sheet_count,
        sha256=sha256,
        _disclaimer=_TAX_DISCLAIMER,
    )


@router.get("/inline/{excel_id}")
async def inline_excel(excel_id: str) -> Any:
    """Local-disk fallback download (used when R2 env is not configured)."""
    from fastapi.responses import FileResponse

    if not excel_id.startswith("xl_") or "/" in excel_id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "invalid excel_id")
    path = f"/tmp/excel_template/{excel_id}.xlsx"  # noqa: S108
    if not os.path.exists(path):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "excel expired or not staged")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{excel_id}.xlsx",
    )
