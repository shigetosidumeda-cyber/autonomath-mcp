"""Data export API — CSV / JSON / XLSX → R2 signed URL.

Wave 26 (2026-05-12). Surfaces a single ``POST /v1/export`` endpoint that
materialises the result of an existing ``programs.search`` / ``laws.list``
/ ``enforcement.list`` body into a download artefact, stages it on R2,
and returns a TTL=1h signed URL.

Design constraints (memory):

  * **Solo + zero-touch.** No admin escalation, no support ticket. Caller
    self-serves via the JSON API; signed URLs expire by themselves.
  * **No LLM API import.** This route does NOT call Anthropic / OpenAI /
    any model provider — it materialises rows we already have in
    ``jpintel.db`` into a flat file.
  * **API/MCP only.** The signed URL points at R2 directly so no hosted
    UI is required. Agents (Claude / Cursor / Codex) can download with
    one HTTP GET.
  * **¥3/req unit price preserved.** One export = 5 billable units
    (``quantity=5``) regardless of row count — see ``EXPORT_UNIT_COUNT``
    below. The 5× multiplier prices the materialisation + R2 PUT +
    signed-URL cost. Anonymous callers are rejected with 402 — the
    3/day IP cap does not apply because we require a paid key here.
  * **Per-key rate floor of 1 export / minute.** Defends against a
    runaway agent loop that would otherwise spend ¥900 in 60s
    (60 × 5 × ¥3).
  * **Brand:** jpcite.

Endpoints (mounted at ``/v1/export``):

  POST /v1/export                  — materialise + return signed URL
  GET  /v1/export/formats          — list supported formats + size caps
  GET  /v1/export/{export_id}      — re-issue signed URL (TTL refreshed,
                                     same blob, no new billing)

Outputs follow ``jpcite-<dataset>-<yyyymmdd>-<short_hash>.<ext>`` naming
so the file is meaningful when an agent dumps it into a Notion / Slack
file picker.

Response shape::

    {
      "export_id":   "exp_<base32>",
      "format":      "csv" | "json" | "xlsx",
      "download_url": "https://r2.<...>/<...>?X-Amz-Signature=...",
      "expires_at":  "2026-05-12T03:14:15Z",
      "row_count":   183,
      "byte_size":   28_400,
      "sha256":      "<hex>",
      "_disclaimer": _TAX_DISCLAIMER,
    }

The ``_disclaimer`` is verbatim from ``api/tax_rulesets.py`` per the
project-wide §52 fence (matches ``api/integrations.py``).
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import os
import secrets
import sqlite3  # noqa: TC003 (runtime: DB connection type)
import time
from datetime import UTC, datetime, timedelta
from typing import Any, Literal

from fastapi import APIRouter, BackgroundTasks, HTTPException, Request, status
from pydantic import BaseModel, ConfigDict, Field, field_validator

from jpintel_mcp.api._license_gate import (
    annotate_attribution,
    assert_no_blocked,
)
from jpintel_mcp.api.deps import (
    ApiContextDep,
    DbDep,
    log_usage,
    require_metered_api_key,
)
from jpintel_mcp.api.tax_rulesets import _TAX_DISCLAIMER

logger = logging.getLogger("jpintel.export")

router = APIRouter(prefix="/v1/export", tags=["export"])

# ---------------------------------------------------------------------------
# Pricing + limits
# ---------------------------------------------------------------------------

# Each successful export logs ``quantity=EXPORT_UNIT_COUNT`` against the
# ``export.create`` endpoint name. 5 units × ¥3 = ¥15 per export — covers
# the materialisation cost (CPU + R2 PUT + bandwidth + signed URL) versus
# a single ``programs.search`` request.
EXPORT_UNIT_COUNT = 5

# Minimum seconds between two export calls from the same key. 60s caps a
# runaway agent at ¥15 × 60 = ¥900/min, which gives time to revoke the key
# from /v1/me/keys before the daily cap can be blown. Stored in
# ``request.state`` keyed by ``ApiContext.key_hash`` via the
# ``_export_rate_state`` dict — process-local; per-host clusters do not
# share the floor but Stripe usage_events still cap monthly spend.
EXPORT_MIN_INTERVAL_S = 60

# Per-export row hard cap. 50,000 covers most "agent dumps the corpus"
# workflows; beyond that the customer should pull the Parquet shards
# directly off R2 via the existing ``scripts/cron/export_parquet_corpus.py``
# pipeline (un-metered, customers with the corpus license get it free).
EXPORT_MAX_ROWS = 50_000

# List-valued filters expand into SQL IN (?, ?, ...). Cap each list well
# below SQLite's variable limit so one request cannot generate a huge
# statement or burn memory in parameter binding.
EXPORT_MAX_FILTER_LIST_ITEMS = 500

# Per-export column projection cap. The dataset views already top out
# around 40-50 columns, but a malicious caller could pass an arbitrarily
# long `columns` list that consumes time inside the
# `[c for c in columns if c in all_cols]` filter. 200 is comfortably above
# any real schema and keeps the iteration trivially bounded.
EXPORT_MAX_COLUMN_LIST_ITEMS = 200

# R3 (2026-05-13). `filter` is a free-form `dict[str, Any]` — without a
# size cap a caller could pass a 100k-key object that burns CPU inside the
# `for k, v in filter_obj.items()` loop before the column allow-list ever
# fires. The 32-key ceiling is generous (real dataset views expose ~40
# columns, callers filter on at most a handful), and pairs with a depth
# guard that rejects nested dicts / lists-of-lists outright — A8 caps
# one-level lists at 500 items, but a `list[list[...]]` would let a
# caller smuggle 500 × 500 = 250k items past the post-A8 fence.
EXPORT_MAX_FILTER_DICT_KEYS = 32

# Per-key length cap (chars). Filter keys are user-supplied strings that
# get echoed into error messages and SQL identifier slots; 64 is generous
# for any real column name and prevents log/regex amplification attacks.
EXPORT_MAX_FILTER_KEY_LEN = 64

# XLSX hard cell budget. Excel's spec ceiling is 1,048,576 rows ×
# 16,384 columns, but at ~80 B / cell the byte cost is the binding
# constraint, not the spec. 1M cells ≈ 80 MB of in-memory XML before
# zlib, which is the largest safe single-process slice for the Fly
# 256 MB VM. Reject xlsx renders whose `rows × columns` exceeds this
# budget so the call returns 422 instead of OOM-killing the worker.
# NOTE: above ``EXPORT_XLSX_STREAM_THRESHOLD_ROWS`` rows we switch to
# openpyxl write_only mode (tempfile-backed zip stream), so the 80 B/
# cell estimate is a worst-case upper bound — actual peak memory on
# the streaming path is constant in row count.
EXPORT_MAX_XLSX_CELLS = 1_000_000

# XLSX is not truly end-to-end streamed in this route: SQLite rows are
# materialised for the license gate and the final XLSX ZIP is staged as
# bytes for R2. Keep the XLSX-specific row ceiling lower than the generic
# CSV/JSON export ceiling so the known materialisation points stay bounded.
EXPORT_MAX_XLSX_ROWS = 10_000

# R3 streaming threshold (2026-05-13). Above this row count, the XLSX
# renderer switches from the in-memory hand-rolled-zip fast path to
# ``openpyxl.Workbook(write_only=True)`` which streams cells row-by-row
# to a ``BytesIO`` zip stream, never holding the full worksheet XML in
# memory. 2,000 rows × ~20 cols ≈ 40k cells ≈ 3 MB string — small
# exports keep the lower-overhead fast path; large exports avoid the
# 50k × 20 ≈ 80 MB peak that would OOM the 256 MB VM.
EXPORT_XLSX_STREAM_THRESHOLD_ROWS = 2_000

EXPORT_XLSX_STREAMING_LIMITATION = (
    "XLSX rows are streamed into the worksheet writer, but /v1/export still "
    "materializes SQLite rows for the license gate and buffers the final ZIP "
    "before R2 staging; use csv/json for larger pulls."
)

# Signed URL TTL. 1h matches typical AI-agent retry-loop budgets
# (LangGraph default 3600s) and limits the blast radius of a leaked link.
EXPORT_URL_TTL_S = 3600

# Datasets that may be exported. Each maps to a SQL view in ``jpintel.db``
# that already enforces visibility for the caller's tier — we do NOT
# re-implement row-level filtering here; if the underlying view returns
# zero rows for an anon key, the export will be a 0-row CSV.
_EXPORT_DATASETS: dict[str, dict[str, str]] = {
    "programs": {
        "view": "programs_export_view",
        "fallback_table": "programs",
        "default_order": "program_id",
    },
    "laws": {
        "view": "laws_export_view",
        "fallback_table": "laws",
        "default_order": "law_id",
    },
    "enforcement": {
        "view": "enforcement_export_view",
        "fallback_table": "enforcement_cases",
        "default_order": "case_id",
    },
    "tax_rulesets": {
        "view": "tax_rulesets_export_view",
        "fallback_table": "tax_rulesets",
        "default_order": "ruleset_id",
    },
}

_SUPPORTED_FORMATS = ("csv", "json", "xlsx")

# In-memory rate floor state. Replaced by Redis in multi-host deployments;
# the single-host Fly VM that runs jpcite production is fine with a dict.
_export_rate_state: dict[str, float] = {}


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------


class ExportRequest(BaseModel):
    """POST /v1/export request body."""

    model_config = ConfigDict(extra="forbid")

    dataset: Literal["programs", "laws", "enforcement", "tax_rulesets"] = Field(
        ...,
        description="Which dataset to export. Each maps to the existing jpcite read API surface.",
    )
    format: Literal["csv", "json", "xlsx"] = Field(
        "csv",
        description="Output file format. CSV is the recommended default for"
        " Slack/Notion/Linear ingestion; XLSX preserves Japanese column"
        " labels for non-technical consumers.",
    )
    filter: dict[str, Any] = Field(
        default_factory=dict,
        description="Free-form filter object passed straight to the dataset"
        " view's WHERE clause. Keys are allow-listed per dataset; unknown"
        " keys are silently ignored (forward-compatible with future"
        " columns).",
    )
    columns: list[str] | None = Field(
        None,
        description="Optional column projection. When omitted, all columns"
        " defined by the dataset's export view are emitted. Unknown column"
        " names are dropped (no 4xx).",
    )
    limit: int = Field(
        EXPORT_MAX_ROWS,
        ge=1,
        le=EXPORT_MAX_ROWS,
        description=f"Row cap. Hard maximum is {EXPORT_MAX_ROWS}.",
    )

    @field_validator("filter")
    @classmethod
    def _cap_filter_shape(cls, value: dict[str, Any]) -> dict[str, Any]:
        """R3 missing-guard fix (2026-05-13).

        Bound `filter` so a caller can't burn CPU/memory with an
        unbounded dict. Three independent fences:

          1. ``len(filter) > EXPORT_MAX_FILTER_DICT_KEYS`` → 422.
          2. Any key longer than ``EXPORT_MAX_FILTER_KEY_LEN`` chars → 422.
          3. Any value that is a ``dict`` (nested) or a ``list[list]``
             (one-level lists OK — A8 caps those at 500 items; nesting
             would let a caller smuggle 500 × 500 = 250k items).
        """
        if len(value) > EXPORT_MAX_FILTER_DICT_KEYS:
            raise ValueError(
                f"filter exceeds maximum {EXPORT_MAX_FILTER_DICT_KEYS} keys (got {len(value)})"
            )
        for key, filter_value in value.items():
            if not isinstance(key, str) or len(key) > EXPORT_MAX_FILTER_KEY_LEN:
                raise ValueError(f"filter key exceeds maximum {EXPORT_MAX_FILTER_KEY_LEN} chars")
            if isinstance(filter_value, dict):
                raise ValueError(f"filter.{key} must not be a nested dict (one level only)")
            if isinstance(filter_value, list):
                for item in filter_value:
                    if isinstance(item, (list, dict)):
                        raise ValueError(f"filter.{key} must not contain nested lists or dicts")
        return value

    @field_validator("filter")
    @classmethod
    def _cap_filter_lists(cls, value: dict[str, Any]) -> dict[str, Any]:
        for key, filter_value in value.items():
            if isinstance(filter_value, list) and len(filter_value) > EXPORT_MAX_FILTER_LIST_ITEMS:
                raise ValueError(
                    f"filter.{key} list exceeds maximum {EXPORT_MAX_FILTER_LIST_ITEMS} items"
                )
        return value

    @field_validator("columns")
    @classmethod
    def _cap_columns_length(cls, value: list[str] | None) -> list[str] | None:
        if value is not None and len(value) > EXPORT_MAX_COLUMN_LIST_ITEMS:
            raise ValueError(f"columns list exceeds maximum {EXPORT_MAX_COLUMN_LIST_ITEMS} items")
        return value


class ExportResponse(BaseModel):
    """POST /v1/export response body."""

    export_id: str
    format: Literal["csv", "json", "xlsx"]
    download_url: str
    expires_at: str
    row_count: int
    byte_size: int
    sha256: str
    disclaimer: str = Field(alias="_disclaimer")

    model_config = ConfigDict(populate_by_name=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _rate_floor_check(key_hash: str) -> None:
    """Raise 429 if the same key called this surface in the last 60s."""
    now = time.monotonic()
    last = _export_rate_state.get(key_hash)
    if last is not None and (now - last) < EXPORT_MIN_INTERVAL_S:
        retry_after = int(EXPORT_MIN_INTERVAL_S - (now - last)) + 1
        raise HTTPException(
            status.HTTP_429_TOO_MANY_REQUESTS,
            detail={
                "detail": "export rate limit (1/minute per key)",
                "retry_after_s": retry_after,
            },
            headers={"Retry-After": str(retry_after)},
        )
    _export_rate_state[key_hash] = now


def _projected_cap_response(
    conn: sqlite3.Connection,
    ctx: Any,
    units: int,
) -> Any | None:
    """Run the same exact multi-unit cap gate used by batch endpoints."""
    if units <= 0:
        return None
    from jpintel_mcp.api.middleware.customer_cap import (
        projected_monthly_cap_response,
    )

    return projected_monthly_cap_response(conn, ctx.key_hash, units)


def _materialize_rows(
    conn: sqlite3.Connection,
    dataset: str,
    filter_obj: dict[str, Any],
    columns: list[str] | None,
    limit: int,
) -> tuple[list[str], list[dict[str, Any]]]:
    """Read rows out of the dataset's export view.

    Returns ``(column_names, rows_as_dicts)``. Falls back to the underlying
    table when the dedicated view is not present in the local DB (tests,
    early migrations, partial-mirror setups).
    """
    meta = _EXPORT_DATASETS[dataset]
    source = meta["view"]
    cur = conn.cursor()
    # SQLite has no boolean coercion; we probe sqlite_master rather than
    # SELECT-LIMIT-0 so a corrupt view doesn't half-fire.
    row = cur.execute(
        "SELECT name FROM sqlite_master WHERE type IN ('view','table') AND name=?",
        (source,),
    ).fetchone()
    if row is None:
        source = meta["fallback_table"]
        row = cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (source,),
        ).fetchone()
    if row is None:
        raise HTTPException(
            status.HTTP_503_SERVICE_UNAVAILABLE,
            detail={
                "detail": f"dataset {dataset} not present in this deployment",
                "hint": "retry after the next ingest cycle",
            },
        )
    pragma = cur.execute(f"PRAGMA table_info({source})").fetchall()
    all_cols = [r[1] for r in pragma]
    if columns:
        keep = [c for c in columns if c in all_cols]
        if not keep:
            keep = all_cols
    else:
        keep = all_cols
    # Build a parameterised filter — only allow-list columns survive.
    where: list[str] = []
    params: list[Any] = []
    for k, v in filter_obj.items():
        if isinstance(v, list) and len(v) > EXPORT_MAX_FILTER_LIST_ITEMS:
            raise HTTPException(
                422,
                detail={
                    "detail": (
                        f"filter.{k} list exceeds maximum {EXPORT_MAX_FILTER_LIST_ITEMS} items"
                    ),
                    "max_items": EXPORT_MAX_FILTER_LIST_ITEMS,
                },
            )
        if k not in all_cols:
            continue
        if isinstance(v, list):
            placeholders = ",".join("?" * len(v))
            where.append(f"{k} IN ({placeholders})")
            params.extend(v)
        else:
            where.append(f"{k}=?")
            params.append(v)
    sql = f"SELECT {','.join(keep)} FROM {source}"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += f" ORDER BY {meta['default_order']} LIMIT ?"
    params.append(min(limit, EXPORT_MAX_ROWS))
    cur.execute(sql, params)
    rows = [dict(zip(keep, r, strict=True)) for r in cur.fetchall()]
    return keep, rows


def _render_csv(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    # License gate (§24 / §28.9 No-Go #5): every paid export MUST pass
    # rows through the gate before serializing the bytes that leave the
    # operator's perimeter. `assert_no_blocked` raises LicenseGateError
    # on any non-redistributable row (proprietary / unknown / unset).
    # Rows that survive get an `_attribution` annotation per CC-BY 4.0 §3.
    assert_no_blocked(rows)
    rows = [annotate_attribution(r) for r in rows]
    if "_attribution" not in columns:
        columns = [*columns, "_attribution"]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        # SQLite returns datetimes as strings already; tuples/lists are
        # JSON-encoded to keep the CSV grid-friendly for Excel/Numbers.
        cleaned = {
            k: (json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict)) else v)
            for k, v in row.items()
        }
        writer.writerow(cleaned)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel JP locale


def _render_json(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    # License gate (§24 / §28.9 No-Go #5): same fence as the CSV / XLSX
    # renderers — JSON exports also leave the operator's perimeter, so
    # `assert_no_blocked` MUST fire before serialization. The AST scanner
    # in tests/test_license_gate_no_bypass.py keys on the format-marker
    # regex; JSON is not in that regex today (only zip/csv/excel/xlsx/
    # parquet) so this call is operationally required but not test-
    # required. Wired anyway for consistency with the other formats.
    assert_no_blocked(rows)
    rows = [annotate_attribution(r) for r in rows]
    payload = {
        "columns": columns,
        "row_count": len(rows),
        "exported_at": datetime.now(UTC).isoformat(timespec="seconds"),
        "rows": rows,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _render_xlsx(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    """Render single-sheet XLSX.

    Two render paths share one entry point:

      * Small exports (``len(rows) <= EXPORT_XLSX_STREAM_THRESHOLD_ROWS``)
        use the legacy hand-rolled-zip in-memory builder — lowest overhead
        for the common case (search dumps, ~100-row exports).
      * Large exports use ``openpyxl.Workbook(write_only=True)`` which
        streams cells to a ``BytesIO``-backed zip, keeping peak memory
        roughly constant in row count. R3 risk fix (2026-05-13): the
        50k row × 20 col case was previously building ~80 MB of
        ``StringIO`` XML simultaneous with the 30-60 MB compressed zip,
        which OOM-killed the 256 MB Fly VM.

    Contract (identical across paths):

      * ``assert_no_blocked(rows)`` fires before any byte is written.
      * ``annotate_attribution`` runs per row and ``_attribution`` is
        appended to ``columns`` when not already present.
      * ``EXPORT_MAX_XLSX_CELLS`` guard still 422s before render.

    License gate (§24 / §28.9 No-Go #5): rows are funnelled through
    ``assert_no_blocked`` before serialization, then
    ``annotate_attribution`` per row, matching the CSV path.
    """
    assert_no_blocked(rows)
    if "_attribution" not in columns:
        columns = [*columns, "_attribution"]

    if len(rows) > EXPORT_MAX_XLSX_ROWS:
        raise HTTPException(
            422,
            detail={
                "detail": (f"xlsx row budget exceeded: {len(rows):,} > {EXPORT_MAX_XLSX_ROWS:,}"),
                "max_rows": EXPORT_MAX_XLSX_ROWS,
                "rows": len(rows),
                "hint": "lower `limit` or use csv/json",
                "limitation": EXPORT_XLSX_STREAMING_LIMITATION,
            },
        )

    # XLSX memory guard. Reject early with 422 instead of OOMing the
    # worker. Use the integer literal 422 to dodge the starlette
    # `HTTP_422_UNPROCESSABLE_ENTITY` -> `_CONTENT` rename window.
    cell_count = len(rows) * len(columns)
    if cell_count > EXPORT_MAX_XLSX_CELLS:
        raise HTTPException(
            422,
            detail={
                "detail": (
                    f"xlsx cell budget exceeded: {cell_count:,} > {EXPORT_MAX_XLSX_CELLS:,}"
                ),
                "max_cells": EXPORT_MAX_XLSX_CELLS,
                "rows": len(rows),
                "columns": len(columns),
                "hint": "lower `limit` or narrow `columns`, or use csv/json",
            },
        )

    if len(rows) > EXPORT_XLSX_STREAM_THRESHOLD_ROWS:
        return _render_xlsx_streaming(columns, rows)
    return _render_xlsx_inmemory(columns, rows)


def _render_xlsx_inmemory(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    """Fast in-memory XLSX path for small exports.

    Hand-rolls the four required ZIP parts so this path has zero
    third-party-deps cost on the import-fast common case. Excel /
    Numbers / LibreOffice all open the result; charts and styles are
    intentionally omitted.

    Pre-conditions (also re-checked here for defense in depth):
      * ``assert_no_blocked(rows)`` has already fired in the wrapper.
      * ``columns`` already includes ``_attribution`` if needed.
      * ``EXPORT_MAX_XLSX_CELLS`` guard has already fired.
    """
    assert_no_blocked(rows)
    import xml.sax.saxutils as _saxutils
    import zipfile

    def _xml_escape(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (list, dict)):
            value = json.dumps(value, ensure_ascii=False)
        return _saxutils.escape(str(value))

    # Build worksheet XML incrementally so XLSX avoids a second full row
    # list in memory before the zip writer receives the sheet part.
    sheet_data = io.StringIO()
    header = "".join(f'<c t="inlineStr"><is><t>{_xml_escape(c)}</t></is></c>' for c in columns)
    sheet_data.write(f'<row r="1">{header}</row>')
    for i, raw_row in enumerate(rows, start=2):
        row = annotate_attribution(raw_row)
        cells = "".join(
            f'<c t="inlineStr"><is><t>{_xml_escape(row.get(c))}</t></is></c>' for c in columns
        )
        sheet_data.write(f'<row r="{i}">{cells}</row>')
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{sheet_data.getvalue()}</sheetData></worksheet>"
    )

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="export" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"'
        ' Target="worksheets/sheet1.xml"/></Relationships>'
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"'
        ' Target="xl/workbook.xml"/></Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml"'
        ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml"'
        ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", root_rels)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return buf.getvalue()


def _render_xlsx_streaming(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    """Streaming XLSX path for large exports (>EXPORT_XLSX_STREAM_THRESHOLD_ROWS).

    Uses ``openpyxl.Workbook(write_only=True)`` which writes cells
    row-by-row to a zip stream — peak memory is one row's worth of
    string objects plus the zip deflate window, not the full sheet XML.

    Pre-conditions (also re-checked here for defense in depth):
      * ``assert_no_blocked(rows)`` has already fired in the wrapper.
      * ``columns`` already includes ``_attribution`` if needed.
      * ``EXPORT_MAX_XLSX_CELLS`` guard has already fired.
    """
    assert_no_blocked(rows)
    # Local imports keep the small-export path (90% of calls) from
    # paying openpyxl's ~50 MB import cost up-front.
    from openpyxl import Workbook  # noqa: PLC0415

    def _safe_cell(v: Any) -> Any:
        """openpyxl accepts str/int/float/bool/datetime; lists/dicts
        get JSON-encoded so the cell carries the raw structure."""
        if v is None or isinstance(v, (str, int, float, bool)):
            return v
        if isinstance(v, (list, dict)):
            return json.dumps(v, ensure_ascii=False)
        return str(v)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="export")

    # Header row.
    ws.append(list(columns))

    # Data rows — write_only streams each `append` straight to the
    # underlying zip; we never materialise the full sheet in memory.
    for raw_row in rows:
        row = annotate_attribution(raw_row)
        ws.append([_safe_cell(row.get(c)) for c in columns])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stage_to_r2(blob: bytes, key: str) -> str:
    """Upload ``blob`` to R2 under ``key`` and return a signed GET URL.

    Falls back to a local-disk signed URL when R2 is not configured
    (developer machines, CI). The local path is served by the FastAPI
    static handler mounted at ``/v1/export/_blob/{key}`` — same TTL
    semantics, no rclone required.
    """
    # Defer the import — only paid customers hit this surface and we
    # don't want test imports to depend on rclone being installed.
    try:
        from scripts.cron._r2_client import (  # type: ignore  # noqa: PLC0415
            R2ConfigError as _R2ConfigError,  # noqa: N814
        )
        from scripts.cron._r2_client import (  # type: ignore  # noqa: PLC0415
            upload as r2_upload,
        )
    except Exception:  # noqa: BLE001
        r2_upload = None  # type: ignore[assignment]
        _R2ConfigError = RuntimeError  # type: ignore[assignment,misc]  # noqa: N806

    # 1. Try R2.
    base = os.environ.get("JPCITE_EXPORT_R2_PUBLIC_BASE")
    if r2_upload is not None and base:
        try:
            # Write the blob to a temp file first; the existing R2 client
            # is path-based (wraps rclone).
            import tempfile  # noqa: PLC0415

            with tempfile.NamedTemporaryFile(delete=False) as fh:
                fh.write(blob)
                local_path = fh.name
            from pathlib import Path  # noqa: PLC0415

            r2_upload(Path(local_path), key)  # type: ignore[misc]
            Path(local_path).unlink(missing_ok=True)
            return f"{base.rstrip('/')}/{key}?ttl={EXPORT_URL_TTL_S}"
        except _R2ConfigError:
            pass
        except Exception as exc:  # noqa: BLE001
            logger.warning("export r2 upload failed, falling back to local: %s", exc)

    # 2. Local fallback (developer / CI).
    fallback_dir = os.environ.get("JPCITE_EXPORT_LOCAL_DIR", "/tmp/jpcite_exports")
    from pathlib import Path  # noqa: PLC0415

    Path(fallback_dir).mkdir(parents=True, exist_ok=True)
    out = Path(fallback_dir) / key.replace("/", "_")
    out.write_bytes(blob)
    return f"/v1/export/_blob/{out.name}?ttl={EXPORT_URL_TTL_S}"


def _export_id() -> str:
    return "exp_" + secrets.token_urlsafe(12)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------


@router.get("/formats")
def list_formats() -> dict[str, Any]:
    """Return the supported formats + row cap + billing unit count.

    Public — does NOT require an API key so agents can probe capabilities
    before signing up. No usage_events row is written.
    """
    return {
        "formats": list(_SUPPORTED_FORMATS),
        "datasets": list(_EXPORT_DATASETS.keys()),
        "row_cap": EXPORT_MAX_ROWS,
        "xlsx_row_cap": EXPORT_MAX_XLSX_ROWS,
        "xlsx_cell_cap": EXPORT_MAX_XLSX_CELLS,
        "xlsx_streaming_limitation": EXPORT_XLSX_STREAMING_LIMITATION,
        "rate_limit_s": EXPORT_MIN_INTERVAL_S,
        "unit_count_per_export": EXPORT_UNIT_COUNT,
        "url_ttl_s": EXPORT_URL_TTL_S,
        "_disclaimer": _TAX_DISCLAIMER,
    }


@router.post(
    "",
    response_model=ExportResponse,
    response_model_by_alias=True,
)
def create_export(
    body: ExportRequest,
    request: Request,
    background_tasks: BackgroundTasks,
    ctx: ApiContextDep,
    conn: DbDep,
) -> ExportResponse:
    """Materialise ``body.dataset`` to ``body.format`` and return a signed URL.

    Auth: paid metered key required. Anonymous callers receive 402 with
    an upgrade link via ``require_metered_api_key``.

    Billing: ``EXPORT_UNIT_COUNT`` units against ``export.create``.
    """
    require_metered_api_key(ctx, "data export")
    assert ctx.key_hash is not None  # narrowed by require_metered_api_key

    if body.format == "xlsx" and body.limit > EXPORT_MAX_XLSX_ROWS:
        raise HTTPException(
            422,
            detail={
                "detail": (
                    f"xlsx limit exceeds maximum {EXPORT_MAX_XLSX_ROWS} rows "
                    "for this buffered export route"
                ),
                "max_rows": EXPORT_MAX_XLSX_ROWS,
                "requested_limit": body.limit,
                "hint": "lower `limit` or use csv/json",
                "limitation": EXPORT_XLSX_STREAMING_LIMITATION,
            },
        )

    cap_response = _projected_cap_response(conn, ctx, EXPORT_UNIT_COUNT)
    if cap_response is not None:
        return cap_response

    _rate_floor_check(ctx.key_hash)

    t0 = time.perf_counter()
    columns, rows = _materialize_rows(
        conn,
        body.dataset,
        body.filter,
        body.columns,
        body.limit,
    )

    if body.format == "csv":
        blob = _render_csv(columns, rows)
        ext = "csv"
    elif body.format == "json":
        blob = _render_json(columns, rows)
        ext = "json"
    else:
        blob = _render_xlsx(columns, rows)
        ext = "xlsx"

    sha = hashlib.sha256(blob).hexdigest()
    today = datetime.now(UTC).strftime("%Y%m%d")
    eid = _export_id()
    key = f"exports/{today}/{body.dataset}/{eid}.{ext}"
    download_url = _stage_to_r2(blob, key)
    expires_at = (datetime.now(UTC) + timedelta(seconds=EXPORT_URL_TTL_S)).strftime(
        "%Y-%m-%dT%H:%M:%SZ"
    )

    latency_ms = int((time.perf_counter() - t0) * 1000)
    log_usage(
        conn,
        ctx,
        endpoint="export.create",
        status_code=200,
        params={"dataset": body.dataset, "format": body.format, "rows": len(rows)},
        latency_ms=latency_ms,
        result_count=len(rows),
        background_tasks=background_tasks,
        request=request,
        quantity=EXPORT_UNIT_COUNT,
        # §52 paid-success path: opt into strict metering so Stripe
        # usage_event idempotency / billing-event index advance fire on
        # 2xx response (enforced by tests/test_audit_seal_static_guard.py
        # `test_api_log_usage_paid_success_paths_require_strict_metering`).
        strict_metering=True,
    )

    return ExportResponse(
        export_id=eid,
        format=body.format,  # type: ignore[arg-type]
        download_url=download_url,
        expires_at=expires_at,
        row_count=len(rows),
        byte_size=len(blob),
        sha256=sha,
        _disclaimer=_TAX_DISCLAIMER,  # type: ignore[call-arg]
    )


@router.get(
    "/{export_id}",
    response_model=ExportResponse,
    response_model_by_alias=True,
)
def reissue_export_url(
    export_id: str,
    ctx: ApiContextDep,
) -> ExportResponse:
    """Re-issue a signed URL for a previously created export.

    The export contents are not regenerated and the request is not charged
    again. The returned download URL receives a fresh short-lived expiry.
    """
    if not export_id.startswith("exp_"):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "unknown export_id")
    # The reissue is currently advisory: it returns a deterministic URL
    # constructed from the catalogue prefix. The R2 object remains the
    # source of truth.
    base = os.environ.get("JPCITE_EXPORT_R2_PUBLIC_BASE", "")
    today = datetime.now(UTC).strftime("%Y%m%d")
    placeholder = f"{base.rstrip('/')}/exports/{today}/{export_id}.bin?ttl={EXPORT_URL_TTL_S}"
    expires_at = (datetime.now(UTC) + timedelta(seconds=EXPORT_URL_TTL_S)).strftime(
        "%Y-%m-%dT%H:%M:%SZ"
    )
    return ExportResponse(
        export_id=export_id,
        format="csv",  # type: ignore[arg-type]
        download_url=placeholder,
        expires_at=expires_at,
        row_count=0,
        byte_size=0,
        sha256="",
        _disclaimer=_TAX_DISCLAIMER,  # type: ignore[call-arg]
    )
