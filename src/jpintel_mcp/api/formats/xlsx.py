"""XLSX renderer (openpyxl write_only mode) — ``?format=xlsx``.

Two-sheet layout:
  Sheet ``data``  — header row 1 (frozen, wrapped) carries §52 disclaimer
                    so the workbook is self-disclosing even when copied.
                    Row 2 is the column header. Rows 3..N are data.
  Sheet ``_meta`` — license summary + per-row source_fetched_at + brand.

write_only mode (``openpyxl.Workbook(write_only=True)``) streams cells
straight to the underlying zip stream, keeping memory constant — important
for the ¥3/req batch case where one search call may spill 100 rows.
"""

from __future__ import annotations

import io
from typing import Any

from fastapi import HTTPException, status
from fastapi.responses import Response

from jpintel_mcp.api._format_dispatch import (
    BRAND_FOOTER,
    DISCLAIMER_HEADER_VALUE,
    DISCLAIMER_JA,
)

REQUIRED_COLUMNS: tuple[str, ...] = (
    "unified_id",
    "source_url",
    "source_fetched_at",
    "license",
)


def _column_order(rows: list[dict[str, Any]]) -> list[str]:
    keys: set[str] = set()
    for r in rows:
        keys.update(r.keys())
    rest = sorted(k for k in keys if k not in REQUIRED_COLUMNS)
    return [*REQUIRED_COLUMNS, *rest]


def _safe_cell(v: Any) -> Any:
    """Return an XLSX-safe primitive.

    openpyxl accepts str/int/float/bool/datetime. Lists / dicts get JSON-
    encoded inline so the cell still carries the raw structure (string
    form). None becomes None (Excel renders blank).
    """
    import json

    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def render_xlsx(rows: list[dict[str, Any]], meta: dict[str, Any]) -> Response:
    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
        from openpyxl.styles import Alignment, Font  # type: ignore[import-untyped]
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=(
                "format=xlsx requires the optional 'formats' extra: "
                "pip install autonomath-mcp[formats]"
            ),
        ) from exc

    columns = _column_order(rows)

    wb = Workbook(write_only=True)
    data_ws = wb.create_sheet(title="data")

    # Row 1 — disclaimer banner (single merged-looking cell at A1; we
    # do not actually merge in write_only mode — Excel renders the long
    # string with wrap_text turned on, which the consumer's row height
    # auto-adjusts to ~3 lines).
    bold_warn = Font(bold=True, color="9C0006")
    wrap = Alignment(wrap_text=True, vertical="top")
    from openpyxl.cell import WriteOnlyCell  # type: ignore[import-untyped]

    banner_cell = WriteOnlyCell(data_ws, value=DISCLAIMER_JA)
    banner_cell.font = bold_warn
    banner_cell.alignment = wrap
    # Pad the banner row out to len(columns) so the visual row width
    # matches the data band.
    data_ws.append([banner_cell] + [None] * (len(columns) - 1))

    # Row 2 — column headers.
    header_font = Font(bold=True)
    header_cells: list[Any] = []
    for c in columns:
        cell = WriteOnlyCell(data_ws, value=c)
        cell.font = header_font
        header_cells.append(cell)
    data_ws.append(header_cells)

    # write_only freeze: openpyxl supports freeze_panes on write_only
    # since 3.1; freeze under header so disclaimer + column header stay
    # pinned when the user scrolls.
    data_ws.freeze_panes = "A3"

    # Data rows.
    for row in rows:
        data_ws.append([_safe_cell(row.get(c)) for c in columns])

    # _meta sheet.
    meta_ws = wb.create_sheet(title="_meta")
    meta_ws.append(["key", "value"])
    meta_ws.append(["disclaimer", DISCLAIMER_JA])
    meta_ws.append(["brand", BRAND_FOOTER])
    meta_ws.append(["license_summary", meta.get("license_summary", "")])
    meta_ws.append(["endpoint", meta.get("endpoint", "")])
    meta_ws.append(["row_count", len(rows)])
    if meta.get("total") is not None:
        meta_ws.append(["total_unfiltered", meta.get("total")])
    if meta.get("as_of") is not None:
        meta_ws.append(["as_of", meta.get("as_of")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{meta.get('filename_stem', 'autonomath_export')}.xlsx"
    return Response(
        content=buf.read(),
        media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-AutonoMath-Disclaimer": DISCLAIMER_HEADER_VALUE,
            "X-AutonoMath-Format": "xlsx",
        },
    )
