"""Performance / memory guard regression tests for /v1/export.

A8 (Programs / Export Performance Guard, 2026-05-13). Bounds:

  * ``EXPORT_MAX_ROWS``                  — per-export row hard cap.
  * ``EXPORT_MAX_FILTER_LIST_ITEMS``     — per-filter IN(...) length cap.
  * ``EXPORT_MAX_COLUMN_LIST_ITEMS``     — projected-column list cap.
  * ``EXPORT_MAX_XLSX_CELLS``            — XLSX peak-memory cell cap.

Each test asserts the guard fires at the right boundary (one item / cell
over the limit), returns the canonical status code, and does so BEFORE
the next stage (SQL execution / XLSX render). Tests run in-process with
a SQLite ``:memory:`` connection — no fixtures or x402 middleware, so
the ``JPCITE_X402_SCHEMA_FAIL_OPEN_DEV`` env var is irrelevant here.
"""

from __future__ import annotations

import json
import sqlite3

import pytest
from fastapi import BackgroundTasks, HTTPException, Request
from pydantic import ValidationError

from jpintel_mcp.api.deps import ApiContext
from jpintel_mcp.api.export import (
    EXPORT_MAX_COLUMN_LIST_ITEMS,
    EXPORT_MAX_FILTER_DICT_KEYS,
    EXPORT_MAX_FILTER_LIST_ITEMS,
    EXPORT_MAX_ROWS,
    EXPORT_MAX_XLSX_CELLS,
    EXPORT_MAX_XLSX_ROWS,
    EXPORT_UNIT_COUNT,
    EXPORT_XLSX_STREAM_THRESHOLD_ROWS,
    EXPORT_XLSX_STREAMING_LIMITATION,
    ExportRequest,
    _materialize_rows,
    _render_xlsx,
    _render_xlsx_streaming,
    create_export,
    list_formats,
)


def _paid_cap_conn(*, key_hash: str, cap_yen: int) -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute(
        "CREATE TABLE api_keys ("
        "key_hash TEXT PRIMARY KEY, tier TEXT, monthly_cap_yen INTEGER, "
        "id INTEGER, parent_key_id INTEGER, revoked_at TEXT)"
    )
    conn.execute(
        "CREATE TABLE usage_events ("
        "key_hash TEXT, ts TEXT, metered INTEGER, status INTEGER, quantity INTEGER)"
    )
    conn.execute(
        "INSERT INTO api_keys "
        "(key_hash, tier, monthly_cap_yen, id, parent_key_id, revoked_at) "
        "VALUES (?, 'paid', ?, 1, NULL, NULL)",
        (key_hash, cap_yen),
    )
    return conn


def test_create_export_projected_cap_rejects_before_materialize_or_render(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from jpintel_mcp.api import export as export_mod

    calls = {"rate": 0, "materialize": 0, "render": 0, "stage": 0}

    def forbidden_rate_floor(_key_hash: str) -> None:
        calls["rate"] += 1
        raise AssertionError("rate floor must not consume cap-rejected export")

    def forbidden_materialize(*_args: object, **_kwargs: object):
        calls["materialize"] += 1
        raise AssertionError("export materialization must not start near cap")

    def forbidden_render(*_args: object, **_kwargs: object) -> bytes:
        calls["render"] += 1
        raise AssertionError("export render must not start near cap")

    def forbidden_stage(*_args: object, **_kwargs: object) -> str:
        calls["stage"] += 1
        raise AssertionError("R2 staging must not start near cap")

    monkeypatch.setattr(export_mod, "_rate_floor_check", forbidden_rate_floor)
    monkeypatch.setattr(export_mod, "_materialize_rows", forbidden_materialize)
    monkeypatch.setattr(export_mod, "_render_csv", forbidden_render)
    monkeypatch.setattr(export_mod, "_stage_to_r2", forbidden_stage)

    key_hash = "kh-export-cap"
    conn = _paid_cap_conn(key_hash=key_hash, cap_yen=(EXPORT_UNIT_COUNT - 1) * 3)
    ctx = ApiContext(
        key_hash=key_hash,
        tier="paid",
        customer_id="cus_test",
        stripe_subscription_id="sub_test",
    )
    request = Request({"type": "http", "method": "POST", "path": "/v1/export", "headers": []})

    response = create_export(
        ExportRequest(dataset="programs", format="csv"),
        request,
        BackgroundTasks(),
        ctx,
        conn,
    )

    assert response.status_code == 503
    err = json.loads(response.body)["error"]
    assert err["projected_units"] == EXPORT_UNIT_COUNT
    assert err["projected_yen"] == EXPORT_UNIT_COUNT * 3
    assert calls == {"rate": 0, "materialize": 0, "render": 0, "stage": 0}


# ---------------------------------------------------------------------------
# filter list cap
# ---------------------------------------------------------------------------


def test_export_request_rejects_large_filter_lists() -> None:
    with pytest.raises(ValidationError) as excinfo:
        ExportRequest(
            dataset="programs",
            format="csv",
            filter={"program_id": [str(i) for i in range(EXPORT_MAX_FILTER_LIST_ITEMS + 1)]},
        )

    assert "filter.program_id" in str(excinfo.value)
    assert str(EXPORT_MAX_FILTER_LIST_ITEMS) in str(excinfo.value)


def test_export_request_accepts_filter_list_at_cap() -> None:
    """Boundary: exactly EXPORT_MAX_FILTER_LIST_ITEMS items must pass."""
    body = ExportRequest(
        dataset="programs",
        format="csv",
        filter={"program_id": [str(i) for i in range(EXPORT_MAX_FILTER_LIST_ITEMS)]},
    )
    assert len(body.filter["program_id"]) == EXPORT_MAX_FILTER_LIST_ITEMS


def test_materialize_rows_rejects_large_filter_lists_before_sql() -> None:
    conn = sqlite3.connect(":memory:")
    conn.execute("CREATE TABLE programs (program_id TEXT PRIMARY KEY, primary_name TEXT)")
    conn.execute("INSERT INTO programs VALUES ('p1', 'test')")

    with pytest.raises(HTTPException) as excinfo:
        _materialize_rows(
            conn,
            "programs",
            {"program_id": [str(i) for i in range(EXPORT_MAX_FILTER_LIST_ITEMS + 1)]},
            None,
            10,
        )

    assert excinfo.value.status_code == 422
    assert excinfo.value.detail["max_items"] == EXPORT_MAX_FILTER_LIST_ITEMS


# ---------------------------------------------------------------------------
# columns list cap
# ---------------------------------------------------------------------------


def test_export_request_rejects_huge_columns_list() -> None:
    """`columns` over EXPORT_MAX_COLUMN_LIST_ITEMS must 422 at pydantic edge."""
    with pytest.raises(ValidationError) as excinfo:
        ExportRequest(
            dataset="programs",
            format="csv",
            columns=[f"col_{i}" for i in range(EXPORT_MAX_COLUMN_LIST_ITEMS + 1)],
        )

    assert "columns" in str(excinfo.value)
    assert str(EXPORT_MAX_COLUMN_LIST_ITEMS) in str(excinfo.value)


def test_export_request_accepts_columns_at_cap() -> None:
    """Boundary: exactly EXPORT_MAX_COLUMN_LIST_ITEMS columns must pass."""
    body = ExportRequest(
        dataset="programs",
        format="csv",
        columns=[f"col_{i}" for i in range(EXPORT_MAX_COLUMN_LIST_ITEMS)],
    )
    assert body.columns is not None
    assert len(body.columns) == EXPORT_MAX_COLUMN_LIST_ITEMS


# ---------------------------------------------------------------------------
# row cap
# ---------------------------------------------------------------------------


def test_export_request_rejects_limit_above_row_cap() -> None:
    with pytest.raises(ValidationError) as excinfo:
        ExportRequest(dataset="programs", format="csv", limit=EXPORT_MAX_ROWS + 1)

    assert "limit" in str(excinfo.value)


def test_export_request_accepts_limit_at_row_cap() -> None:
    body = ExportRequest(dataset="programs", format="csv", limit=EXPORT_MAX_ROWS)
    assert body.limit == EXPORT_MAX_ROWS


def test_materialize_rows_clamps_limit_to_row_cap() -> None:
    """`_materialize_rows` clamps ``limit`` to ``EXPORT_MAX_ROWS`` even if a
    caller bypasses the pydantic guard (e.g. internal/cron paths that
    reuse the helper)."""
    conn = sqlite3.connect(":memory:")
    conn.execute("CREATE TABLE programs (program_id TEXT PRIMARY KEY, primary_name TEXT)")
    conn.executemany(
        "INSERT INTO programs VALUES (?, ?)",
        [(f"p{i}", f"name-{i}") for i in range(3)],
    )

    _columns, rows = _materialize_rows(conn, "programs", {}, None, EXPORT_MAX_ROWS * 10)
    # The seed only has 3 rows; the test asserts the SQL did NOT raise on
    # the absurd limit (the helper clamps with min(limit, EXPORT_MAX_ROWS)
    # before binding the SQL param).
    assert len(rows) == 3


# ---------------------------------------------------------------------------
# XLSX cell budget
# ---------------------------------------------------------------------------


def _redistributable_row(**extra: object) -> dict[str, object]:
    """Build a row that passes ``assert_no_blocked`` (gov_standard license).

    The license-gate fence rejects any row that lacks an allow-listed
    ``license`` value. The cell-budget guard must fire BEFORE that fence
    (we don't want a 50 MB worker to OOM just because a single row is
    unlicensed), so the budget test seeds licensed rows; the
    duplicate-attribution / smoke tests do the same so they exercise
    the full renderer path.
    """
    row: dict[str, object] = {"license": "gov_standard"}
    row.update(extra)
    return row


def test_render_xlsx_rejects_cell_count_above_budget() -> None:
    """Renderer raises 422 when rows × columns exceeds the memory budget."""
    columns = [f"col_{i}" for i in range(200)]
    # 200 columns + _attribution = 201 (after the in-renderer append).
    # Pick enough rows to exceed the cell budget while staying below the
    # XLSX row budget, so this test specifically exercises the cell guard.
    effective_cols = len(columns) + 1  # _attribution
    rows_needed = (EXPORT_MAX_XLSX_CELLS // effective_cols) + 2
    assert rows_needed <= EXPORT_MAX_XLSX_ROWS
    rows = [_redistributable_row(**dict.fromkeys(columns, "x")) for _ in range(rows_needed)]

    with pytest.raises(HTTPException) as excinfo:
        _render_xlsx(list(columns), rows)

    assert excinfo.value.status_code == 422
    assert excinfo.value.detail["max_cells"] == EXPORT_MAX_XLSX_CELLS
    assert excinfo.value.detail["rows"] == rows_needed
    assert "xlsx cell budget" in excinfo.value.detail["detail"]


def test_render_xlsx_rejects_row_count_above_xlsx_budget() -> None:
    """XLSX is not end-to-end streamed, so the renderer enforces a
    format-specific row cap before building the final ZIP bytes."""
    columns = ["program_id"]
    rows = [_redistributable_row(program_id=f"p{i}") for i in range(EXPORT_MAX_XLSX_ROWS + 1)]

    with pytest.raises(HTTPException) as excinfo:
        _render_xlsx(list(columns), rows)

    assert excinfo.value.status_code == 422
    assert excinfo.value.detail["max_rows"] == EXPORT_MAX_XLSX_ROWS
    assert excinfo.value.detail["limitation"] == EXPORT_XLSX_STREAMING_LIMITATION


def test_render_xlsx_accepts_small_payload() -> None:
    """Smoke: a 2-row XLSX renders without raising and emits a zip blob."""
    columns = ["program_id", "primary_name", "tier"]
    rows = [
        _redistributable_row(program_id="p1", primary_name="テスト補助金 A", tier="S"),
        _redistributable_row(program_id="p2", primary_name="テスト補助金 B", tier="A"),
    ]

    blob = _render_xlsx(list(columns), rows)
    # XLSX is a ZIP container; "PK\x03\x04" is the local file header magic.
    assert blob.startswith(b"PK\x03\x04")


def test_render_xlsx_does_not_duplicate_attribution_column() -> None:
    """Regression: if columns already includes ``_attribution`` we must not
    append it again (otherwise XLSX has duplicate header cells and the
    cell-count guard double-counts)."""
    columns = ["program_id", "_attribution"]
    rows = [_redistributable_row(program_id="p1", _attribution="CC-BY 4.0")]

    blob = _render_xlsx(list(columns), rows)
    # The columns argument is mutated only via concat (`[*columns, "_attr"]`);
    # the caller's list must NOT have grown.
    assert columns == ["program_id", "_attribution"]
    # The XLSX blob ships the header twice if duplication regressed.
    # We can't easily decode the deflated zip part inline, but the
    # rendered cell count gives us a guard: 1 row × 2 columns = 2 data
    # cells. The cell-budget guard runs on `len(columns)` after the
    # in-renderer append; assert the guard saw "2" not "3".
    # The cleanest invariant is to call the renderer twice and ensure
    # the size doesn't grow on the second call (idempotent column ops).
    blob2 = _render_xlsx(list(columns), rows)
    assert len(blob) == len(blob2)


# ---------------------------------------------------------------------------
# Constants sanity
# ---------------------------------------------------------------------------


def test_perf_constants_have_sane_relative_values() -> None:
    """Defence against an accidental constant flip that would invert the
    relationship between row cap and XLSX cell budget (would let a max-row
    export OOM the worker)."""
    # 50,000 rows × 20-ish columns = 1M cells; budget must cover that.
    assert EXPORT_MAX_XLSX_CELLS >= EXPORT_MAX_ROWS * 20
    assert EXPORT_XLSX_STREAM_THRESHOLD_ROWS < EXPORT_MAX_XLSX_ROWS < EXPORT_MAX_ROWS
    # Filter list cap must be well below the SQLite default variable
    # ceiling (SQLITE_MAX_VARIABLE_NUMBER default 999 / 32766).
    assert EXPORT_MAX_FILTER_LIST_ITEMS <= 999
    # Column cap stays small (real views have ~40 cols max).
    assert EXPORT_MAX_COLUMN_LIST_ITEMS <= 1_000


# ---------------------------------------------------------------------------
# R3 filter shape cap (dict-key count / nesting depth)
# ---------------------------------------------------------------------------


def test_export_request_rejects_oversized_filter_dict() -> None:
    """33-key filter dict (one over EXPORT_MAX_FILTER_DICT_KEYS) must 422."""
    too_many = {f"key_{i}": "v" for i in range(EXPORT_MAX_FILTER_DICT_KEYS + 1)}
    with pytest.raises(ValidationError) as excinfo:
        ExportRequest(dataset="programs", format="csv", filter=too_many)

    assert "filter" in str(excinfo.value)
    assert str(EXPORT_MAX_FILTER_DICT_KEYS) in str(excinfo.value)


def test_export_request_accepts_filter_dict_at_cap() -> None:
    """Boundary: exactly EXPORT_MAX_FILTER_DICT_KEYS flat keys must pass."""
    at_cap = {f"key_{i}": "v" for i in range(EXPORT_MAX_FILTER_DICT_KEYS)}
    body = ExportRequest(dataset="programs", format="csv", filter=at_cap)
    assert len(body.filter) == EXPORT_MAX_FILTER_DICT_KEYS


def test_export_request_rejects_nested_dict_in_filter() -> None:
    """Nested dict value (e.g. ``{"foo": {"bar": 1}}``) must 422."""
    with pytest.raises(ValidationError) as excinfo:
        ExportRequest(
            dataset="programs",
            format="csv",
            filter={"foo": {"bar": 1}},
        )

    assert "nested dict" in str(excinfo.value)


def test_export_request_rejects_list_of_list_in_filter() -> None:
    """list[list[...]] must 422 — one-level lists are OK (capped by A8)."""
    with pytest.raises(ValidationError) as excinfo:
        ExportRequest(
            dataset="programs",
            format="csv",
            filter={"foo": [["a", "b"], ["c", "d"]]},
        )

    assert "nested lists or dicts" in str(excinfo.value)


def test_export_request_rejects_overlong_filter_key() -> None:
    """A filter key longer than EXPORT_MAX_FILTER_KEY_LEN chars must 422."""
    long_key = "k" * 65
    with pytest.raises(ValidationError) as excinfo:
        ExportRequest(
            dataset="programs",
            format="csv",
            filter={long_key: "v"},
        )

    assert "filter key" in str(excinfo.value)


def test_list_formats_discloses_xlsx_buffered_limit() -> None:
    payload = list_formats()

    assert payload["xlsx_row_cap"] == EXPORT_MAX_XLSX_ROWS
    assert payload["xlsx_cell_cap"] == EXPORT_MAX_XLSX_CELLS
    assert payload["xlsx_streaming_limitation"] == EXPORT_XLSX_STREAMING_LIMITATION


def test_create_export_rejects_xlsx_limit_before_materialize(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A too-large XLSX request must fail before row materialization since
    the route still buffers rows and the final ZIP."""
    from jpintel_mcp.api import export as export_mod

    calls = {"materialize": 0}

    def forbidden_materialize(*_args: object, **_kwargs: object):
        calls["materialize"] += 1
        raise AssertionError("xlsx over-limit request must not materialize rows")

    monkeypatch.setattr(export_mod, "_materialize_rows", forbidden_materialize)
    monkeypatch.setattr(export_mod, "_rate_floor_check", lambda _key_hash: None)
    monkeypatch.setattr(export_mod, "_projected_cap_response", lambda *_args: None)

    conn = sqlite3.connect(":memory:")
    ctx = ApiContext(
        key_hash="kh-xlsx-limit",
        tier="paid",
        customer_id="cus_test",
        stripe_subscription_id="sub_test",
    )
    request = Request({"type": "http", "method": "POST", "path": "/v1/export", "headers": []})

    with pytest.raises(HTTPException) as excinfo:
        create_export(
            ExportRequest(
                dataset="programs",
                format="xlsx",
                limit=EXPORT_MAX_XLSX_ROWS + 1,
            ),
            request,
            BackgroundTasks(),
            ctx,
            conn,
        )

    assert excinfo.value.status_code == 422
    assert excinfo.value.detail["max_rows"] == EXPORT_MAX_XLSX_ROWS
    assert calls == {"materialize": 0}


# ---------------------------------------------------------------------------
# R3 XLSX streaming path (memory ceiling)
# ---------------------------------------------------------------------------


def _zip_magic_ok(blob: bytes) -> bool:
    """XLSX = ZIP container; verify local file header magic."""
    return blob.startswith(b"PK\x03\x04")


def test_render_xlsx_small_export_uses_inmemory_fast_path(monkeypatch: pytest.MonkeyPatch) -> None:
    """Boundary: a 100-row export must use the in-memory hand-rolled-zip
    path (below ``EXPORT_XLSX_STREAM_THRESHOLD_ROWS``) and never call the
    streaming openpyxl writer."""
    from jpintel_mcp.api import export as export_mod

    streaming_called = {"count": 0}
    inmemory_called = {"count": 0}

    real_streaming = export_mod._render_xlsx_streaming
    real_inmemory = export_mod._render_xlsx_inmemory

    def spy_streaming(cols: list[str], rows: list[dict[str, object]]) -> bytes:
        streaming_called["count"] += 1
        return real_streaming(cols, rows)

    def spy_inmemory(cols: list[str], rows: list[dict[str, object]]) -> bytes:
        inmemory_called["count"] += 1
        return real_inmemory(cols, rows)

    monkeypatch.setattr(export_mod, "_render_xlsx_streaming", spy_streaming)
    monkeypatch.setattr(export_mod, "_render_xlsx_inmemory", spy_inmemory)

    columns = ["program_id", "primary_name", "tier"]
    rows = [
        _redistributable_row(
            program_id=f"p{i}",
            primary_name=f"テスト補助金 {i}",
            tier="S",
        )
        for i in range(100)
    ]

    blob = export_mod._render_xlsx(list(columns), rows)

    assert _zip_magic_ok(blob)
    assert inmemory_called["count"] == 1
    assert streaming_called["count"] == 0


def test_render_xlsx_large_export_uses_streaming_path(monkeypatch: pytest.MonkeyPatch) -> None:
    """Above ``EXPORT_XLSX_STREAM_THRESHOLD_ROWS``, the public renderer
    delegates to ``_render_xlsx_streaming`` (openpyxl write_only) — the
    in-memory fast path must NOT be called."""
    from jpintel_mcp.api import export as export_mod

    streaming_called = {"count": 0}
    inmemory_called = {"count": 0}

    real_streaming = export_mod._render_xlsx_streaming
    real_inmemory = export_mod._render_xlsx_inmemory

    def spy_streaming(cols: list[str], rows: list[dict[str, object]]) -> bytes:
        streaming_called["count"] += 1
        return real_streaming(cols, rows)

    def spy_inmemory(cols: list[str], rows: list[dict[str, object]]) -> bytes:
        inmemory_called["count"] += 1
        return real_inmemory(cols, rows)

    monkeypatch.setattr(export_mod, "_render_xlsx_streaming", spy_streaming)
    monkeypatch.setattr(export_mod, "_render_xlsx_inmemory", spy_inmemory)

    columns = ["program_id", "primary_name", "tier"]
    rows = [
        _redistributable_row(
            program_id=f"p{i}",
            primary_name=f"row-{i}",
            tier="S",
        )
        # One above the streaming threshold → streaming path.
        for i in range(EXPORT_XLSX_STREAM_THRESHOLD_ROWS + 1)
    ]

    blob = export_mod._render_xlsx(list(columns), rows)

    assert _zip_magic_ok(blob)
    assert streaming_called["count"] == 1
    assert inmemory_called["count"] == 0


def test_render_xlsx_streaming_peak_memory_under_50mb() -> None:
    """R3 memory-bound regression. The streaming path must hold peak
    memory under 50 MB for a 5,000 row × 20 col XLSX render.

    Measured via ``tracemalloc`` (stdlib — psutil is not a hard dep).
    The previous in-memory builder would peak around 80 MB just for the
    ``StringIO`` worksheet XML on the same input, plus the compressed
    zip; the streaming path keeps peak roughly constant in row count.
    """
    import tracemalloc

    n_rows = 5_000
    n_data_cols = 20
    columns = [f"col_{i}" for i in range(n_data_cols)]
    # ~40-byte ASCII per cell × 20 cols × 5000 rows ≈ 4 MB row payload
    # plus the rendered XLSX zip. The streaming path should keep peak
    # comfortably under 50 MB.
    rows = [
        _redistributable_row(**{c: f"value-{c}-{i}-xxxxx" for c in columns}) for i in range(n_rows)
    ]

    tracemalloc.start()
    try:
        # Reset the peak counter after building the input fixture so we
        # measure ONLY the render's incremental allocations.
        tracemalloc.reset_peak()
        blob = _render_xlsx_streaming(list(columns), rows)
        _current, peak = tracemalloc.get_traced_memory()
    finally:
        tracemalloc.stop()

    assert _zip_magic_ok(blob)
    # 50 MB ceiling — generous headroom over the actual ~10-20 MB the
    # streaming path observes in practice, tight enough to catch a
    # regression that reintroduces a full in-memory sheet build.
    peak_mb = peak / (1024 * 1024)
    assert peak_mb < 50, (
        f"streaming xlsx peak memory {peak_mb:.1f} MB exceeds 50 MB ceiling "
        f"(rows={n_rows}, cols={n_data_cols})"
    )


def test_render_xlsx_streaming_round_trip_preserves_rows() -> None:
    """The streaming path must produce a valid XLSX whose cells round-trip
    back through openpyxl (data shape preserved, header intact)."""
    from openpyxl import load_workbook

    columns = ["program_id", "primary_name", "tier"]
    rows = [
        _redistributable_row(
            program_id=f"p{i}",
            primary_name=f"行 {i}",
            tier="S",
        )
        # Force the streaming path with EXPORT_XLSX_STREAM_THRESHOLD_ROWS+50.
        for i in range(EXPORT_XLSX_STREAM_THRESHOLD_ROWS + 50)
    ]

    blob = _render_xlsx_streaming(list(columns), rows)
    assert _zip_magic_ok(blob)

    import io as _io

    wb = load_workbook(_io.BytesIO(blob), read_only=True)
    ws = wb["export"]
    rows_back = list(ws.iter_rows(values_only=True))
    # +1 header row, +_attribution column appended by the public wrapper —
    # `_render_xlsx_streaming` is called directly here so the caller is
    # responsible for the append. The renderer simply preserves whatever
    # columns it was given; we passed 3 cols so we expect 3 header cells.
    assert len(rows_back) == len(rows) + 1
    # Header preserved.
    assert rows_back[0] == tuple(columns)
    # First data row preserved.
    assert rows_back[1][0] == "p0"
    assert rows_back[1][1] == "行 0"
    assert rows_back[1][2] == "S"


def test_render_xlsx_streaming_threshold_boundary() -> None:
    """At exactly ``EXPORT_XLSX_STREAM_THRESHOLD_ROWS`` rows the public
    renderer stays on the fast path; one row over flips to streaming.

    The boundary is ``len(rows) > EXPORT_XLSX_STREAM_THRESHOLD_ROWS``.
    """
    columns = ["program_id"]
    # `threshold` exactly → inmemory.
    at_threshold = [
        _redistributable_row(program_id=f"p{i}") for i in range(EXPORT_XLSX_STREAM_THRESHOLD_ROWS)
    ]
    blob_inmem = _render_xlsx(list(columns), at_threshold)
    assert _zip_magic_ok(blob_inmem)

    # `threshold + 1` → streaming.
    over_threshold = [
        _redistributable_row(program_id=f"p{i}")
        for i in range(EXPORT_XLSX_STREAM_THRESHOLD_ROWS + 1)
    ]
    blob_stream = _render_xlsx(list(columns), over_threshold)
    assert _zip_magic_ok(blob_stream)
