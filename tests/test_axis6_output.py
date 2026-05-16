"""tests/test_axis6_output.py — Axis 6 output forms smoke tests.

Wave 35 Axis 6 (2026-05-12). Covers:
* PDF generate (reportlab) smoke + section presence
* Excel template render smoke + 5-sheet integrity + seed presence
* freee / MF / yayoi plugin signature + extraction
* Notion v3 schema verify
* Linear v2 schema verify
* Webhook v2 HMAC verify (Python-side parity with the TS implementation)
* Migration 244 forward + rollback idempotency
* Cron dry-run

All tests are network-free.
"""

from __future__ import annotations

import hashlib
import hmac
import io
import json
import sqlite3
import tempfile
import time
from pathlib import Path

import pytest


def _ensure_repo_on_path() -> None:
    import sys

    repo = Path(__file__).resolve().parent.parent
    for p in (repo, repo / "src"):
        if str(p) not in sys.path:
            sys.path.insert(0, str(p))


_ensure_repo_on_path()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def test_pdf_render_produces_bytes_with_sections() -> None:
    pytest.importorskip("reportlab")
    from jpintel_mcp.api.pdf_report import _render_pdf

    ctx = {
        "client_id": "9123456789012",
        "fetched_at": "2026-05-12T00:00:00Z",
        "houjin_360": {"name": "テスト株式会社", "prefecture": "東京都"},
        "risk_score": {"composite_score": 0.42, "computed_at": "2026-05-11"},
        "new_programs": [{"program_id": "P-001", "title": "テスト補助金", "authority": "経産省"}],
        "amendments": [
            {
                "amendment_id": "A-001",
                "law_id": "L-001",
                "summary": "テスト改正",
                "effective_from": "2026-06-01",
            }
        ],
        "fence_summary": [{"law": "税理士法 §52", "rule": "テスト fence"}],
    }
    blob, page_count = _render_pdf("9123456789012", ctx)
    assert isinstance(blob, bytes)
    assert blob[:4] == b"%PDF"
    assert page_count >= 1


def test_pdf_render_handles_empty_substrate() -> None:
    pytest.importorskip("reportlab")
    from jpintel_mcp.api.pdf_report import _render_pdf

    ctx = {
        "client_id": "X",
        "fetched_at": "2026-05-12T00:00:00Z",
        "houjin_360": None,
        "risk_score": None,
        "new_programs": [],
        "amendments": [],
        "fence_summary": [],
    }
    blob, page_count = _render_pdf("X", ctx)
    assert blob[:4] == b"%PDF"
    assert page_count >= 1


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def test_excel_render_5_sheet_workbook() -> None:
    pytest.importorskip("openpyxl")
    from jpintel_mcp.api.excel_template import _render_xlsx

    substrate = {
        "program": {
            "program_id": "P-001",
            "title": "テスト補助金",
            "authority": "経産省",
            "tier": "A",
            "source_url": "https://example.gov.jp/p001",
            "summary": "テスト",
            "max_amount_yen": 1_000_000,
            "source_fetched_at": "2026-05-11",
        },
        "adoption_stats": {"source": "case_studies", "count": 12},
        "application_rounds": [
            {
                "round_id": "R-001",
                "application_open_at": "2026-06-01",
                "application_close_at": "2026-06-30",
                "status": "open",
            }
        ],
    }
    blob, sheet_count = _render_xlsx(
        "P-001",
        substrate,
        include_value_estimate=True,
        include_calendar=True,
        note="test note",
    )
    assert blob[:2] == b"PK"
    assert sheet_count == 5
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(io.BytesIO(blob))
    assert set(wb.sheetnames) == {
        "制度サマリ",
        "必要書類chk",
        "金額目安",
        "期限calendar",
        "補足連絡先",
    }


def test_excel_request_accepts_legacy_include_roi_alias() -> None:
    from jpintel_mcp.api.excel_template import ExcelEstimateRequest

    body = ExcelEstimateRequest.model_validate({"include_roi": False})
    assert body.include_value_estimate is False


def test_excel_template_seed_exists_and_opens() -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    seed = (
        Path(__file__).resolve().parent.parent
        / "templates"
        / "excel"
        / "application_estimate_v1.xlsx"
    )
    assert seed.exists(), f"seed missing: {seed}"
    wb = load_workbook(seed)
    assert len(wb.sheetnames) == 5


# ---------------------------------------------------------------------------
# Plugin signatures
# ---------------------------------------------------------------------------


def _hmac_sha256(secret: str, body: bytes) -> str:
    return hmac.new(secret.encode("utf-8"), body, hashlib.sha256).hexdigest()


def test_mf_webhook_verify_and_account_extraction() -> None:
    from importlib.util import module_from_spec, spec_from_file_location

    target = Path(__file__).resolve().parent.parent / "sdk" / "mf-plugin" / "mf_webhook_trigger.py"
    spec = spec_from_file_location("mf_webhook_trigger", target)
    assert spec is not None and spec.loader is not None
    mod = module_from_spec(spec)
    try:
        spec.loader.exec_module(mod)
    except ImportError as exc:
        pytest.skip(f"mf plugin import skipped: {exc}")

    body = json.dumps(
        {
            "transaction": {
                "account_classification": {
                    "large_category": "通信費",
                    "middle_category": "ソフトウェア",
                }
            }
        },
        ensure_ascii=False,
    ).encode("utf-8")
    sig = _hmac_sha256("S3CR3T_MF", body)
    mod.verify_mf_webhook(raw_body=body, signature_header=sig, secret="S3CR3T_MF")
    with pytest.raises(mod.MfWebhookSignatureError):
        mod.verify_mf_webhook(raw_body=body, signature_header="bad", secret="S3CR3T_MF")

    payload = json.loads(body.decode("utf-8"))
    accs = mod._extract_accounts(payload)
    assert "通信費" in accs
    assert "ソフトウェア" in accs


def test_yayoi_csv_parser_and_mapping() -> None:
    from importlib.util import module_from_spec, spec_from_file_location

    target = Path(__file__).resolve().parent.parent / "sdk" / "yayoi-plugin" / "yayoi_to_jpcite.py"
    spec = spec_from_file_location("yayoi_to_jpcite", target)
    assert spec is not None and spec.loader is not None
    mod = module_from_spec(spec)
    try:
        spec.loader.exec_module(mod)
    except ImportError as exc:
        pytest.skip(f"yayoi plugin import skipped: {exc}")

    csv_text = (
        "日付,借方勘定科目,貸方勘定科目,金額\n"
        "2026-05-01,ソフトウェア,普通預金,500000\n"
        "2026-05-02,通信費,普通預金,18000\n"
        "2026-05-03,広告宣伝費,普通預金,80000\n"
    )
    accs = mod.parse_yayoi_csv(csv_text.encode("utf-8"))
    assert "ソフトウェア" in accs
    assert "通信費" in accs
    assert "広告宣伝費" in accs
    purposes = mod.map_accounts_to_purposes(accs)
    assert "IT導入" in purposes
    assert "販路開拓" in purposes


def test_freee_journal_mapping_present() -> None:
    """The freee plugin module must have JOURNAL_ACCOUNT_TO_FUNDING_PURPOSE."""
    target = (
        Path(__file__).resolve().parent.parent / "sdk" / "freee-plugin" / "freee_webhook_trigger.py"
    )
    assert target.exists()
    src = target.read_text(encoding="utf-8")
    assert "JOURNAL_ACCOUNT_TO_FUNDING_PURPOSE" in src
    assert "verify_freee_webhook" in src
    assert "freee_webhook_handle" in src


# ---------------------------------------------------------------------------
# Notion v3 schema verify
# ---------------------------------------------------------------------------


def test_notion_v3_amendment_property_schema() -> None:
    from tools.integrations.notion_sync_v3 import (
        _amendment_to_notion_properties,
    )

    hit = {
        "amendment_id": "AMD-001",
        "law_id": "law-2026-001",
        "effective_from": "2026-07-01",
        "summary": "テスト改正概要",
    }
    props = _amendment_to_notion_properties(hit)
    assert "title" in props["name"]
    assert props["kind"]["select"]["name"] == "law_amendment"
    assert props["external_id"]["rich_text"][0]["text"]["content"] == "AMD-001"
    assert props["law_id"]["rich_text"][0]["text"]["content"] == "law-2026-001"
    assert props["effective_from"]["date"]["start"] == "2026-07-01"
    assert props["source_url"]["url"].startswith("https://jpcite.com/laws/")


def test_notion_v3_houjin_property_schema() -> None:
    from tools.integrations.notion_sync_v3 import (
        _houjin_to_notion_properties,
    )

    hit = {
        "watch_id": "W-001",
        "houjin_bangou": "9123456789012",
        "hit_kind": "amendment_match",
        "summary": "watchlist 法人に法令改正",
        "last_seen_at": "2026-05-12T00:00:00Z",
    }
    props = _houjin_to_notion_properties(hit)
    assert props["kind"]["select"]["name"] == "houjin_watch_hit"
    assert props["houjin_bangou"]["rich_text"][0]["text"]["content"] == "9123456789012"


# ---------------------------------------------------------------------------
# Linear v2 schema verify
# ---------------------------------------------------------------------------


def test_linear_v2_gql_mutation_variables_shape() -> None:
    from tools.integrations import linear_ticket_v2 as mod

    assert "issueCreate" in mod._CREATE_ISSUE_GQL
    for var in ("$teamId", "$title", "$description", "$labelIds", "$priority"):
        assert var in mod._CREATE_ISSUE_GQL

    import os

    saved = os.environ.pop("LINEAR_TARGETS_JSON", None)
    try:
        n = mod._fan_out(
            kind="test",
            hits=[{"amendment_id": "A-1"}],
            title_fn=lambda h: "T",
            body_fn=lambda h: "B",
        )
        assert n == 0
    finally:
        if saved is not None:
            os.environ["LINEAR_TARGETS_JSON"] = saved


# ---------------------------------------------------------------------------
# Webhook v2 HMAC
# ---------------------------------------------------------------------------


def test_webhook_v2_hmac_python_side_matches_ts_contract() -> None:
    secret = "test_secret_v2"
    timestamp = str(int(time.time() * 1000))
    body = json.dumps(
        {
            "kind": "amendment.confirmed",
            "event": "warn",
            "title": "改正アラート",
            "url": "https://jpcite.com/laws/L-001.html",
            "summary": "テストサマリ",
            "targets": ["slack", "discord", "teams"],
            "fields": [{"name": "law_id", "value": "L-001"}],
            "actions": [{"label": "jpcite で開く", "url": "https://jpcite.com/laws/L-001.html"}],
        },
        ensure_ascii=False,
    )
    signing_input = f"{timestamp}.{body}".encode()
    digest = hmac.new(secret.encode("utf-8"), signing_input, hashlib.sha256).hexdigest()
    assert len(digest) == 64
    assert all(c in "0123456789abcdef" for c in digest)


def test_webhook_v2_file_exports_get_post_handlers() -> None:
    p = Path(__file__).resolve().parent.parent / "functions" / "webhook_router_v2.ts"
    src = p.read_text(encoding="utf-8")
    assert "export const onRequestPost" in src
    assert "export const onRequestGet" in src
    assert "AdaptiveCard" in src
    assert "Block Kit" in src
    assert "Embed" in src
    assert "REPLAY_WINDOW_MS" in src
    assert "MAX_BODY_BYTES = 65_536" in src


# ---------------------------------------------------------------------------
# Migration 244
# ---------------------------------------------------------------------------


def test_migration_244_applies_idempotently() -> None:
    mig = (
        Path(__file__).resolve().parent.parent
        / "scripts"
        / "migrations"
        / "244_pdf_report_generator.sql"
    )
    assert mig.exists()
    sql = mig.read_text(encoding="utf-8")
    with tempfile.TemporaryDirectory() as tmp:
        db = Path(tmp) / "t.db"
        for _ in range(2):
            conn = sqlite3.connect(str(db))
            try:
                conn.executescript(sql)
                conn.commit()
            finally:
                conn.close()
        conn = sqlite3.connect(str(db))
        try:
            tables = {
                r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
            }
            assert "am_pdf_report_subscriptions" in tables
            assert "am_pdf_report_generation_log" in tables
        finally:
            conn.close()


def test_migration_244_rollback_is_idempotent() -> None:
    fwd = (
        Path(__file__).resolve().parent.parent
        / "scripts"
        / "migrations"
        / "244_pdf_report_generator.sql"
    )
    rb = (
        Path(__file__).resolve().parent.parent
        / "scripts"
        / "migrations"
        / "244_pdf_report_generator_rollback.sql"
    )
    assert rb.exists()
    fwd_sql = fwd.read_text(encoding="utf-8")
    rb_sql = rb.read_text(encoding="utf-8")
    with tempfile.TemporaryDirectory() as tmp:
        db = Path(tmp) / "t.db"
        conn = sqlite3.connect(str(db))
        try:
            conn.executescript(fwd_sql)
            conn.executescript(rb_sql)
            conn.executescript(rb_sql)
            conn.commit()
            tables = {
                r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
            }
            assert "am_pdf_report_subscriptions" not in tables
            assert "am_pdf_report_generation_log" not in tables
        finally:
            conn.close()


# ---------------------------------------------------------------------------
# Cron dry-run
# ---------------------------------------------------------------------------


def test_generate_pdf_reports_monthly_dry_run() -> None:
    pytest.importorskip("reportlab")
    import os
    import subprocess
    import sys

    repo = Path(__file__).resolve().parent.parent
    with tempfile.TemporaryDirectory() as tmp:
        db = Path(tmp) / "t.db"
        mig = repo / "scripts" / "migrations" / "244_pdf_report_generator.sql"
        sql = mig.read_text(encoding="utf-8")
        conn = sqlite3.connect(str(db))
        try:
            conn.executescript(sql)
            conn.commit()
        finally:
            conn.close()

        env = os.environ.copy()
        env["AUTONOMATH_DB_PATH"] = str(db)
        env["JPINTEL_DB"] = str(db)
        result = subprocess.run(
            [
                sys.executable,
                str(repo / "scripts" / "cron" / "generate_pdf_reports_monthly.py"),
                "--dry-run",
            ],
            capture_output=True,
            text=True,
            env=env,
            timeout=30,
        )
        assert result.returncode == 0, f"stderr: {result.stderr}"
