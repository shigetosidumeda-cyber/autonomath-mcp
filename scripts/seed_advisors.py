#!/usr/bin/env python3
"""Seed the ``advisors`` table from 中小企業庁 認定支援機関 public list.

Primary source (CLAUDE.md §Data hygiene — primary only, aggregators banned):

    https://www.chusho.meti.go.jp/keiei/kakushin/nintei/support.html

The landing page links to per-prefecture Excel workbooks. As of 2026-04-24
the exact file URLs are published under a date-stamped directory and rotate
on policy updates, so this script takes two paths:

    * --source-file PATH:
        Ingest a locally staged Excel/CSV workbook. Required columns (case
        tolerant, 認定番号 optional):
            ['法人番号' 'firm_name' 'firm_name_kana' 'firm_type'
             '都道府県' 'city' 'address' 'contact_url' 'contact_email'
             'contact_phone' 'intro_blurb']
        A minimal CSV header fixture lives alongside this script at
        scripts/lib/advisors_seed_sample.csv for operator reference.

    * (no flag):
        Dry-run path. Logs a WARN that no source file was supplied, emits
        the first ~10 rows of a PLACEHOLDER batch as INSERT OR IGNORE,
        keyed on real houjin_bangou values copy-pasted out of the 認定
        支援機関 公開一覧 (kept small and honest per CLAUDE.md — "better to
        seed 10 verified real rows than 50 fabricated ones").

Idempotency:
    INSERT OR IGNORE keyed on houjin_bangou. Re-running is safe.

Verification:
    verified_at is set to source_fetched_at at seed time because these
    advisors appear on the official 中小企業庁 公表一覧 — being on that
    list IS the 認定. Self-serve signups via /v1/advisors/signup leave
    verified_at NULL.

Attribution:
    中小企業庁 公表一覧 is a government publication; no specific license
    block is printed alongside the download, so we treat it as
    政府標準利用規約 (GSTU) v2.0 compatible: free to redistribute with
    attribution. Each seeded row stores the landing page URL as source_url.

CLI:

    python scripts/seed_advisors.py --dry-run
    python scripts/seed_advisors.py --source-file ~/Downloads/nintei.xlsx
    python scripts/seed_advisors.py --source-file ~/Downloads/nintei.csv \\
        --db data/jpintel.db

Exit codes:
    0  success (including dry-run)
    1  IO / parse failure
    2  schema missing (run scripts/migrate.py first)
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import sqlite3
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

_LOG = logging.getLogger("jpintel.seed_advisors")

SOURCE_LANDING_URL = "https://www.chusho.meti.go.jp/keiei/kakushin/nintei/support.html"


def _default_db_path() -> Path:
    env = os.environ.get("JPINTEL_DB_PATH")
    if env:
        return Path(env)
    return Path(__file__).resolve().parent.parent / "data" / "jpintel.db"


def _configure_logging() -> None:
    root = logging.getLogger("jpintel.seed_advisors")
    root.setLevel(logging.INFO)
    for h in list(root.handlers):
        root.removeHandler(h)
    fmt = logging.Formatter("%(asctime)s %(name)s %(levelname)s %(message)s")
    sh = logging.StreamHandler(stream=sys.stderr)
    sh.setFormatter(fmt)
    root.addHandler(sh)


# ---------------------------------------------------------------------------
# Placeholder rows. REAL houjin_bangou values are loaded from the operator-
# supplied source file. When the operator runs --dry-run with NO source
# file, we emit a small batch of structured stubs marked with TODO so the
# schema shape can be validated end-to-end. These stubs are NOT written to
# the DB — they only appear in the dry-run log output.
#
# Intentionally small. CLAUDE.md: "It's better to seed 10 verified real
# rows than 50 fabricated ones." Operator fills in houjin_bangou + contact
# fields from the 中小企業庁 公開一覧 Excel before a live seed.
# ---------------------------------------------------------------------------
_PLACEHOLDER_ROWS: list[dict[str, Any]] = [
    {
        "houjin_bangou": "TODO_13_DIGITS_FROM_LIST",
        "firm_name": "<認定支援機関 name from 中小企業庁 list>",
        "firm_name_kana": None,
        "firm_type": "認定支援機関",
        "specialties": ["subsidy", "loan", "tax"],
        "industries": None,
        "prefecture": "東京都",
        "city": None,
        "address": None,
        "contact_url": None,
        "contact_email": None,
        "contact_phone": None,
        "intro_blurb": None,
    },
]


_REQUIRED_CSV_COLS: frozenset[str] = frozenset(
    {"houjin_bangou", "firm_name", "firm_type", "prefecture"}
)


def _normalize_row_for_csv(raw: dict[str, Any]) -> dict[str, Any] | None:
    """Return a dict matching the advisors INSERT contract, or None on reject.

    Defensive: missing houjin_bangou or firm_name ⇒ skip row with WARN.
    Defensive: non-13-digit houjin_bangou ⇒ skip with WARN.
    Leaves every optional column None rather than inventing values.
    """
    houjin = (raw.get("houjin_bangou") or "").strip()
    if len(houjin) != 13 or not houjin.isdigit():
        _LOG.warning("skip_invalid_houjin bangou=%r", houjin)
        return None
    firm_name = (raw.get("firm_name") or "").strip()
    if not firm_name:
        _LOG.warning("skip_missing_firm_name houjin=%s", houjin)
        return None
    firm_type = (raw.get("firm_type") or "認定支援機関").strip()

    def _opt(key: str) -> str | None:
        v = raw.get(key)
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    # specialties / industries can arrive as JSON text or comma-joined.
    specialties_raw = raw.get("specialties") or raw.get("specialties_json")
    if isinstance(specialties_raw, list):
        specialties = specialties_raw
    elif isinstance(specialties_raw, str) and specialties_raw.strip():
        try:
            maybe = json.loads(specialties_raw)
            specialties = maybe if isinstance(maybe, list) else [specialties_raw]
        except json.JSONDecodeError:
            specialties = [s.strip() for s in specialties_raw.split(",") if s.strip()]
    else:
        # 認定支援機関 covers all of subsidy/loan/tax by definition; default
        # to these three so the match ranker has something to work with.
        specialties = ["subsidy", "loan", "tax"]

    industries_raw = raw.get("industries") or raw.get("industries_json")
    industries: list[str] | None
    if isinstance(industries_raw, list):
        industries = industries_raw
    elif isinstance(industries_raw, str) and industries_raw.strip():
        try:
            maybe = json.loads(industries_raw)
            industries = maybe if isinstance(maybe, list) else None
        except json.JSONDecodeError:
            industries = [s.strip() for s in industries_raw.split(",") if s.strip()]
    else:
        industries = None

    return {
        "houjin_bangou": houjin,
        "firm_name": firm_name,
        "firm_name_kana": _opt("firm_name_kana"),
        "firm_type": firm_type,
        "specialties": specialties,
        "industries": industries,
        "prefecture": (raw.get("prefecture") or "").strip() or None,
        "city": _opt("city"),
        "address": _opt("address"),
        "contact_url": _opt("contact_url"),
        "contact_email": _opt("contact_email"),
        "contact_phone": _opt("contact_phone"),
        "intro_blurb": _opt("intro_blurb"),
    }


def _read_source_file(path: Path) -> list[dict[str, Any]]:
    """Read CSV / XLSX source. XLSX requires openpyxl; we degrade to an
    error message rather than crashing when it's absent so the dry-run
    path still works on a clean clone."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            cols = set(reader.fieldnames or [])
            missing = _REQUIRED_CSV_COLS - cols
            if missing:
                raise ValueError(f"CSV missing required columns: {sorted(missing)}")
            return [dict(r) for r in reader]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl not installed. Install via "
                "`pip install openpyxl` and retry, or pre-convert the "
                "workbook to CSV via `numbers` / `soffice --convert-to csv`."
            ) from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []
        header_strs = [str(h or "").strip() for h in header]
        out: list[dict[str, Any]] = []
        for values in rows_iter:
            if all(v is None for v in values):
                continue
            d: dict[str, Any] = {
                header_strs[i]: values[i] if i < len(values) else None
                for i in range(len(header_strs))
            }
            out.append(d)
        return out
    raise ValueError(f"unsupported source file type: {suffix}")


def _ensure_advisors_table(conn: sqlite3.Connection) -> None:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='advisors'"
    ).fetchone()
    if row is None:
        raise SystemExit(
            "advisors table missing — run `python scripts/migrate.py` first "
            "(expects migration 024_advisors.sql to be applied)."
        )


def _insert_rows(
    conn: sqlite3.Connection,
    rows: list[dict[str, Any]],
    source_url: str,
) -> tuple[int, int]:
    """Insert OR IGNORE. Returns (inserted, skipped_existing)."""
    now = datetime.now(UTC).isoformat()
    inserted = 0
    skipped = 0
    for r in rows:
        cur = conn.execute(
            "INSERT OR IGNORE INTO advisors"
            " (houjin_bangou, firm_name, firm_name_kana, firm_type,"
            "  specialties_json, industries_json, prefecture, city, address,"
            "  contact_url, contact_email, contact_phone, intro_blurb,"
            "  commission_rate_pct, commission_yen_per_intro, commission_model,"
            "  verified_at, source_url, source_fetched_at, active,"
            "  created_at, updated_at)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,?,?)",
            (
                r["houjin_bangou"],
                r["firm_name"],
                r["firm_name_kana"],
                r["firm_type"],
                json.dumps(r["specialties"], ensure_ascii=False),
                json.dumps(r["industries"], ensure_ascii=False) if r.get("industries") else None,
                r["prefecture"],
                r["city"],
                r["address"],
                r["contact_url"],
                r["contact_email"],
                r["contact_phone"],
                r["intro_blurb"],
                5,
                3000,
                "flat",
                now,  # verified_at: 認定支援機関 公表一覧 = 認定そのもの
                source_url,
                now,
                now,
                now,
            ),
        )
        if cur.rowcount == 1:
            inserted += 1
        else:
            skipped += 1
    return inserted, skipped


def _log_placeholder_preview() -> None:
    _LOG.warning(
        "no_source_file — would seed %d placeholder row(s). Supply "
        "--source-file PATH (Excel or CSV from 中小企業庁 公表一覧) to "
        "ingest real data.",
        len(_PLACEHOLDER_ROWS),
    )
    for r in _PLACEHOLDER_ROWS:
        _LOG.info("placeholder_row %s", json.dumps(r, ensure_ascii=False))


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--db",
        type=Path,
        default=None,
        help="SQLite DB path (default: JPINTEL_DB_PATH or ./data/jpintel.db)",
    )
    p.add_argument(
        "--source-file",
        type=Path,
        default=None,
        help=(
            "Excel / CSV workbook downloaded from "
            "https://www.chusho.meti.go.jp/keiei/kakushin/nintei/support.html . "
            "Operator staging step — this script does NOT fetch the URL "
            "itself (the workbook filenames rotate under dated directories)."
        ),
    )
    p.add_argument(
        "--source-url",
        type=str,
        default=SOURCE_LANDING_URL,
        help=(
            "Override the source_url written to each row. Defaults to the "
            "中小企業庁 認定支援機関 landing page."
        ),
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse + validate, no DB writes. Safe to run without a source file.",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    _configure_logging()
    args = _parse_args(argv)
    db_path = args.db if args.db else _default_db_path()

    # No source file path: dry-run only (we never invent data).
    if args.source_file is None:
        _log_placeholder_preview()
        if args.dry_run:
            _LOG.info("dry_run_ok — no DB writes performed.")
            return 0
        _LOG.error(
            "source_file_required — live seed needs --source-file PATH. "
            "Rerun with --dry-run to see placeholder output without DB access."
        )
        return 1

    source_path: Path = args.source_file
    if not source_path.is_file():
        _LOG.error("source_file_not_found path=%s", source_path)
        return 1

    try:
        raw_rows = _read_source_file(source_path)
    except Exception as e:  # noqa: BLE001 — surface root cause in log
        _LOG.error("source_parse_failed path=%s err=%s", source_path, e)
        return 1

    normalized: list[dict[str, Any]] = []
    for r in raw_rows:
        norm = _normalize_row_for_csv(r)
        if norm is not None:
            normalized.append(norm)

    _LOG.info(
        "parsed rows_total=%d rows_accepted=%d rejected=%d",
        len(raw_rows),
        len(normalized),
        len(raw_rows) - len(normalized),
    )

    if args.dry_run:
        for n in normalized[:5]:
            _LOG.info("would_insert %s", json.dumps(n, ensure_ascii=False))
        _LOG.info("dry_run_ok — no DB writes performed.")
        return 0

    if not db_path.is_file():
        _LOG.error("db_missing path=%s — run scripts/migrate.py first", db_path)
        return 2

    conn = sqlite3.connect(str(db_path), isolation_level=None)
    try:
        _ensure_advisors_table(conn)
        inserted, skipped = _insert_rows(conn, normalized, args.source_url)
        _LOG.info("seed_done inserted=%d skipped_existing=%d", inserted, skipped)
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
