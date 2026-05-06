#!/usr/bin/env python3
"""Ingest 食品衛生法 行政処分 / 食中毒事例 into ``am_enforcement_detail``.

Sources (primary 一次資料 only — no aggregators):
  1. 厚生労働省 食中毒統計 食中毒事件一覧 (xlsx)
       https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/
       kenkou_iryou/shokuhin/syokuchu/04.html
     - R6 (2024): /content/001681445.xlsx — 1,043 incidents, 47 prefectures
     - R5 (2023): /content/001472342.xlsx — ~1,000 incidents
     Each row = 都道府県 / 発生月日 / 発生場所 / 原因食品 / 病因物質 /
                原因施設 / 摂食者数 / 患者数 / 死者数。
     都道府県 column is "東京都区部" / "北海道" / "横浜市" 等 政令市保健所単位。
     これは 食中毒 incident で、自治体が把握 → 全国集計、即ち 食品衛生法
     第63条/第69条 公表 系統の primary feed。

  2. 各都道府県 食品衛生法違反事例ページ (年度別 PDF) — 福岡県の半期 PDF が
     最も construct がはっきりしている。Phase 2 で巡回。

Schema mapping:
  - enforcement_kind:
      * 営業停止 / 改善指示 系統 (将来 phase 2 で取り込む) → business_improvement
      * 営業禁止 / 許可取消 → license_revoke
      * 廃棄命令 / 食中毒 incident のみ → other
    Phase 1 (this commit) は all 食中毒 → 'other'。
  - issuing_authority: 都道府県名そのまま ('東京都区部', '北海道', '横浜市').
    これは MHLW 統計が市区分けで集計する単位なので保健所 authority と一致。
  - related_law_ref: '食品衛生法' (固定。条文は事例単位では未確定なので
    汎用ラベルで保持。reason_summary に病因物質 + 原因食品 を畳む)。
  - amount_yen: NULL.
  - reason_summary: 病因物質 + 原因食品 + 原因施設 + 患者数 + 死者数 を一行に。
  - target_name: 原因施設 (e.g. '飲食店', '事業場-給食施設-老人ホーム') —
    特定企業名は MHLW 統計には含まれないので施設カテゴリを target proxy に
    する (公表名なし=NULL は CHECK 制約に当たる前に '不明施設' で fallback)。

Parallel-safe:
  - BEGIN IMMEDIATE + PRAGMA busy_timeout=300000.
  - 1 source 1 commit (small batch).

Idempotent on (issuing_authority, issuance_date, target_name) tuple
(both DB existing + within-batch).

CLI:
    python scripts/ingest/ingest_enforcement_food_hygiene.py \\
        [--db autonomath.db] [--max-rows 250] [--dry-run]
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import logging
import re
import sqlite3
import sys
import unicodedata
from dataclasses import dataclass
from datetime import UTC, date as date_cls, datetime
from pathlib import Path
from urllib.parse import urlparse

try:
    import openpyxl  # type: ignore
except ImportError as exc:  # pragma: no cover
    print(f"ERROR: openpyxl not installed: {exc}", file=sys.stderr)
    raise

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

from scripts.lib.http import HttpClient  # noqa: E402

_LOG = logging.getLogger("autonomath.ingest.food_hygiene")

DEFAULT_DB = REPO_ROOT / "autonomath.db"
USER_AGENT = "AutonoMath/0.1.0 (+https://bookyou.net)"

# ---------------------------------------------------------------------------
# MHLW 食中毒事件一覧 xlsx feed
#   The page lists fiscal-year files going back to H8. We pull R6 + R5
#   (令和6 + 令和5 = calendar 2024 + 2023) — both contain ~1k incidents
#   each so a single combined run easily clears the +200 target.
# ---------------------------------------------------------------------------

MHLW_INDEX_URL = (
    "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/kenkou_iryou/shokuhin/syokuchu/04.html"
)

# (label, xlsx_url, calendar_year_for_dedup_seed)
MHLW_XLSX_FEEDS: list[tuple[str, str, int]] = [
    ("令和6年食中毒発生事例", "https://www.mhlw.go.jp/content/001681445.xlsx", 2024),
    ("令和5年食中毒発生事例", "https://www.mhlw.go.jp/content/001472342.xlsx", 2023),
]


# ---------------------------------------------------------------------------
# Row dataclass
# ---------------------------------------------------------------------------


@dataclass
class FoodEnfRow:
    target_name: str  # 原因施設 / fallback
    issuance_date: str  # ISO yyyy-mm-dd
    issuing_authority: str  # 都道府県名
    enforcement_kind: str  # 'other' (food-poisoning incident)
    reason_summary: str
    related_law_ref: str  # '食品衛生法'
    source_url: str
    extra: dict | None = None


def _normalize(s: str) -> str:
    if not s:
        return ""
    return unicodedata.normalize("NFKC", str(s)).strip()


def _date_to_iso(value) -> str | None:  # type: ignore[no-untyped-def]
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date_cls):
        return value.isoformat()
    s = _normalize(value)
    # Excel sometimes leaves dates as text "2024/5/1" or "R6.5.1"
    m = re.search(r"(20\d{2})[/\-.](\d{1,2})[/\-.](\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    m = re.search(r"R\s*(\d+)[.\-](\d{1,2})[.\-](\d{1,2})", s)
    if m:
        y = 2018 + int(m.group(1))
        mo, d = int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


# ---------------------------------------------------------------------------
# Excel parsing (MHLW 食中毒事件一覧)
# ---------------------------------------------------------------------------


def parse_mhlw_food_xlsx(
    xlsx_bytes: bytes,
    *,
    source_url: str,
) -> list[FoodEnfRow]:
    """Parse one MHLW 食中毒事件一覧 .xlsx and return enforcement rows.

    Sheet layout (single sheet per workbook):
      Row 1: title.
      Row 2: header (都道府県名等 / 発生月日 / 発生場所 / 原因食品 /
                     病因物質 / 原因施設 / 摂食者数 / 患者数 / 死者数).
      Row 3+: data rows. col[0] is empty (number column), col[1..9] = data.
    """
    rows: list[FoodEnfRow] = []
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:  # noqa: BLE001
        _LOG.warning("xlsx open failed %s: %s", source_url, exc)
        return rows

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for raw in ws.iter_rows(values_only=True, min_row=3):
            if not raw or len(raw) < 10:
                continue
            # Layout: (None, pref, date, place, food, agent, facility,
            #          ate, patients, dead)
            pref = _normalize(raw[1])
            date_iso = _date_to_iso(raw[2])
            place = _normalize(raw[3])
            food = _normalize(raw[4])
            agent = _normalize(raw[5])
            facility = _normalize(raw[6])
            ate = raw[7]
            patients = raw[8]
            dead = raw[9]
            if not pref or not date_iso:
                continue
            # Skip aggregate / summary rows masquerading as data
            if pref in ("計", "合計", "総計"):
                continue
            target = facility or "不明施設"
            # Build a compact reason cite-able as 食品衛生法 (always present)
            patients_n = int(patients) if isinstance(patients, (int, float)) else 0
            dead_n = int(dead) if isinstance(dead, (int, float)) else 0
            reason = (
                f"食品衛生法に基づく食中毒事例 / 病因物質: {agent or '不明'} / "
                f"原因食品: {food or '不明'} / 原因施設: {facility or '不明'} / "
                f"発生場所: {place or '不明'} / 患者数: {patients_n}名"
                + (f" / 死者数: {dead_n}名" if dead_n else "")
            )[:1500]
            rows.append(
                FoodEnfRow(
                    target_name=target,
                    issuance_date=date_iso,
                    issuing_authority=pref,
                    enforcement_kind="other",
                    reason_summary=reason,
                    related_law_ref="食品衛生法",
                    source_url=source_url,
                    extra={
                        "feed": "mhlw_shokuchudoku_itiran",
                        "occurrence_place": place or None,
                        "causal_food": food or None,
                        "pathogen": agent or None,
                        "facility_category": facility or None,
                        "patients": patients_n,
                        "deaths": dead_n,
                    },
                )
            )
    return rows


def fetch_mhlw_food_feeds(http: HttpClient) -> list[FoodEnfRow]:
    out: list[FoodEnfRow] = []
    for label, url, _yr in MHLW_XLSX_FEEDS:
        res = http.get(url, max_bytes=10 * 1024 * 1024)
        if not res.ok:
            _LOG.warning("[mhlw-food] fetch fail %s status=%s", url, res.status)
            continue
        rs = parse_mhlw_food_xlsx(res.body, source_url=url)
        _LOG.info("[mhlw-food] %s rows=%d url=%s", label, len(rs), url)
        out.extend(rs)
    return out


# ---------------------------------------------------------------------------
# DB layer
# ---------------------------------------------------------------------------


def _slug8(name: str, date: str, extra: str = "") -> str:
    h = hashlib.sha1(f"{name}|{date}|{extra}".encode("utf-8")).hexdigest()
    return h[:8]


# Map a few common 都道府県 / 政令市 labels to slugs for canonical_id stability.
# Anything not in this map gets a sha1-derived slug fallback.
PREF_SLUG_MAP: dict[str, str] = {
    "北海道": "hokkaido",
    "札幌市": "sapporo",
    "青森県": "aomori",
    "岩手県": "iwate",
    "宮城県": "miyagi",
    "仙台市": "sendai",
    "秋田県": "akita",
    "山形県": "yamagata",
    "福島県": "fukushima",
    "いわき市": "iwaki",
    "郡山市": "koriyama",
    "茨城県": "ibaraki",
    "栃木県": "tochigi",
    "宇都宮市": "utsunomiya",
    "群馬県": "gunma",
    "高崎市": "takasaki",
    "前橋市": "maebashi",
    "埼玉県": "saitama",
    "川越市": "kawagoe",
    "さいたま市": "saitamashi",
    "越谷市": "koshigaya",
    "川口市": "kawaguchi",
    "千葉県": "chiba",
    "千葉市": "chibashi",
    "船橋市": "funabashi",
    "柏市": "kashiwa",
    "東京都": "tokyo",
    "東京都区部": "tokyo23",
    "八王子市": "hachioji",
    "町田市": "machida",
    "神奈川県": "kanagawa",
    "横浜市": "yokohama",
    "川崎市": "kawasaki",
    "相模原市": "sagamihara",
    "横須賀市": "yokosuka",
    "藤沢市": "fujisawa",
    "茅ヶ崎市": "chigasaki",
    "新潟県": "niigata",
    "新潟市": "niigatashi",
    "富山県": "toyama",
    "富山市": "toyamashi",
    "石川県": "ishikawa",
    "金沢市": "kanazawa",
    "福井県": "fukui",
    "山梨県": "yamanashi",
    "甲府市": "kofu",
    "長野県": "nagano",
    "長野市": "naganoshi",
    "松本市": "matsumoto",
    "岐阜県": "gifu",
    "岐阜市": "gifushi",
    "静岡県": "shizuoka",
    "静岡市": "shizuokashi",
    "浜松市": "hamamatsu",
    "愛知県": "aichi",
    "名古屋市": "nagoya",
    "豊田市": "toyota",
    "豊橋市": "toyohashi",
    "岡崎市": "okazaki",
    "三重県": "mie",
    "滋賀県": "shiga",
    "大津市": "otsu",
    "京都府": "kyoto",
    "京都市": "kyotoshi",
    "大阪府": "osaka",
    "大阪市": "osakashi",
    "堺市": "sakai",
    "東大阪市": "higashiosaka",
    "高槻市": "takatsuki",
    "兵庫県": "hyogo",
    "神戸市": "kobe",
    "姫路市": "himeji",
    "尼崎市": "amagasaki",
    "西宮市": "nishinomiya",
    "明石市": "akashi",
    "奈良県": "nara",
    "奈良市": "narashi",
    "和歌山県": "wakayama",
    "和歌山市": "wakayamashi",
    "鳥取県": "tottori",
    "島根県": "shimane",
    "松江市": "matsue",
    "岡山県": "okayama",
    "岡山市": "okayamashi",
    "倉敷市": "kurashiki",
    "広島県": "hiroshima",
    "広島市": "hiroshimashi",
    "福山市": "fukuyama",
    "呉市": "kure",
    "山口県": "yamaguchi",
    "下関市": "shimonoseki",
    "徳島県": "tokushima",
    "香川県": "kagawa",
    "高松市": "takamatsu",
    "愛媛県": "ehime",
    "松山市": "matsuyama",
    "高知県": "kochi",
    "高知市": "kochishi",
    "福岡県": "fukuoka",
    "福岡市": "fukuokashi",
    "北九州市": "kitakyushu",
    "久留米市": "kurume",
    "佐賀県": "saga",
    "長崎県": "nagasaki",
    "長崎市": "nagasakishi",
    "佐世保市": "sasebo",
    "熊本県": "kumamoto",
    "熊本市": "kumamotoshi",
    "大分県": "oita",
    "大分市": "oitashi",
    "宮崎県": "miyazaki",
    "宮崎市": "miyazakishi",
    "鹿児島県": "kagoshima",
    "鹿児島市": "kagoshimashi",
    "沖縄県": "okinawa",
    "那覇市": "naha",
}


def _pref_slug(label: str) -> str:
    if label in PREF_SLUG_MAP:
        return PREF_SLUG_MAP[label]
    h = hashlib.sha1(label.encode("utf-8")).hexdigest()
    return f"pref-{h[:6]}"


def ensure_tables(conn: sqlite3.Connection) -> None:
    for tbl in ("am_entities", "am_enforcement_detail"):
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (tbl,),
        ).fetchone()
        if not row:
            raise SystemExit(f"missing table '{tbl}' — apply migrations first")


def existing_dedup_keys(
    conn: sqlite3.Connection,
) -> set[tuple[str, str, str]]:
    """Return (issuing_authority, issuance_date, target_name) keys for
    rows we'd consider duplicates of MHLW 食品衛生 feed entries.

    Scope it tightly to the food-hygiene universe to avoid touching the
    47 都道府県 労働局 rows already loaded:
       reason_summary LIKE '%食品衛生法%' OR
       reason_summary LIKE '%食中毒%'
    """
    out: set[tuple[str, str, str]] = set()
    cur = conn.execute(
        """
        SELECT issuing_authority, issuance_date, target_name
        FROM am_enforcement_detail
        WHERE reason_summary LIKE '%食品衛生法%'
           OR reason_summary LIKE '%食中毒%'
        """,
    )
    for a, d, n in cur.fetchall():
        if a and d and n:
            out.add((a, d, n))
    return out


def upsert_entity(
    conn: sqlite3.Connection,
    canonical_id: str,
    primary_name: str,
    url: str,
    raw_json: str,
    now_iso: str,
) -> None:
    domain = urlparse(url).netloc or None
    conn.execute(
        """
        INSERT INTO am_entities (
            canonical_id, record_kind, source_topic, source_record_index,
            primary_name, authority_canonical, confidence,
            source_url, source_url_domain, fetched_at, raw_json,
            canonical_status, citation_status
        ) VALUES (?, 'enforcement', 'mhlw_food_shokuchudoku', NULL,
                  ?, NULL, 0.85, ?, ?, ?, ?, 'active', 'ok')
        ON CONFLICT(canonical_id) DO UPDATE SET
            primary_name      = excluded.primary_name,
            source_url        = excluded.source_url,
            source_url_domain = excluded.source_url_domain,
            fetched_at        = excluded.fetched_at,
            raw_json          = excluded.raw_json,
            updated_at        = datetime('now')
        """,
        (
            canonical_id,
            primary_name[:500],
            url,
            domain,
            now_iso,
            raw_json,
        ),
    )


def insert_enforcement(
    conn: sqlite3.Connection,
    entity_id: str,
    row: FoodEnfRow,
    now_iso: str,
) -> None:
    conn.execute(
        """
        INSERT INTO am_enforcement_detail (
            entity_id, houjin_bangou, target_name, enforcement_kind,
            issuing_authority, issuance_date, exclusion_start, exclusion_end,
            reason_summary, related_law_ref, amount_yen,
            source_url, source_fetched_at
        ) VALUES (?, NULL, ?, ?, ?, ?, NULL, NULL, ?, ?, NULL, ?, ?)
        """,
        (
            entity_id,
            row.target_name[:500],
            row.enforcement_kind,
            row.issuing_authority,
            row.issuance_date,
            row.reason_summary[:4000],
            row.related_law_ref[:1000],
            row.source_url,
            now_iso,
        ),
    )


def write_rows(
    conn: sqlite3.Connection,
    rows: list[FoodEnfRow],
    *,
    now_iso: str,
    max_rows: int | None,
) -> tuple[int, int, int, dict[str, int]]:
    """Insert rows in a BEGIN IMMEDIATE block.

    Returns (inserted, dup_db, dup_batch, by_authority).
    Stops once max_rows new rows are inserted.
    """
    if not rows:
        return 0, 0, 0, {}
    db_keys = existing_dedup_keys(conn)
    # Within-batch dedup uses a finer key (pref, date, target, pathogen,
    # food) because MHLW 統計 includes many same-day same-prefecture
    # アニサキス incidents that all collapse to (pref, date, '不明施設')
    # otherwise. The DB-side dedup tuple stays (pref, date, target) for
    # storage compatibility.
    batch_keys_fine: set[tuple[str, str, str, str, str]] = set()
    inserted = 0
    dup_db = 0
    dup_batch = 0
    by_auth: dict[str, int] = {}
    try:
        conn.execute("BEGIN IMMEDIATE")
        for r in rows:
            if max_rows is not None and inserted >= max_rows:
                break
            key_db = (r.issuing_authority, r.issuance_date, r.target_name)
            ext = r.extra or {}
            key_batch = (
                r.issuing_authority,
                r.issuance_date,
                r.target_name,
                str(ext.get("pathogen") or ""),
                str(ext.get("causal_food") or ""),
            )
            if key_db in db_keys:
                dup_db += 1
                continue
            if key_batch in batch_keys_fine:
                dup_batch += 1
                continue
            batch_keys_fine.add(key_batch)

            slug = _pref_slug(r.issuing_authority)
            ext_for_slug = r.extra or {}
            seq_seed = "|".join(
                [
                    str(ext_for_slug.get("pathogen") or ""),
                    str(ext_for_slug.get("causal_food") or ""),
                    str(ext_for_slug.get("occurrence_place") or ""),
                    str(ext_for_slug.get("patients") or ""),
                ]
            )
            seq = _slug8(r.target_name, r.issuance_date, seq_seed)
            canonical_id = (
                f"enforcement:food-hygiene-{slug}-{r.issuance_date.replace('-', '')}-{seq}"
            )
            primary_name = f"{r.target_name} ({r.issuance_date}) - {r.issuing_authority} 食中毒事例"
            raw_json = json.dumps(
                {
                    "issuing_authority": r.issuing_authority,
                    "target_name": r.target_name,
                    "issuance_date": r.issuance_date,
                    "enforcement_kind": r.enforcement_kind,
                    "related_law_ref": r.related_law_ref,
                    "reason_summary": r.reason_summary,
                    "source_url": r.source_url,
                    "extra": r.extra or {},
                    "source_attribution": "厚生労働省ウェブサイト",
                    "license": "政府機関の著作物（出典明記で転載引用可）",
                },
                ensure_ascii=False,
            )
            try:
                upsert_entity(
                    conn,
                    canonical_id,
                    primary_name,
                    r.source_url,
                    raw_json,
                    now_iso,
                )
                insert_enforcement(conn, canonical_id, r, now_iso)
                inserted += 1
                by_auth[r.issuing_authority] = by_auth.get(r.issuing_authority, 0) + 1
            except sqlite3.Error as exc:
                _LOG.error(
                    "DB insert err pref=%s name=%s date=%s: %s",
                    r.issuing_authority,
                    r.target_name,
                    r.issuance_date,
                    exc,
                )
                continue
        conn.commit()
    except sqlite3.Error as exc:
        _LOG.error("BEGIN/commit failed: %s", exc)
        try:
            conn.rollback()
        except sqlite3.Error:
            pass
    return inserted, dup_db, dup_batch, by_auth


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--db", type=Path, default=DEFAULT_DB)
    ap.add_argument(
        "--max-rows", type=int, default=250, help="cap inserts at this number (default 250)"
    )
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--verbose", "-v", action="store_true")
    return ap.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )

    http = HttpClient(user_agent=USER_AGENT)
    now_iso = (
        datetime.now(UTC)
        .isoformat(timespec="seconds")
        .replace(
            "+00:00",
            "Z",
        )
    )

    rows = fetch_mhlw_food_feeds(http)
    _LOG.info("total parsed rows=%d", len(rows))

    if args.dry_run:
        for r in rows[:5]:
            _LOG.info(
                "sample: pref=%s date=%s target=%s reason=%s",
                r.issuing_authority,
                r.issuance_date,
                r.target_name,
                r.reason_summary[:100],
            )
        http.close()
        return 0

    if not args.db.exists():
        _LOG.error("autonomath.db missing: %s", args.db)
        http.close()
        return 2

    conn = sqlite3.connect(str(args.db))
    conn.execute("PRAGMA busy_timeout=300000")
    conn.execute("PRAGMA foreign_keys=ON")
    ensure_tables(conn)

    inserted, dup_db, dup_batch, by_auth = write_rows(
        conn,
        rows,
        now_iso=now_iso,
        max_rows=args.max_rows,
    )
    try:
        conn.close()
    except sqlite3.Error:
        pass
    http.close()

    _LOG.info(
        "done parsed=%d inserted=%d dup_db=%d dup_batch=%d",
        len(rows),
        inserted,
        dup_db,
        dup_batch,
    )
    print(
        f"food-hygiene ingest: parsed={len(rows)} inserted={inserted} "
        f"dup_db={dup_db} dup_batch={dup_batch}"
    )
    print("by authority (top 15):")
    for k, v in sorted(by_auth.items(), key=lambda x: -x[1])[:15]:
        print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
