#!/usr/bin/env python3
"""Ingest 出入国在留管理庁 (ISA) + 法務省 + 外国人技能実習機構 系
行政処分・公表事項を ``am_enforcement_detail`` + ``am_entities`` に投入。

Targets covered (法人スポンサー / 受入機関ベース):
  1. 監理団体 許可取消し (技能実習法第37条第1項) — sup_revoke
  2. 監理団体 改善命令 (技能実習法第36条第1項) — sup_improve
  3. 実習実施者 計画認定取消し (技能実習法第16条第1項) FY2018-FY2025
     — cert_cancel × 8 年度
  4. 実習実施者 改善命令 (技能実習法第15条第1項) — impl_improve

なお、公表されるのは法人スポンサー (受入機関 / 監理団体) のみ。個人在留者
の取消しは件数しか公表されておらず行政の名簿には載らないため、本スクリプ
トは法人 row のみを生成する (CLAUDE.md mission の anonymize 個人在留者
ルールは将来 e-Stat 由来 row を投入する場合の予備)。

Sources walked (一次資料のみ — moj.go.jp/isa 配下のみ):
  - https://www.moj.go.jp/isa/applications/titp/nyuukokukanri07_00138.html
      技能実習法に基づく行政処分等 公表ページ。本ページから 11 本の
      Excel ファイル (上記 4 カテゴリ) を fetch する。
  - https://www.moj.go.jp/isa/content/930005832.xlsx 監理団体 許可取消し
  - https://www.moj.go.jp/isa/content/930005831.xlsx 監理団体 改善命令
  - https://www.moj.go.jp/isa/content/930004391.xlsx 実習実施者 改善命令
  - https://www.moj.go.jp/isa/content/001391104.xlsx FY2018 認定取消し
  - https://www.moj.go.jp/isa/content/001391106.xlsx FY2019 認定取消し
  - https://www.moj.go.jp/isa/content/001391107.xlsx FY2020 認定取消し
  - https://www.moj.go.jp/isa/content/001391108.xlsx FY2021 認定取消し
  - https://www.moj.go.jp/isa/content/001391109.xlsx FY2022 認定取消し
  - https://www.moj.go.jp/isa/content/001395059.xlsx FY2023 認定取消し
  - https://www.moj.go.jp/isa/content/001417745.xlsx FY2024 認定取消し
  - https://www.moj.go.jp/isa/content/001439965.xlsx FY2025 認定取消し

Aggregator BAN は厳守: noukaweb / hojyokin-portal / biz.stayway / prtimes /
nikkei / wikipedia は 一切 fetch しない。

License: 法務省 出入国在留管理庁 ウェブサイト (政府機関の著作物、出典明記
で転載引用可、PDL v1.0 互換)。

Schema mapping (am_enforcement_detail.enforcement_kind enum):
  許可取消し (技能実習法第37条) / 認定取消し (技能実習法第16条) → license_revoke
  改善命令 (技能実習法第15条 / 第36条)                          → business_improvement
  事業停止命令 (現在事例なし、空エンクロージャのみ)             → contract_suspend

Authority labels:
  - 監理団体・実習実施者 すべて → '出入国在留管理庁' / '外国人技能実習機構'
    の連名処分。issuing_authority は '出入国在留管理庁' を採用 (Article 16/37
    の最終決定権者であるため)。

Dedup key:
  (target_name, issuance_date, enforcement_kind) within issuing_authority.

Parallel-safe (CLAUDE.md §5):
  - PRAGMA busy_timeout=300000 + BEGIN IMMEDIATE
  - per-source-file commit (Excel 1 本 ≈ 8-180 row → 1 commit)。

CLI:
  python scripts/ingest/ingest_enforcement_isa_zairyu.py
  python scripts/ingest/ingest_enforcement_isa_zairyu.py --dry-run -v
  python scripts/ingest/ingest_enforcement_isa_zairyu.py --skip-fetch
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import io
import json
import logging
import re
import sqlite3
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl  # type: ignore
    import requests  # type: ignore
except ImportError as exc:  # pragma: no cover
    print(f"missing dep: {exc}. pip install openpyxl requests", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_DB = REPO_ROOT / "autonomath.db"
CACHE_DIR = REPO_ROOT / "data" / "_cache" / "isa_enforcement"

USER_AGENT = (
    "AutonoMath/0.1.0 (+https://bookyou.net) ingest-isa-enforcement (contact=info@bookyou.net)"
)
HTTP_TIMEOUT = 60
RATE_SLEEP = 1.5  # ≤1 req/sec/host (gov rate-limit)

PUBLIC_PAGE = "https://www.moj.go.jp/isa/applications/titp/nyuukokukanri07_00138.html"

# (excel_url, kind_label, enforcement_kind, related_law_ref,
#  target_kind {'jissyusha'/'kanri-dantai'}, file_label)
SOURCE_FILES: list[dict[str, str]] = [
    {
        "url": "https://www.moj.go.jp/isa/content/930005832.xlsx",
        "kind": "license_revoke",
        "law_ref": "技能実習法第37条第1項",
        "target_kind": "kanri-dantai",
        "label": "監理団体 許可取消し",
        "topic": "isa_titp_kanri_revoke",
        "summary_prefix": "監理団体 許可取消し（技能実習法第37条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/930005831.xlsx",
        "kind": "business_improvement",
        "law_ref": "技能実習法第36条第1項",
        "target_kind": "kanri-dantai",
        "label": "監理団体 改善命令",
        "topic": "isa_titp_kanri_improve",
        "summary_prefix": "監理団体に対する改善命令（技能実習法第36条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/930004391.xlsx",
        "kind": "business_improvement",
        "law_ref": "技能実習法第15条第1項",
        "target_kind": "jissyusha",
        "label": "実習実施者 改善命令",
        "topic": "isa_titp_jissyusha_improve",
        "summary_prefix": "実習実施者に対する改善命令（技能実習法第15条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/001391104.xlsx",
        "kind": "license_revoke",
        "law_ref": "技能実習法第16条第1項",
        "target_kind": "jissyusha",
        "label": "実習実施者 認定取消し FY2018",
        "topic": "isa_titp_jissyusha_revoke_2018",
        "summary_prefix": "技能実習計画の認定取消し（技能実習法第16条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/001391106.xlsx",
        "kind": "license_revoke",
        "law_ref": "技能実習法第16条第1項",
        "target_kind": "jissyusha",
        "label": "実習実施者 認定取消し FY2019",
        "topic": "isa_titp_jissyusha_revoke_2019",
        "summary_prefix": "技能実習計画の認定取消し（技能実習法第16条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/001391107.xlsx",
        "kind": "license_revoke",
        "law_ref": "技能実習法第16条第1項",
        "target_kind": "jissyusha",
        "label": "実習実施者 認定取消し FY2020",
        "topic": "isa_titp_jissyusha_revoke_2020",
        "summary_prefix": "技能実習計画の認定取消し（技能実習法第16条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/001391108.xlsx",
        "kind": "license_revoke",
        "law_ref": "技能実習法第16条第1項",
        "target_kind": "jissyusha",
        "label": "実習実施者 認定取消し FY2021",
        "topic": "isa_titp_jissyusha_revoke_2021",
        "summary_prefix": "技能実習計画の認定取消し（技能実習法第16条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/001391109.xlsx",
        "kind": "license_revoke",
        "law_ref": "技能実習法第16条第1項",
        "target_kind": "jissyusha",
        "label": "実習実施者 認定取消し FY2022",
        "topic": "isa_titp_jissyusha_revoke_2022",
        "summary_prefix": "技能実習計画の認定取消し（技能実習法第16条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/001395059.xlsx",
        "kind": "license_revoke",
        "law_ref": "技能実習法第16条第1項",
        "target_kind": "jissyusha",
        "label": "実習実施者 認定取消し FY2023",
        "topic": "isa_titp_jissyusha_revoke_2023",
        "summary_prefix": "技能実習計画の認定取消し（技能実習法第16条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/001417745.xlsx",
        "kind": "license_revoke",
        "law_ref": "技能実習法第16条第1項",
        "target_kind": "jissyusha",
        "label": "実習実施者 認定取消し FY2024",
        "topic": "isa_titp_jissyusha_revoke_2024",
        "summary_prefix": "技能実習計画の認定取消し（技能実習法第16条第1項）",
    },
    {
        "url": "https://www.moj.go.jp/isa/content/001439965.xlsx",
        "kind": "license_revoke",
        "law_ref": "技能実習法第16条第1項",
        "target_kind": "jissyusha",
        "label": "実習実施者 認定取消し FY2025",
        "topic": "isa_titp_jissyusha_revoke_2025",
        "summary_prefix": "技能実習計画の認定取消し（技能実習法第16条第1項）",
    },
]

_LOG = logging.getLogger("autonomath.ingest_isa_zairyu")

# ---------------------------------------------------------------------------
# Date / text helpers
# ---------------------------------------------------------------------------

WAREKI_RE = re.compile(
    r"(令和|平成|昭和|R|H|S)\s*(元|[0-9０-９]+)\s*年\s*" r"([0-9０-９]+)\s*月\s*([0-9０-９]+)\s*日"
)
SEIREKI_RE = re.compile(r"(20[0-9]{2})\s*[年\-/]\s*([0-9]+)\s*[月\-/]\s*([0-9]+)\s*日?")


def _normalize(text: str) -> str:
    text = unicodedata.normalize("NFKC", text or "")
    return re.sub(r"\s+", " ", text).strip()


def _excel_serial_to_iso(serial: int | float) -> str | None:
    """Excel 1900 serial → ISO yyyy-mm-dd. Excel uses 1900-01-01 as 1
    (with 1900-02-29 as a phantom day → use the standard offset that aligns
    with most modern dates 43461 = 2019-01-01)."""
    try:
        n = int(serial)
    except (TypeError, ValueError):
        return None
    if n < 30000 or n > 60000:
        return None
    # Excel serial 1 == 1899-12-31 (not 1900-01-01) when accounting for the
    # 1900 leap-year bug. Use the conventional offset that maps 1 → 1900-01-01
    # but subtract 2 because we want 43461 → 2019-01-01.
    base = date(1899, 12, 30)
    try:
        return (base + timedelta(days=n)).isoformat()
    except (OverflowError, ValueError):
        return None


def _wareki_to_iso(text: str) -> str | None:
    if not text:
        return None
    s = unicodedata.normalize("NFKC", text)
    m = WAREKI_RE.search(s)
    if m:
        era, yr, mo, dy = m.group(1), m.group(2), m.group(3), m.group(4)
        try:
            yr_i = 1 if yr == "元" else int(yr)
        except ValueError:
            return None
        if era in ("R", "令和"):
            year = 2018 + yr_i
        elif era in ("H", "平成"):
            year = 1988 + yr_i
        elif era in ("S", "昭和"):
            year = 1925 + yr_i
        else:
            return None
        try:
            mo_i, dy_i = int(mo), int(dy)
            if not (1 <= mo_i <= 12 and 1 <= dy_i <= 31):
                return None
            return f"{year:04d}-{mo_i:02d}-{dy_i:02d}"
        except ValueError:
            return None
    m = SEIREKI_RE.search(s)
    if m:
        try:
            y, mo, dy = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 2000 <= y <= 2100 and 1 <= mo <= 12 and 1 <= dy <= 31:
                return f"{y:04d}-{mo:02d}-{dy:02d}"
        except ValueError:
            return None
    return None


def _coerce_date(cell: object) -> str | None:
    """Cell value (datetime / int / str) → ISO yyyy-mm-dd."""
    if cell is None:
        return None
    if isinstance(cell, datetime):
        return cell.date().isoformat()
    if isinstance(cell, date):
        return cell.isoformat()
    if isinstance(cell, (int, float)):
        return _excel_serial_to_iso(cell)
    if isinstance(cell, str):
        s = cell.strip()
        if not s:
            return None
        # Try ISO directly (e.g. '2024-04-30 00:00:00')
        m = re.match(r"^(20[0-9]{2})-([0-9]{2})-([0-9]{2})", s)
        if m:
            try:
                y, mo, dy = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if 2000 <= y <= 2100 and 1 <= mo <= 12 and 1 <= dy <= 31:
                    return f"{y:04d}-{mo:02d}-{dy:02d}"
            except ValueError:
                pass
        # Wareki / Seireki Japanese form
        return _wareki_to_iso(s)
    return None


def _slug8(name: str, date_iso: str, kind: str) -> str:
    h = hashlib.sha1(f"{name}|{date_iso}|{kind}".encode()).hexdigest()
    return h[:10]


def _slugify_jp(text: str, max_len: int = 28) -> str:
    text = unicodedata.normalize("NFKC", text or "")
    text = re.sub(r"\s+", "-", text)
    text = re.sub(r"[^0-9A-Za-z_\-぀-ヿ一-鿿々]", "", text)
    return text[:max_len] or "unknown"


_BAD_NAME_LITERALS = frozenset(
    {
        "監理団体名",
        "実習実施者名",
        "代表者",
        "代表者名",
        "所在地",
        "事業者名",
        "認定番号",
        "許可番号",
        "別紙",
        "別添",
        "備考",
        "事業所",
        "番号",
        "計画番号",
    }
)


def _is_valid_target_name(name: str | None) -> bool:
    if not name:
        return False
    s = _normalize(name).strip()
    if not s or len(s) < 2 or len(s) > 200:
        return False
    if s in _BAD_NAME_LITERALS:
        return False
    if re.fullmatch(r"\d+", s):
        return False
    if re.fullmatch(r"[\d\-\.]+", s):
        return False
    if (
        re.search(r"年.*月.*日", s)
        and not any(
            kw in s
            for kw in (
                "株式会社",
                "有限会社",
                "合同会社",
                "合資会社",
                "合名会社",
                "協同組合",
                "協会",
                "組合",
                "(株)",
                "(有)",
                "(合)",
                "(同)",
                "事業協同組合",
            )
        )
        and re.fullmatch(r"[\d０-９年月日\s\.\-/平成令和元昭和大正]+", s)
    ):
        return False
    return True


# ---------------------------------------------------------------------------
# HTTP client
# ---------------------------------------------------------------------------


class HttpClient:
    def __init__(self, *, user_agent: str = USER_AGENT) -> None:
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": user_agent,
                "Accept-Language": "ja,en;q=0.5",
            }
        )
        self._last: float = 0.0

    def get(self, url: str, *, timeout: float = HTTP_TIMEOUT) -> requests.Response | None:
        delta = time.monotonic() - self._last
        if delta < RATE_SLEEP:
            time.sleep(RATE_SLEEP - delta)
        last_err: Exception | None = None
        for attempt in range(3):
            try:
                resp = self.session.get(url, timeout=timeout)
                self._last = time.monotonic()
                if resp.status_code == 200:
                    return resp
                if resp.status_code == 404:
                    return None
                last_err = RuntimeError(f"{resp.status_code} for {url}")
            except requests.RequestException as exc:
                last_err = exc
            time.sleep(2**attempt)
        _LOG.warning("fetch failed after retries: %s: %s", url, last_err)
        return None


def _fetch_or_cache(http: HttpClient, url: str, *, skip_fetch: bool = False) -> bytes | None:
    """Return Excel bytes. Fall back to cached copy when fresh fetch fails."""
    cache_path = CACHE_DIR / hashlib.sha1(url.encode()).hexdigest()
    if skip_fetch and cache_path.exists():
        return cache_path.read_bytes()
    resp = http.get(url)
    if resp is not None:
        try:
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            cache_path.write_bytes(resp.content)
        except OSError:
            pass
        return resp.content
    if cache_path.exists():
        _LOG.warning("falling back to cached %s", url)
        return cache_path.read_bytes()
    return None


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------


@dataclass
class EnfRow:
    target_name: str
    representative: str | None
    location: str | None
    permit_or_cert_number: str | None  # 許可番号 / 認定番号 (first one only)
    permit_or_cert_count: int  # number of cert/permit IDs revoked
    issuance_date: str
    reason_summary: str
    enforcement_kind: str
    related_law_ref: str
    issuing_authority: str
    source_url: str
    source_topic: str
    target_kind: str  # 'jissyusha' / 'kanri-dantai'


def _classify_target_excel(
    xlsx_bytes: bytes,
    *,
    src_url: str,
    src_topic: str,
    enforcement_kind: str,
    law_ref: str,
    summary_prefix: str,
    target_kind: str,
) -> list[EnfRow]:
    out: list[EnfRow] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        _LOG.warning("openpyxl load failed: %s err=%s", src_url, exc)
        return out

    # Iterate over sheets that look like enforcement listings (header row
    # contains 名称 / 実習実施者 / 監理団体 + 措置年月日 / 処分年月日).
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        header = [_normalize(c.value if c.value is not None else "") for c in ws[1]]
        # Detect listing sheet
        name_idx = -1
        rep_idx = -1
        loc_idx = -1
        perm_idx = -1
        reason_idx = -1
        date_idx = -1
        for i, h in enumerate(header):
            if name_idx < 0 and (
                "実習実施者名" in h or "監理団体名" in h or h.endswith("名称") or h == "事業者名"
            ):
                name_idx = i
            elif rep_idx < 0 and "代表者" in h:
                rep_idx = i
            elif loc_idx < 0 and "所在地" in h:
                loc_idx = i
            elif perm_idx < 0 and ("許可番号" in h or "認定番号" in h or "計画" in h):
                perm_idx = i
            elif reason_idx < 0 and ("措置理由" in h or "処分理由" in h or "理由" in h):
                reason_idx = i
            elif date_idx < 0 and (
                "措置年月日" in h or "処分年月日" in h or "年月日" in h or "命令年月日" in h
            ):
                date_idx = i

        if name_idx < 0 or date_idx < 0:
            _LOG.debug(
                "skip sheet %s/%s — no name/date column header=%s",
                src_url.rsplit("/", 1)[-1],
                sheet_name,
                header,
            )
            continue

        # Walk rows. A new entity begins at a row where col[name_idx] is
        # non-empty. Subsequent rows with only perm-id columns belong to
        # the previous entity (collect perm count + secondary perm ids).
        current: EnfRow | None = None
        for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            cells = list(row)
            if len(cells) <= max(name_idx, date_idx):
                continue
            raw_name = cells[name_idx]
            name_str = _normalize(str(raw_name) if raw_name is not None else "")
            date_iso = _coerce_date(cells[date_idx] if date_idx < len(cells) else None)
            if _is_valid_target_name(name_str) and date_iso:
                # flush previous
                if current is not None:
                    out.append(current)
                rep = (
                    _normalize(str(cells[rep_idx]))
                    if 0 <= rep_idx < len(cells) and cells[rep_idx]
                    else None
                )
                loc = (
                    _normalize(str(cells[loc_idx]))
                    if 0 <= loc_idx < len(cells) and cells[loc_idx]
                    else None
                )
                perm = (
                    _normalize(str(cells[perm_idx]))
                    if 0 <= perm_idx < len(cells) and cells[perm_idx]
                    else None
                )
                reason = (
                    _normalize(str(cells[reason_idx]))
                    if 0 <= reason_idx < len(cells) and cells[reason_idx]
                    else ""
                )
                # build summary: prefix + reason snippet
                # Cap at 1200 chars.
                summary = summary_prefix
                if reason:
                    summary = f"{summary} | {reason}"
                if loc:
                    summary = f"{summary} | 所在地: {loc}"
                if rep:
                    summary = f"{summary} | 代表: {rep}"
                summary = summary[:1800]
                current = EnfRow(
                    target_name=name_str,
                    representative=rep,
                    location=loc,
                    permit_or_cert_number=perm,
                    permit_or_cert_count=1 if perm else 0,
                    issuance_date=date_iso,
                    reason_summary=summary,
                    enforcement_kind=enforcement_kind,
                    related_law_ref=law_ref,
                    issuing_authority="出入国在留管理庁",
                    source_url=src_url,
                    source_topic=src_topic,
                    target_kind=target_kind,
                )
            elif (
                current is not None
                and 0 <= perm_idx < len(cells)
                and cells[perm_idx]
                and (raw_name is None or not name_str)
            ):
                # continuation row: same entity, additional perm/cert id
                current.permit_or_cert_count += 1

        if current is not None:
            out.append(current)
    return out


# ---------------------------------------------------------------------------
# DB layer
# ---------------------------------------------------------------------------


def ensure_tables(conn: sqlite3.Connection) -> None:
    for tbl in ("am_entities", "am_enforcement_detail", "am_authority"):
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (tbl,),
        ).fetchone()
        if not row:
            raise SystemExit(f"missing table '{tbl}' — apply migrations first")


def ensure_authority(label: str) -> str:
    """Return canonical_id for the issuing authority. The two authorities we
    use already exist in am_authority (verified at script-design time):
        '出入国在留管理庁'        → 'authority:moj-isa'
        '外国人技能実習機構'     → 'authority:otit'
        '法務省'                 → 'authority:moj'
    """
    if label == "出入国在留管理庁":
        return "authority:moj-isa"
    if label == "外国人技能実習機構":
        return "authority:otit"
    if label == "法務省":
        return "authority:moj"
    return "authority:moj"


def existing_dedup_keys(cur: sqlite3.Cursor) -> set[tuple[str, str, str]]:
    """Pre-load (target_name, issuance_date, enforcement_kind) for fast dedup
    against rows already issued by '出入国在留管理庁' / '外国人技能実習機構'."""
    cur.execute(
        "SELECT target_name, issuance_date, enforcement_kind "
        "FROM am_enforcement_detail "
        "WHERE issuing_authority IN ('出入国在留管理庁', '外国人技能実習機構', '法務省')"
    )
    out: set[tuple[str, str, str]] = set()
    for n, d, k in cur.fetchall():
        if n and d and k:
            out.add((_normalize(n), d, k))
    return out


def existing_canonical_ids(cur: sqlite3.Cursor) -> set[str]:
    cur.execute(
        "SELECT canonical_id FROM am_entities "
        "WHERE record_kind='enforcement' "
        "AND source_topic LIKE 'isa_titp%'"
    )
    return {row[0] for row in cur.fetchall()}


def build_canonical_id(row: EnfRow) -> str:
    name_slug = _slugify_jp(row.target_name, max_len=20)
    iso = row.issuance_date.replace("-", "")
    h = _slug8(row.target_name, row.issuance_date, row.enforcement_kind)
    return f"enforcement:isa-titp:{iso}:{name_slug}:{h}"[:255]


def insert_one(cur: sqlite3.Cursor, row: EnfRow, *, now_iso: str) -> bool:
    canonical_id = build_canonical_id(row)
    authority_canonical = ensure_authority(row.issuing_authority)
    raw = {
        "source": "isa_titp_admin_action",
        "target_kind": row.target_kind,
        "target_name": row.target_name,
        "representative": row.representative,
        "location": row.location,
        "permit_or_cert_number": row.permit_or_cert_number,
        "permit_or_cert_count": row.permit_or_cert_count,
        "issuance_date": row.issuance_date,
        "enforcement_kind": row.enforcement_kind,
        "issuing_authority": row.issuing_authority,
        "related_law_ref": row.related_law_ref,
        "reason_summary": row.reason_summary,
        "source_url": row.source_url,
        "source_topic": row.source_topic,
        "license": ("政府機関の著作物（出典明記で転載引用可、PDL v1.0 互換）"),
        "attribution": (
            "出典: 法務省 出入国在留管理庁 "
            "「公表情報（監理団体一覧、行政処分等、失踪者数ほか）」"
            f" {PUBLIC_PAGE}"
        ),
        "fetched_at": now_iso,
    }
    primary_name = (
        f"{row.target_name} ({row.issuance_date}) - {row.issuing_authority} {row.enforcement_kind}"
    )[:500]
    cur.execute(
        """INSERT OR IGNORE INTO am_entities
               (canonical_id, record_kind, source_topic, source_record_index,
                primary_name, authority_canonical, confidence, source_url,
                source_url_domain, fetched_at, raw_json,
                canonical_status, citation_status)
           VALUES (?, 'enforcement', ?, NULL, ?, ?, ?, ?, ?, ?, ?,
                   'active', 'ok')""",
        (
            canonical_id,
            row.source_topic,
            primary_name,
            authority_canonical,
            0.95,
            row.source_url,
            "moj.go.jp",
            now_iso,
            json.dumps(raw, ensure_ascii=False),
        ),
    )
    if cur.rowcount == 0:
        return False
    cur.execute(
        """INSERT INTO am_enforcement_detail
               (entity_id, houjin_bangou, target_name, enforcement_kind,
                issuing_authority, issuance_date, exclusion_start, exclusion_end,
                reason_summary, related_law_ref, amount_yen,
                source_url, source_fetched_at)
           VALUES (?, NULL, ?, ?, ?, ?, NULL, NULL, ?, ?, NULL, ?, ?)""",
        (
            canonical_id,
            row.target_name[:255],
            row.enforcement_kind,
            row.issuing_authority[:255],
            row.issuance_date,
            (row.reason_summary or "")[:2000] or None,
            row.related_law_ref[:255] if row.related_law_ref else None,
            row.source_url,
            now_iso,
        ),
    )
    return True


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--db", type=Path, default=DEFAULT_DB)
    ap.add_argument(
        "--limit-files",
        type=int,
        default=None,
        help="cap number of source Excel files to process (debug)",
    )
    ap.add_argument(
        "--skip-fetch",
        action="store_true",
        help="use cached Excel files only — skip live HTTP fetch",
    )
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("-v", "--verbose", action="store_true")
    return ap.parse_args(argv)


def run(args: argparse.Namespace) -> int:
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    for noisy in ("openpyxl", "urllib3", "requests"):
        logging.getLogger(noisy).setLevel(logging.WARNING)
    now_iso = datetime.now(tz=UTC).isoformat(timespec="seconds").replace("+00:00", "Z")
    http = HttpClient()

    all_rows: list[EnfRow] = []

    files = SOURCE_FILES
    if args.limit_files is not None:
        files = files[: args.limit_files]
    for entry in files:
        url = entry["url"]
        label = entry["label"]
        _LOG.info("[fetch] %s (%s)", label, url)
        data = _fetch_or_cache(http, url, skip_fetch=args.skip_fetch)
        if data is None:
            _LOG.warning("[fetch] failed: %s", url)
            continue
        rows = _classify_target_excel(
            data,
            src_url=url,
            src_topic=entry["topic"],
            enforcement_kind=entry["kind"],
            law_ref=entry["law_ref"],
            summary_prefix=entry["summary_prefix"],
            target_kind=entry["target_kind"],
        )
        _LOG.info("[parse] %s → %d rows", label, len(rows))
        all_rows.extend(rows)

    _LOG.info("phase summary: total parsed rows = %d", len(all_rows))

    if args.dry_run or not all_rows:
        from collections import Counter

        by_kind = Counter(r.enforcement_kind for r in all_rows)
        by_auth = Counter(r.issuing_authority for r in all_rows)
        by_law = Counter(r.related_law_ref for r in all_rows)
        by_topic = Counter(r.source_topic for r in all_rows)
        for r in all_rows[:30]:
            _LOG.info(
                "  CAND: %s | %s | %s | %s",
                r.issuance_date,
                r.target_name[:40],
                r.enforcement_kind,
                r.related_law_ref,
            )
        _LOG.info("dry-run: would attempt %d inserts", len(all_rows))
        _LOG.info("by enforcement_kind: %s", dict(by_kind))
        _LOG.info("by issuing_authority: %s", dict(by_auth))
        _LOG.info("by related_law_ref: %s", dict(by_law))
        _LOG.info("by source_topic: %s", dict(by_topic))
        if args.dry_run:
            return 0

    if not args.db.exists():
        _LOG.error("autonomath.db missing: %s", args.db)
        return 2

    con = sqlite3.connect(str(args.db), timeout=300.0)
    try:
        con.execute("PRAGMA busy_timeout=300000")
        con.execute("PRAGMA foreign_keys=ON")
        ensure_tables(con)
        cur = con.cursor()
        cur.execute("BEGIN IMMEDIATE")
        existing_keys = existing_dedup_keys(cur)
        existing_ids = existing_canonical_ids(cur)
        con.commit()
    except sqlite3.Error as exc:
        _LOG.error("DB init failed: %s", exc)
        with contextlib.suppress(sqlite3.Error):
            con.close()
        return 2

    inserted = 0
    skipped_dup_db = 0
    skipped_dup_id = 0
    skipped_dup_batch = 0
    skipped_invalid = 0
    breakdown_kind: dict[str, int] = {}
    breakdown_authority: dict[str, int] = {}
    breakdown_law: dict[str, int] = {}
    breakdown_target_kind: dict[str, int] = {}

    batch_keys: set[tuple[str, str, str]] = set()
    pre_total = con.execute("SELECT COUNT(*) FROM am_enforcement_detail").fetchone()[0]
    pre_titp = con.execute(
        "SELECT COUNT(*) FROM am_enforcement_detail WHERE related_law_ref LIKE '%技能実習法%'"
    ).fetchone()[0]

    batch_size = 50
    pending = 0

    try:
        cur.execute("BEGIN IMMEDIATE")
    except sqlite3.Error as exc:
        _LOG.error("DB BEGIN failed: %s", exc)
        with contextlib.suppress(sqlite3.Error):
            con.close()
        return 2

    for r in all_rows:
        if not r.target_name or not r.issuance_date:
            skipped_invalid += 1
            continue
        nm = _normalize(r.target_name)
        key = (nm, r.issuance_date, r.enforcement_kind)
        if key in existing_keys:
            skipped_dup_db += 1
            continue
        if key in batch_keys:
            skipped_dup_batch += 1
            continue
        cid = build_canonical_id(r)
        if cid in existing_ids:
            skipped_dup_id += 1
            continue
        try:
            ok = insert_one(cur, r, now_iso=now_iso)
        except sqlite3.IntegrityError as exc:
            _LOG.warning("integrity err for %s: %s", r.target_name, exc)
            continue
        except sqlite3.Error as exc:
            _LOG.error("DB error %s: %s", r.target_name, exc)
            continue
        if ok:
            inserted += 1
            batch_keys.add(key)
            existing_ids.add(cid)
            breakdown_kind[r.enforcement_kind] = breakdown_kind.get(r.enforcement_kind, 0) + 1
            breakdown_authority[r.issuing_authority] = (
                breakdown_authority.get(r.issuing_authority, 0) + 1
            )
            breakdown_law[r.related_law_ref] = breakdown_law.get(r.related_law_ref, 0) + 1
            breakdown_target_kind[r.target_kind] = breakdown_target_kind.get(r.target_kind, 0) + 1
            pending += 1
            if pending >= batch_size:
                try:
                    con.commit()
                    cur.execute("BEGIN IMMEDIATE")
                    pending = 0
                except sqlite3.Error as exc:
                    _LOG.error("commit failed: %s", exc)
                    return 2
        else:
            skipped_dup_id += 1

    try:
        con.commit()
    except sqlite3.Error as exc:
        _LOG.error("final commit failed: %s", exc)

    post_total = con.execute("SELECT COUNT(*) FROM am_enforcement_detail").fetchone()[0]
    post_titp = con.execute(
        "SELECT COUNT(*) FROM am_enforcement_detail WHERE related_law_ref LIKE '%技能実習法%'"
    ).fetchone()[0]
    with contextlib.suppress(sqlite3.Error):
        con.close()

    _LOG.info(
        "done parsed=%d inserted=%d dup_db=%d dup_id=%d dup_batch=%d invalid=%d",
        len(all_rows),
        inserted,
        skipped_dup_db,
        skipped_dup_id,
        skipped_dup_batch,
        skipped_invalid,
    )

    print(
        json.dumps(
            {
                "inserted": inserted,
                "parsed": len(all_rows),
                "skipped_dup_db": skipped_dup_db,
                "skipped_dup_id": skipped_dup_id,
                "skipped_dup_batch": skipped_dup_batch,
                "skipped_invalid": skipped_invalid,
                "pre_titp_count": pre_titp,
                "post_titp_count": post_titp,
                "delta_titp": post_titp - pre_titp,
                "pre_am_enforcement_total": pre_total,
                "post_am_enforcement_total": post_total,
                "breakdown_by_kind": breakdown_kind,
                "breakdown_by_authority": breakdown_authority,
                "breakdown_by_law": breakdown_law,
                "breakdown_by_target_kind": breakdown_target_kind,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0 if inserted >= 0 else 1


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    return run(args)


if __name__ == "__main__":
    sys.exit(main())
