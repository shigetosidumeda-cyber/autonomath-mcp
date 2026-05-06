#!/usr/bin/env python3
"""Ingest MLIT 行政処分 (建設業 / 宅建業 / 指名停止 / 旅行業) into
``am_entities`` + ``am_enforcement_detail``.

Sources (primary only — aggregators BANNED):
    * https://www.mlit.go.jp/nega-inf/cgi-bin/search.cgi   (POST + GET, EID=search)
    * https://www.mlit.go.jp/kankocho/content/001913487.xlsx  (観光庁 旅行業)

Recon: analysis_wave18/data_collection_log/p2_recon_mlit.md

Per recon:
  * EID=search 必須 — 空だと form 再表示 (silent 0 result)
  * 10 件/page, page= GET param
  * <h3 class="title">検索結果：N件</h3> で総件数
  * 一覧 row: <td class="name">商号<span>（法人番号13桁）</span></td>
                <td class="address">...</td>
                <td class="date">YYYY年MM月DD日</td>
                <td class="name">処分を行った者</td>
                <td class="punish">許可取消/営業停止/指示/勧告</td>
                <td class="detail"><a href="search.cgi?...&no=NNNN">詳細</a></td>
  * 詳細 page: <dl class="overview__list"> dt/dd pairs
    商号又は名称 / 代表者 / 主たる営業所の所在地 / 許可番号 /
    建設業の種類 / 処分年月日 / 処分を行った者 / 根拠法令 /
    処分の内容（詳細） / 処分の原因となった事実 / その他参考となる事項

旅行業 xlsx structure (63 rows × 12 cols, single 様式 sheet):
    header row 2 (0-indexed). cols: 処分日 / 行政庁 / 種別 /
    登録番号 / 事業者名 / 営業所名 / 営業所住所 / 処分内容 /
    期間 / 根拠法令 / 違反行為の概要
    法人番号は未掲載 (登録番号 = 観光庁長官登録第101号 形式) — NULL OK.

Schema target (autonomath.db):
    * am_entities(canonical_id = 'enforcement:mlit:<yyyy-mm-dd>:<houjin>:<kind>',
                  record_kind='enforcement', primary_name, source_url,
                  raw_json)
    * am_enforcement_detail(entity_id, houjin_bangou, target_name,
                            enforcement_kind, issuing_authority,
                            issuance_date, reason_summary, related_law_ref)

enforcement_kind mapping (MLIT 処分種別 → am_enforcement_detail CHECK enum):
    許可取消  -> license_revoke
    登録取消  -> license_revoke
    営業停止  -> business_improvement   (license_suspend が無いので近似)
    業務停止  -> business_improvement
    指示      -> business_improvement
    勧告      -> business_improvement
    指名停止  -> contract_suspend

dedup key: (houjin_bangou, issuance_date, enforcement_kind, target_name)
  — houjin_bangou が NULL (旅行業 xlsx) のときは target_name + issuance_date
    + enforcement_kind で dedup.

CLI:
    python scripts/ingest/ingest_enforcement_mlit.py \\
        --db autonomath.db \\
        [--categories kensetugyousya,takuti,shimeiteishi,travel]  # default 全部
        [--year-from 2020] [--year-to 2025]  # default: 直近 5 年
        [--limit N]  # per-category
        [--dry-run]
        [--skip-detail]  # 一覧のみ (速度優先、違反事実 prose 無し)

Exit codes:
    0 success
    1 network / parse failure
    2 DB lock / missing schema
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import logging
import re
import sqlite3
import sys
import time
import urllib.parse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    import httpx
except ImportError as e:  # pragma: no cover
    sys.exit(f"httpx required: {e}")

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_LOG = logging.getLogger("autonomath.ingest.enforcement_mlit")

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_DB = REPO_ROOT / "autonomath.db"

MLIT_SEARCH_URL = "https://www.mlit.go.jp/nega-inf/cgi-bin/search.cgi"
KANKOCHO_TRAVEL_XLSX = "https://www.mlit.go.jp/kankocho/content/001913487.xlsx"

USER_AGENT = "jpintel-mcp-ingest/1.0 (+https://jpcite.com; contact=ops@jpcite.com)"

PER_REQUEST_DELAY_SEC = 1.0
HTTP_TIMEOUT_SEC = 30.0
MAX_RETRIES = 3

CATEGORIES_ALL = ("kensetugyousya", "takuti", "shimeiteishi", "travel")
CATEGORY_LABEL_JA = {
    "kensetugyousya": "建設業",
    "takuti": "宅地建物取引業",
    "shimeiteishi": "国交省指名停止",
    "travel": "旅行業",
}

# MLIT 処分内容 → am_enforcement_detail.enforcement_kind enum
PUNISH_KIND_MAP: dict[str, str] = {
    "許可取消": "license_revoke",
    "登録取消": "license_revoke",
    "登録の取消し": "license_revoke",
    "営業停止": "business_improvement",
    "業務停止": "business_improvement",
    "指示": "business_improvement",
    "勧告": "business_improvement",
    "指名停止": "contract_suspend",
    "競争参加資格停止": "contract_suspend",
    "指名の停止": "contract_suspend",
}

ISSUING_AUTHORITY_DEFAULT = {
    "kensetugyousya": "国土交通省",
    "takuti": "国土交通省",
    "shimeiteishi": "国土交通省",
    "travel": "観光庁",
}

# Regex for 一覧 table parsing (fast; BeautifulSoup optional)
ROW_RE = re.compile(
    r'<tr>\s*<td class="name">([^<]+?)(?:<span>\(([^)]*)\)</span>)?</td>'
    r'\s*<td class="address">([^<]*)</td>'
    r'\s*<td class="date">([^<]+)</td>'
    r'\s*<td class="name">([^<]*)</td>'
    r'\s*<td class="punish">([^<]+)</td>'
    r'\s*<td class="detail"><a[^>]+href="([^"]+)"[^>]*>詳細</a></td>\s*</tr>',
    re.DOTALL,
)
# The <span>（法人番号）</span> uses 全角 parens; match those too
ROW_RE_FW = re.compile(
    r'<tr>\s*<td class="name">([^<]+?)(?:<span>（([^）]*)）</span>)?</td>'
    r'\s*<td class="address">([^<]*)</td>'
    r'\s*<td class="date">([^<]+)</td>'
    r'\s*<td class="name">([^<]*)</td>'
    r'\s*<td class="punish">([^<]+)</td>'
    r'\s*<td class="detail"><a[^>]+href="([^"]+)"[^>]*>詳細</a></td>\s*</tr>',
    re.DOTALL,
)
RESULT_COUNT_RE = re.compile(r'<h3 class="title">検索結果：\s*(\d+)\s*件</h3>')
HOUJIN_13_RE = re.compile(r"\d{13}")

# Detail page dt/dd parser
DL_RE = re.compile(
    r'<dt class="title">([^<]+)</dt><dd class="text">([^<]*)</dd>',
    re.DOTALL,
)

DATE_YEAR_RE = re.compile(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日")


# ---------------------------------------------------------------------------
# HTTP client with rate-limit
# ---------------------------------------------------------------------------


class MlitHttpClient:
    """Rate-limited httpx client for MLIT endpoints."""

    def __init__(self) -> None:
        self._client = httpx.Client(
            headers={
                "User-Agent": USER_AGENT,
                "Accept-Language": "ja,en;q=0.5",
            },
            timeout=HTTP_TIMEOUT_SEC,
            follow_redirects=True,
        )
        self._last_fetch: float = 0.0

    def _pace(self) -> None:
        now = time.monotonic()
        wait = PER_REQUEST_DELAY_SEC - (now - self._last_fetch)
        if wait > 0:
            time.sleep(wait)
        self._last_fetch = time.monotonic()

    def get(self, url: str) -> tuple[int, str]:
        """GET url; return (status, text)."""
        last_exc: Exception | None = None
        for attempt in range(1, MAX_RETRIES + 1):
            self._pace()
            try:
                r = self._client.get(url)
                return r.status_code, r.text
            except httpx.HTTPError as exc:
                last_exc = exc
                if attempt == MAX_RETRIES:
                    break
                time.sleep(2**attempt)
        _LOG.warning("GET failed url=%s err=%s", url, last_exc)
        return 0, ""

    def post(self, url: str, data: dict[str, str]) -> tuple[int, str]:
        """POST form-encoded; return (status, text)."""
        last_exc: Exception | None = None
        for attempt in range(1, MAX_RETRIES + 1):
            self._pace()
            try:
                r = self._client.post(url, data=data)
                return r.status_code, r.text
            except httpx.HTTPError as exc:
                last_exc = exc
                if attempt == MAX_RETRIES:
                    break
                time.sleep(2**attempt)
        _LOG.warning("POST failed url=%s err=%s", url, last_exc)
        return 0, ""

    def get_bytes(self, url: str) -> tuple[int, bytes]:
        """GET binary (for xlsx)."""
        last_exc: Exception | None = None
        for attempt in range(1, MAX_RETRIES + 1):
            self._pace()
            try:
                r = self._client.get(url)
                return r.status_code, r.content
            except httpx.HTTPError as exc:
                last_exc = exc
                if attempt == MAX_RETRIES:
                    break
                time.sleep(2**attempt)
        _LOG.warning("GET-bytes failed url=%s err=%s", url, last_exc)
        return 0, b""

    def close(self) -> None:
        self._client.close()


# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------


@dataclass
class EnforcementRow:
    """One normalized MLIT enforcement event (to be written to
    am_entities + am_enforcement_detail)."""

    category: str  # kensetugyousya / takuti / ...
    target_name: str
    houjin_bangou: str | None  # 13-digit
    address: str | None
    issuance_date: str  # ISO yyyy-mm-dd
    issuing_authority: str  # 大阪府 / 国交省 / 観光庁 ...
    punishment_raw: str  # 許可取消 / 営業停止 / ...
    enforcement_kind: str  # license_revoke / business_improvement / contract_suspend
    source_url: str  # canonical permalink (detail page for mlit, xlsx for travel)
    # From detail page (may be None if skip-detail)
    representative: str | None = None
    license_no: str | None = None
    license_types: str | None = None
    detail_content: str | None = None  # 処分の内容（詳細）
    reason_summary: str | None = None  # 処分の原因となった事実
    related_law_ref: str | None = None
    period_raw: str | None = None
    other_notes: str | None = None
    # For xlsx (no detail fetch)
    registration_no: str | None = None


def make_canonical_id(row: EnforcementRow) -> str:
    """canonical_id = enforcement:mlit:<yyyy-mm-dd>:<houjin|hash>:<kind>.

    If houjin_bangou is missing (旅行業 xlsx), fall back to short sha1 of
    target_name + registration_no.
    """
    key = (
        row.houjin_bangou
        or hashlib.sha1(f"{row.target_name}|{row.registration_no or ''}".encode()).hexdigest()[:12]
    )
    # enforcement_kind may repeat on same day for same houjin (e.g.
    # 許可取消 + 指示 issued together); include raw punishment too.
    punish_slug = hashlib.sha1(row.punishment_raw.encode()).hexdigest()[:6]
    return f"enforcement:mlit:{row.issuance_date}:{key}:{row.enforcement_kind}:{punish_slug}"


# ---------------------------------------------------------------------------
# Date normalization
# ---------------------------------------------------------------------------


def ja_date_to_iso(s: str) -> str | None:
    """'2024年12月25日' → '2024-12-25'."""
    m = DATE_YEAR_RE.search(s)
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return dt.date(y, mo, d).isoformat()
    except ValueError:
        return None


def excel_serial_to_iso(v: Any) -> str | None:
    """openpyxl returns datetime for most cells but some sheets emit the
    raw serial int. Handle both. Also accept 'R7.7.2' Japanese era strings."""
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, (int, float)):
        # Excel 1900-based serial
        try:
            base = dt.date(1899, 12, 30)
            return (base + dt.timedelta(days=int(v))).isoformat()
        except Exception:
            return None
    if isinstance(v, str):
        iso = ja_date_to_iso(v)
        if iso:
            return iso
        # Reiwa: R7.3.27 -> 2025-03-27 (R1=2019)
        m = re.match(r"R\s*(\d+)\.(\d+)\.(\d+)", v)
        if m:
            y = 2018 + int(m.group(1))
            try:
                return dt.date(y, int(m.group(2)), int(m.group(3))).isoformat()
            except ValueError:
                return None
        # Heisei: H30.3.27 -> 2018-03-27
        m = re.match(r"H\s*(\d+)\.(\d+)\.(\d+)", v)
        if m:
            y = 1988 + int(m.group(1))
            try:
                return dt.date(y, int(m.group(2)), int(m.group(3))).isoformat()
            except ValueError:
                return None
        # Plain YYYY-MM-DD
        m = re.match(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", v)
        if m:
            try:
                return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
            except ValueError:
                return None
    return None


# ---------------------------------------------------------------------------
# MLIT search walker (建設業 / 宅建業 / 指名停止)
# ---------------------------------------------------------------------------


def map_enforcement_kind(punish: str, category: str) -> str:
    """Normalize 処分内容 → enforcement_kind enum."""
    p = punish.strip()
    # Prefer direct match; fallback to substring.
    if p in PUNISH_KIND_MAP:
        return PUNISH_KIND_MAP[p]
    for kw, kind in PUNISH_KIND_MAP.items():
        if kw in p:
            return kind
    # Unknown 処分 for shimeiteishi → contract_suspend by category convention
    if category == "shimeiteishi":
        return "contract_suspend"
    return "business_improvement"


def parse_list_page(html: str) -> tuple[int | None, list[dict[str, str]]]:
    """Return (total_count, rows).

    rows have raw fields; further normalization happens downstream.
    """
    total: int | None = None
    m = RESULT_COUNT_RE.search(html)
    if m:
        total = int(m.group(1))

    rows: list[dict[str, str]] = []
    # Try full-width paren first (actual format), then half-width fallback.
    matches = list(ROW_RE_FW.finditer(html)) or list(ROW_RE.finditer(html))
    for rm in matches:
        target_name_raw = rm.group(1).strip()
        houjin_raw = (rm.group(2) or "").strip()
        houjin_digits = "".join(ch for ch in houjin_raw if ch.isdigit())
        if len(houjin_digits) == 13:
            houjin = houjin_digits
        else:
            houjin = ""
        address = rm.group(3).strip()
        date_ja = rm.group(4).strip()
        issuer = rm.group(5).strip()
        punish = rm.group(6).strip()
        detail_href = rm.group(7).strip()
        rows.append(
            {
                "target_name": target_name_raw,
                "houjin_bangou": houjin,
                "address": address,
                "date_ja": date_ja,
                "issuer": issuer,
                "punish": punish,
                "detail_href": detail_href,
            }
        )
    return total, rows


def walk_category(
    http: MlitHttpClient,
    category: str,
    year_from: int,
    year_to: int,
    limit: int | None,
) -> list[dict[str, str]]:
    """Walk all pages of a single category within a year range.

    Uses POST for page 1 then follows GET pagination links (page=N).
    """
    all_rows: list[dict[str, str]] = []
    # Query each year separately to keep result count below 1000-entry
    # pagination walls (MLIT CGI behaves but this is defensive).
    for year in range(year_from, year_to + 1):
        form = {
            "jigyoubunya": category,
            "EID": "search",  # critical — silent 0 without this
            "start_year": str(year),
            "start_month": "1",
            "end_year": str(year),
            "end_month": "12",
            "disposal_name1": "",
            "disposal_name2": "",
            "reason_con": "",
            "reason1": "",
            "reason2": "",
            "reason3": "",
            "shobun": "",
            "address": "",
            "agency": "",
        }
        status, html = http.post(MLIT_SEARCH_URL, form)
        if status != 200 or not html:
            _LOG.warning(
                "cat=%s year=%d POST failed status=%s len=%d",
                category,
                year,
                status,
                len(html),
            )
            continue
        total, rows = parse_list_page(html)
        if total is None:
            _LOG.warning(
                "cat=%s year=%d could not parse result count; skipping",
                category,
                year,
            )
            continue
        _LOG.info(
            "cat=%s year=%d total=%d page1_rows=%d",
            category,
            year,
            total,
            len(rows),
        )
        all_rows.extend(rows)
        if total <= 10:
            continue

        # Pagination: 10 per page.
        pages = (total + 9) // 10
        for page in range(2, pages + 1):
            if limit is not None and len(all_rows) >= limit:
                break
            qs = {
                "jigyoubunya": category,
                "EID": "search",
                "start_year": str(year),
                "start_month": "1",
                "end_year": str(year),
                "end_month": "12",
                "disposal_name1": "",
                "disposal_name2": "",
                "reason_con": "",
                "reason1": "",
                "reason2": "",
                "reason3": "",
                "shobun": "",
                "address": "",
                "agency": "",
                "page": str(page),
            }
            page_url = MLIT_SEARCH_URL + "?" + urllib.parse.urlencode(qs)
            status, html = http.get(page_url)
            if status != 200:
                _LOG.warning(
                    "cat=%s year=%d page=%d GET failed status=%s",
                    category,
                    year,
                    page,
                    status,
                )
                continue
            _, prows = parse_list_page(html)
            all_rows.extend(prows)
            if not prows:
                _LOG.warning(
                    "cat=%s year=%d page=%d zero rows — stopping pagination",
                    category,
                    year,
                    page,
                )
                break
        if limit is not None and len(all_rows) >= limit:
            break

    return all_rows


def fetch_detail(
    http: MlitHttpClient,
    category: str,
    detail_href: str,
) -> dict[str, str]:
    """Fetch + parse detail page. Return dt/dd dict."""
    # detail_href is relative: search.cgi?jigyoubunya=xxx&EID=search&no=NNNN
    if detail_href.startswith("http"):
        url = detail_href
    else:
        url = "https://www.mlit.go.jp/nega-inf/cgi-bin/" + detail_href.lstrip("/")
    status, html = http.get(url)
    if status != 200:
        return {}
    out: dict[str, str] = {}
    for m in DL_RE.finditer(html):
        k = m.group(1).strip()
        v = m.group(2).strip()
        out[k] = v
    return out


# ---------------------------------------------------------------------------
# Travel 観光庁 xlsx
# ---------------------------------------------------------------------------


def parse_travel_xlsx(xlsx_bytes: bytes) -> list[EnforcementRow]:
    """Parse 観光庁 旅行業 xlsx into EnforcementRow list."""
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed; pip install openpyxl")
    import io

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[EnforcementRow] = []
    # Header at row 3 (1-indexed). Data from row 4 onward.
    for raw in ws.iter_rows(min_row=4, values_only=True):
        # col layout (1-indexed, 0 of tuple is None pad):
        # 0=pad, 1=処分日, 2=行政庁, 3=種別, 4=登録番号, 5=事業者名,
        # 6=営業所名, 7=営業所住所, 8=処分内容, 9=期間, 10=根拠法令, 11=違反行為の概要
        if not raw or len(raw) < 12:
            continue
        issuance = excel_serial_to_iso(raw[1])
        issuer = str(raw[2]).strip() if raw[2] else None
        # tier = raw[3]  # 1種/2種/地域/... ignored for now
        reg_no = str(raw[4]).replace("\n", "").strip() if raw[4] else None
        target_name = str(raw[5]).strip() if raw[5] else None
        # office_name = raw[6]  # 営業所名 (not stored directly)
        address = str(raw[7]).strip() if raw[7] else None
        punishment_raw = str(raw[8]).strip() if raw[8] else ""
        period = str(raw[9]).strip() if raw[9] else None
        law = str(raw[10]).strip() if raw[10] else None
        violation = str(raw[11]).strip() if raw[11] else None

        if not issuance or not target_name or not punishment_raw:
            continue
        kind = map_enforcement_kind(punishment_raw, "travel")
        rows.append(
            EnforcementRow(
                category="travel",
                target_name=target_name,
                houjin_bangou=None,  # 観光庁 xlsx に法人番号なし
                address=address,
                issuance_date=issuance,
                issuing_authority=issuer or "観光庁",
                punishment_raw=punishment_raw,
                enforcement_kind=kind,
                source_url=KANKOCHO_TRAVEL_XLSX,
                registration_no=reg_no,
                period_raw=period,
                related_law_ref=law,
                reason_summary=violation,
            )
        )
    return rows


# ---------------------------------------------------------------------------
# DB layer
# ---------------------------------------------------------------------------


def open_db(db_path: Path) -> sqlite3.Connection:
    """Open autonomath.db with large busy_timeout + immediate transaction."""
    if not db_path.exists():
        raise SystemExit(f"DB not found: {db_path}")
    conn = sqlite3.connect(str(db_path), timeout=300.0)
    conn.execute("PRAGMA busy_timeout = 300000")
    conn.execute("PRAGMA journal_mode = WAL")
    # Verify tables exist.
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='am_enforcement_detail'"
    ).fetchone()
    if not row:
        conn.close()
        raise SystemExit("am_enforcement_detail table missing")
    return conn


def load_dedup_set(conn: sqlite3.Connection) -> set[tuple[str, str, str, str]]:
    """Preload existing (houjin_bangou, issuance_date, enforcement_kind, target_name)
    tuples to avoid re-inserting."""
    dedup: set[tuple[str, str, str, str]] = set()
    for row in conn.execute(
        "SELECT IFNULL(houjin_bangou, ''), issuance_date, IFNULL(enforcement_kind, ''), IFNULL(target_name, '') FROM am_enforcement_detail"
    ):
        dedup.add((row[0], row[1], row[2], row[3]))
    return dedup


def upsert_enforcement(
    conn: sqlite3.Connection,
    row: EnforcementRow,
    fetched_at: str,
) -> str:
    """Insert am_entities + am_enforcement_detail. Returns 'insert'|'skip'|'update'.

    Uses INSERT OR IGNORE on am_entities (canonical_id PK) and matches by the
    same canonical_id for am_enforcement_detail. Duplicate canonical_id means
    we already have the row — 'skip'.
    """
    canonical_id = make_canonical_id(row)

    raw_json = {
        "category": row.category,
        "category_label": CATEGORY_LABEL_JA.get(row.category, row.category),
        "target_name": row.target_name,
        "houjin_bangou": row.houjin_bangou,
        "address": row.address,
        "issuance_date": row.issuance_date,
        "issuing_authority": row.issuing_authority,
        "punishment_raw": row.punishment_raw,
        "enforcement_kind": row.enforcement_kind,
        "representative": row.representative,
        "license_no": row.license_no,
        "license_types": row.license_types,
        "detail_content": row.detail_content,
        "reason_summary": row.reason_summary,
        "related_law_ref": row.related_law_ref,
        "period_raw": row.period_raw,
        "other_notes": row.other_notes,
        "registration_no": row.registration_no,
        "fetched_at": fetched_at,
        "source": "mlit_nega_inf" if row.category != "travel" else "kankocho_xlsx",
    }

    source_url_domain = urllib.parse.urlparse(row.source_url).netloc

    # am_entities upsert (INSERT OR IGNORE — canonical_id is stable dedup).
    cur = conn.execute(
        """INSERT OR IGNORE INTO am_entities (
            canonical_id, record_kind, source_topic, primary_name,
            confidence, source_url, source_url_domain, fetched_at, raw_json
        ) VALUES (?, 'enforcement', ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            canonical_id,
            f"mlit_{row.category}",
            row.target_name,
            0.95,
            row.source_url,
            source_url_domain,
            fetched_at,
            json.dumps(raw_json, ensure_ascii=False, separators=(",", ":")),
        ),
    )
    entity_inserted = cur.rowcount > 0

    # am_enforcement_detail: see if the exact same canonical_id + kind
    # already has a row; skip to stay idempotent.
    existing = conn.execute(
        "SELECT enforcement_id FROM am_enforcement_detail WHERE entity_id = ?",
        (canonical_id,),
    ).fetchone()

    if existing:
        return "skip"

    conn.execute(
        """INSERT INTO am_enforcement_detail (
            entity_id, houjin_bangou, target_name, enforcement_kind,
            issuing_authority, issuance_date, reason_summary,
            related_law_ref, source_url, source_fetched_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            canonical_id,
            row.houjin_bangou,
            row.target_name,
            row.enforcement_kind,
            row.issuing_authority,
            row.issuance_date,
            row.reason_summary,
            row.related_law_ref,
            row.source_url,
            fetched_at,
        ),
    )
    return "insert" if entity_inserted else "update"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def normalize_mlit_row(
    raw: dict[str, str],
    category: str,
    detail: dict[str, str] | None,
) -> EnforcementRow | None:
    """Convert raw 一覧 row (+ optional detail) to EnforcementRow."""
    issuance = ja_date_to_iso(raw["date_ja"])
    if not issuance:
        _LOG.warning(
            "cat=%s cannot parse date=%r target=%r",
            category,
            raw["date_ja"],
            raw["target_name"],
        )
        return None
    punishment = raw["punish"].strip()
    kind = map_enforcement_kind(punishment, category)
    issuer = raw["issuer"].strip() or ISSUING_AUTHORITY_DEFAULT.get(category, "国土交通省")

    # For 指名停止 the 処分を行った者 IS the 整備局 / 地整 — keep that as authority.
    # For 建設業 / 宅建業 the 処分を行った者 = 都道府県知事 or 国交大臣 (国交省表記)

    target_clean = raw["target_name"].strip()
    houjin = raw.get("houjin_bangou") or ""
    if houjin and not HOUJIN_13_RE.fullmatch(houjin):
        # half-width digits only; shouldn't happen but be strict
        houjin = ""

    # Build source_url = the detail permalink (stable within掲載期限)
    detail_href = raw["detail_href"].strip()
    if detail_href.startswith("http"):
        source_url = detail_href
    else:
        source_url = "https://www.mlit.go.jp/nega-inf/cgi-bin/" + detail_href.lstrip("/")

    row = EnforcementRow(
        category=category,
        target_name=target_clean,
        houjin_bangou=houjin or None,
        address=raw.get("address") or None,
        issuance_date=issuance,
        issuing_authority=issuer,
        punishment_raw=punishment,
        enforcement_kind=kind,
        source_url=source_url,
    )

    if detail:
        row.representative = detail.get("代表者") or None
        row.license_no = detail.get("許可番号") or None
        row.license_types = detail.get("許可を受けている建設業の種類") or None
        row.detail_content = detail.get("処分の内容（詳細）") or None
        row.reason_summary = detail.get("処分の原因となった事実") or None
        row.related_law_ref = detail.get("根拠法令") or None
        row.other_notes = detail.get("その他参考となる事項") or None
    return row


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--db", type=Path, default=DEFAULT_DB, help=f"SQLite DB (default {DEFAULT_DB})")
    ap.add_argument(
        "--categories",
        type=str,
        default=",".join(CATEGORIES_ALL),
        help="comma-separated: kensetugyousya,takuti,shimeiteishi,travel",
    )
    ap.add_argument(
        "--year-from",
        type=int,
        default=dt.date.today().year - 4,
        help="start year (inclusive); default = today-4",
    )
    ap.add_argument(
        "--year-to",
        type=int,
        default=dt.date.today().year,
        help="end year (inclusive); default = today",
    )
    ap.add_argument(
        "--limit", type=int, default=None, help="per-category row cap (for smoke tests)"
    )
    ap.add_argument("--dry-run", action="store_true", help="parse only; no DB writes")
    ap.add_argument(
        "--skip-detail",
        action="store_true",
        help="one-shot一覧だけ取得, 違反事実 prose 無し (faster)",
    )
    ap.add_argument(
        "--log-file", type=Path, default=None, help="append progress summary to this file"
    )
    ap.add_argument("--verbose", "-v", action="store_true")
    return ap.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    categories = [c.strip() for c in args.categories.split(",") if c.strip()]
    unknown = [c for c in categories if c not in CATEGORIES_ALL]
    if unknown:
        _LOG.error("unknown categories: %s (allowed: %s)", unknown, CATEGORIES_ALL)
        return 2

    fetched_at = dt.datetime.now(dt.UTC).isoformat(timespec="seconds").replace("+00:00", "Z")

    http = MlitHttpClient()
    conn: sqlite3.Connection | None = None
    if not args.dry_run:
        conn = open_db(args.db)
        # BEGIN IMMEDIATE lets us hold write-lock promptly; otherwise WAL
        # readers can keep us waiting.
        conn.execute("BEGIN IMMEDIATE")

    stats: dict[str, dict[str, int]] = {}
    try:
        for cat in categories:
            cat_stats = {
                "walked": 0,
                "detail_fetched": 0,
                "built": 0,
                "insert": 0,
                "skip": 0,
                "update": 0,
                "parse_skip": 0,
            }
            stats[cat] = cat_stats

            if cat == "travel":
                _LOG.info("fetching 観光庁 旅行業 xlsx...")
                status, body = http.get_bytes(KANKOCHO_TRAVEL_XLSX)
                if status != 200 or not body:
                    _LOG.error("travel xlsx fetch failed status=%s", status)
                    continue
                try:
                    trows = parse_travel_xlsx(body)
                except Exception as exc:  # noqa: BLE001
                    _LOG.exception("travel xlsx parse failed: %s", exc)
                    continue
                # Filter year range
                y0, y1 = args.year_from, args.year_to
                trows = [
                    r for r in trows if r.issuance_date and y0 <= int(r.issuance_date[:4]) <= y1
                ]
                cat_stats["walked"] = len(trows)
                cat_stats["built"] = len(trows)
                if args.limit:
                    trows = trows[: args.limit]
                if args.dry_run or conn is None:
                    for r in trows[:5]:
                        _LOG.info(
                            "DRY travel %s | %s | %s | %s",
                            r.issuance_date,
                            r.target_name,
                            r.punishment_raw,
                            r.enforcement_kind,
                        )
                    continue
                for r in trows:
                    try:
                        verdict = upsert_enforcement(conn, r, fetched_at)
                        cat_stats[verdict] = cat_stats.get(verdict, 0) + 1
                    except sqlite3.Error as exc:
                        _LOG.error("DB insert failed %s: %s", r.target_name, exc)
                        cat_stats["parse_skip"] += 1
                _LOG.info("travel done: %s", cat_stats)
                continue

            # MLIT search for kensetugyousya / takuti / shimeiteishi.
            _LOG.info(
                "walking cat=%s years=%d..%d",
                cat,
                args.year_from,
                args.year_to,
            )
            list_rows = walk_category(
                http,
                cat,
                args.year_from,
                args.year_to,
                args.limit,
            )
            cat_stats["walked"] = len(list_rows)
            _LOG.info("cat=%s walked=%d", cat, len(list_rows))

            for lr in list_rows:
                detail: dict[str, str] | None = None
                if not args.skip_detail:
                    detail = fetch_detail(http, cat, lr["detail_href"])
                    if detail:
                        cat_stats["detail_fetched"] += 1

                row = normalize_mlit_row(lr, cat, detail)
                if row is None:
                    cat_stats["parse_skip"] += 1
                    continue
                cat_stats["built"] += 1

                if args.dry_run or conn is None:
                    if cat_stats["built"] <= 3:
                        _LOG.info(
                            "DRY %s %s | %s | houjin=%s | %s | kind=%s",
                            cat,
                            row.issuance_date,
                            row.target_name,
                            row.houjin_bangou,
                            row.punishment_raw,
                            row.enforcement_kind,
                        )
                    continue
                try:
                    verdict = upsert_enforcement(conn, row, fetched_at)
                    cat_stats[verdict] = cat_stats.get(verdict, 0) + 1
                except sqlite3.Error as exc:
                    _LOG.error("DB insert failed %s: %s", row.target_name, exc)
                    cat_stats["parse_skip"] += 1

                # Periodic flush every 50 inserts to survive mid-run crashes.
                if (cat_stats["insert"] + cat_stats["update"]) % 50 == 0 and (
                    cat_stats["insert"] + cat_stats["update"]
                ) > 0:
                    conn.commit()
                    conn.execute("BEGIN IMMEDIATE")

            _LOG.info("cat=%s done: %s", cat, cat_stats)
    finally:
        http.close()
        if conn is not None:
            conn.commit()
            conn.close()

    # Progress summary
    total_insert = sum(s.get("insert", 0) for s in stats.values())
    total_skip = sum(s.get("skip", 0) for s in stats.values())
    total_walked = sum(s.get("walked", 0) for s in stats.values())
    _LOG.info(
        "SUMMARY walked=%d insert=%d skip=%d per_cat=%s",
        total_walked,
        total_insert,
        total_skip,
        stats,
    )

    if args.log_file is not None:
        with open(args.log_file, "a") as f:
            f.write(
                f"\n## {fetched_at} MLIT enforcement ingest\n"
                f"  years={args.year_from}..{args.year_to} "
                f"categories={categories}\n"
                f"  walked={total_walked} insert={total_insert} "
                f"skip={total_skip}\n"
                f"  per_category={json.dumps(stats, ensure_ascii=False)}\n"
            )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
