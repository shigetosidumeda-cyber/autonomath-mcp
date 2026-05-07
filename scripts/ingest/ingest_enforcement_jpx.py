#!/usr/bin/env python3
"""Ingest JPX 自主規制 / 取引所 上場会社措置 + CPAAOB 監査法人検査勧告 into
``am_enforcement_detail`` + ``am_entities``.

Coverage (2026-04-25):
  1. 上場廃止銘柄一覧
       https://www.jpx.co.jp/listing/stocks/delisted/index.html
       (~68 rows, 上場廃止日 / 銘柄名 / コード / 市場区分 / 上場廃止理由)
       authority='東京証券取引所', law='有価証券上場規程第601条',
       enforcement_kind='license_revoke'
  2. 改善報告書・改善状況報告書徴求会社一覧
       https://www.jpx.co.jp/listing/measures/improvement-reports/index.html
       (~22 rows)
       authority='日本取引所自主規制法人', law='有価証券上場規程第502条',
       enforcement_kind='business_improvement'
  3. 特別注意銘柄一覧 (現行)
       https://www.jpx.co.jp/listing/measures/alert/index.html (~9 rows)
     特別注意銘柄指定履歴
       https://www.jpx.co.jp/listing/measures/alert/01.html (~1 row)
     +Excel 履歴 2019-2023:
       https://www.jpx.co.jp/listing/measures/alert/tvdivq0000000i86-att/jp_History_2019-2023.xlsx
       authority='日本取引所自主規制法人', law='有価証券上場規程第503条',
       enforcement_kind='contract_suspend'
  4. 公表措置銘柄一覧
       https://www.jpx.co.jp/listing/measures/public-announce/index.html
       (~5 rows)
       authority='日本取引所自主規制法人', law='有価証券上場規程第508条',
       enforcement_kind='other'
  5. 上場契約違約金徴求会社一覧
       https://www.jpx.co.jp/listing/measures/listing-agreement-violation/index.html
       (~1 row)
       authority='日本取引所自主規制法人', law='有価証券上場規程第509条',
       enforcement_kind='fine'
  6. 監理銘柄・整理銘柄一覧
       https://www.jpx.co.jp/listing/market-alerts/supervision/index.html
       (~57 rows)
       authority='東京証券取引所', law='有価証券上場規程第501条',
       enforcement_kind='contract_suspend'
  7. 上場廃止基準に係る猶予期間入り銘柄等一覧
       https://www.jpx.co.jp/listing/market-alerts/grace-period/index.html
       +Excel:
       https://www.jpx.co.jp/listing/market-alerts/grace-period/tvdivq0000000kwb-att/GracePeriod_JP.xlsx
       (~134 rows)
       authority='東京証券取引所', law='有価証券上場規程第501条',
       enforcement_kind='contract_suspend'
  8. 改善期間該当銘柄等一覧
       https://www.jpx.co.jp/listing/market-alerts/improvement-period/index.html
       +Excel:
       https://www.jpx.co.jp/listing/market-alerts/improvement-period/aocfb400000029l8-att/Companies_ImprovementPeriod_JP_260422.xlsx
       (~75 rows)
       authority='東京証券取引所', law='有価証券上場規程第501条',
       enforcement_kind='contract_suspend'
  9. CPAAOB 監査法人 検査結果に基づく勧告
       https://www.fsa.go.jp/cpaaob/shinsakensa/kankoku/index.html
       (24+ rows, 全PDF)
       authority='公認会計士・監査審査会', law='公認会計士法第41条の2',
       enforcement_kind='business_improvement'

Idempotency:
  - Dedup key: (issuing_authority, issuance_date, target_name).
  - canonical_id format: AM-ENF-JPX-{slug8}-{seq04}.
  - Cross-source dedup vs FSA #21 done via target_name + date + amount.
    JPX-side authorities differ from FSA so practical collisions are rare.

Parallel-write SQLite:
  - BEGIN IMMEDIATE + PRAGMA busy_timeout=300000
  - Single bulk commit at end (PMDA / FSA pattern).

CLI:
    python scripts/ingest/ingest_enforcement_jpx.py
    python scripts/ingest/ingest_enforcement_jpx.py --dry-run
    python scripts/ingest/ingest_enforcement_jpx.py --max-rows 500 -v
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import io
import json
import logging
import re
import sqlite3
import sys
import time
import unicodedata
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

try:
    import requests  # type: ignore
    from bs4 import BeautifulSoup  # type: ignore
    from openpyxl import load_workbook  # type: ignore
except ImportError as exc:  # pragma: no cover
    print(
        f"missing dep: {exc}. pip install requests beautifulsoup4 openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

_LOG = logging.getLogger("autonomath.ingest.jpx")

DEFAULT_DB = REPO_ROOT / "autonomath.db"

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
)
BASE_JPX = "https://www.jpx.co.jp"
BASE_FSA = "https://www.fsa.go.jp"
HTTP_TIMEOUT = 60
RATE_SLEEP = 1.0  # 1 req/sec/host


# ---------------------------------------------------------------------------
# Source list
# ---------------------------------------------------------------------------


@dataclass
class JpxSource:
    name: str
    url: str
    authority: str
    law_basis: str
    enforcement_kind: str
    parser: str  # one of 'delisted', 'improvement_report', 'alert',
    # 'alert_history', 'public_announce', 'agreement_violation',
    # 'supervision', 'grace_period', 'improvement_period',
    # 'cpaaob_kankoku', 'alert_excel', 'grace_excel',
    # 'improvement_period_excel'


SOURCES: list[JpxSource] = [
    JpxSource(
        name="上場廃止銘柄",
        url=f"{BASE_JPX}/listing/stocks/delisted/index.html",
        authority="東京証券取引所",
        law_basis="東京証券取引所有価証券上場規程第601条",
        enforcement_kind="license_revoke",
        parser="delisted",
    ),
    JpxSource(
        name="改善報告書徴求",
        url=f"{BASE_JPX}/listing/measures/improvement-reports/index.html",
        authority="日本取引所自主規制法人",
        law_basis="東京証券取引所有価証券上場規程第502条",
        enforcement_kind="business_improvement",
        parser="improvement_report",
    ),
    JpxSource(
        name="特別注意銘柄(現行)",
        url=f"{BASE_JPX}/listing/measures/alert/index.html",
        authority="日本取引所自主規制法人",
        law_basis="東京証券取引所有価証券上場規程第503条",
        enforcement_kind="contract_suspend",
        parser="alert",
    ),
    JpxSource(
        name="特別注意銘柄(履歴)",
        url=f"{BASE_JPX}/listing/measures/alert/01.html",
        authority="日本取引所自主規制法人",
        law_basis="東京証券取引所有価証券上場規程第503条",
        enforcement_kind="contract_suspend",
        parser="alert_history",
    ),
    JpxSource(
        name="特別注意銘柄(2019-2023 Excel)",
        url=(f"{BASE_JPX}/listing/measures/alert/tvdivq0000000i86-att/jp_History_2019-2023.xlsx"),
        authority="日本取引所自主規制法人",
        law_basis="東京証券取引所有価証券上場規程第503条",
        enforcement_kind="contract_suspend",
        parser="alert_excel",
    ),
    JpxSource(
        name="公表措置銘柄",
        url=f"{BASE_JPX}/listing/measures/public-announce/index.html",
        authority="日本取引所自主規制法人",
        law_basis="東京証券取引所有価証券上場規程第508条",
        enforcement_kind="other",
        parser="public_announce",
    ),
    JpxSource(
        name="上場契約違約金徴求",
        url=(f"{BASE_JPX}/listing/measures/listing-agreement-violation/index.html"),
        authority="日本取引所自主規制法人",
        law_basis="東京証券取引所有価証券上場規程第509条",
        enforcement_kind="fine",
        parser="agreement_violation",
    ),
    JpxSource(
        name="監理・整理銘柄",
        url=f"{BASE_JPX}/listing/market-alerts/supervision/index.html",
        authority="東京証券取引所",
        law_basis="東京証券取引所有価証券上場規程第501条",
        enforcement_kind="contract_suspend",
        parser="supervision",
    ),
    JpxSource(
        name="猶予期間入り銘柄",
        url=f"{BASE_JPX}/listing/market-alerts/grace-period/index.html",
        authority="東京証券取引所",
        law_basis="東京証券取引所有価証券上場規程第501条",
        enforcement_kind="contract_suspend",
        parser="grace_period",
    ),
    JpxSource(
        name="猶予期間入り銘柄(履歴Excel)",
        url=(
            f"{BASE_JPX}/listing/market-alerts/grace-period/"
            "tvdivq0000000kwb-att/GracePeriod_JP.xlsx"
        ),
        authority="東京証券取引所",
        law_basis="東京証券取引所有価証券上場規程第501条",
        enforcement_kind="contract_suspend",
        parser="grace_excel",
    ),
    JpxSource(
        name="改善期間該当銘柄",
        url=f"{BASE_JPX}/listing/market-alerts/improvement-period/index.html",
        authority="東京証券取引所",
        law_basis="東京証券取引所有価証券上場規程第501条",
        enforcement_kind="contract_suspend",
        parser="improvement_period",
    ),
    JpxSource(
        name="改善期間該当銘柄(Excel)",
        url=(
            f"{BASE_JPX}/listing/market-alerts/improvement-period/"
            "aocfb400000029l8-att/Companies_ImprovementPeriod_JP_260422.xlsx"
        ),
        authority="東京証券取引所",
        law_basis="東京証券取引所有価証券上場規程第501条",
        enforcement_kind="contract_suspend",
        parser="improvement_period_excel",
    ),
    JpxSource(
        name="CPAAOB 監査法人勧告",
        url=f"{BASE_FSA}/cpaaob/shinsakensa/kankoku/index.html",
        authority="公認会計士・監査審査会",
        law_basis="公認会計士法第41条の2",
        enforcement_kind="business_improvement",
        parser="cpaaob_kankoku",
    ),
]


# ---------------------------------------------------------------------------
# Date / text helpers
# ---------------------------------------------------------------------------

WAREKI_RE = re.compile(
    r"(令和|平成|昭和)\s*(元|[0-9０-９]+)\s*年\s*" r"([0-9０-９]+)\s*月\s*([0-9０-９]+)\s*日"
)
ERA_OFFSET = {"令和": 2018, "平成": 1988, "昭和": 1925}
SLASH_DATE_RE = re.compile(r"(\d{4})/(\d{1,2})/(\d{1,2})")
HOUJIN_RE = re.compile(r"法人番号\s*[:：]?\s*([0-9]{13})")
_TR_DIGITS = str.maketrans("０１２３４５６７８９", "0123456789")


def _normalize(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", s)).strip()


def _slash_to_iso(text: str) -> str | None:
    """Match the first YYYY/MM/DD date and return ISO yyyy-mm-dd."""
    if not text:
        return None
    m = SLASH_DATE_RE.search(text)
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if 1990 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


def _wareki_to_iso(text: str) -> str | None:
    if not text:
        return None
    s = _normalize(text)
    m = WAREKI_RE.search(s)
    if not m:
        return None
    era = m.group(1)
    yr_raw = m.group(2).translate(_TR_DIGITS)
    mo = int(m.group(3).translate(_TR_DIGITS))
    d = int(m.group(4).translate(_TR_DIGITS))
    yr = 1 if yr_raw == "元" else int(yr_raw)
    year = ERA_OFFSET[era] + yr
    if 1990 <= year <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
        return f"{year:04d}-{mo:02d}-{d:02d}"
    return None


def _excel_date_to_iso(val) -> str | None:
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, str):
        # Try slash form first, then 令和
        return _slash_to_iso(val) or _wareki_to_iso(val)
    return None


def _slug8(*parts: str) -> str:
    return hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:8]


def _clean_company(text: str) -> str:
    """Trim whitespace, full-width spaces, trailing 注 marks from company cell."""
    if not text:
        return ""
    s = unicodedata.normalize("NFKC", text)
    s = re.sub(r"\s+", "", s)  # collapse all internal whitespace (incl ZWS)
    s = s.replace("　", "")
    return s[:500]


def _parse_yen(text: str) -> int | None:
    if not text:
        return None
    s = unicodedata.normalize("NFKC", text).translate(_TR_DIGITS).replace(",", "")
    m = re.search(
        r"(?:金|料|金額)?\s*(\d{1,4})\s*(?:億)?\s*(\d{1,4})?\s*(?:万)?\s*(\d{1,4})?\s*円", s
    )
    if not m:
        return None
    g = [int(x) if x else 0 for x in m.groups()]
    if "億" in s and "万" in s:
        return g[0] * 100_000_000 + g[1] * 10_000 + g[2]
    if "億" in s:
        return g[0] * 100_000_000 + (g[1] or 0)
    if "万" in s:
        return g[0] * 10_000 + (g[1] or 0)
    return g[0]


# ---------------------------------------------------------------------------
# Row dataclass
# ---------------------------------------------------------------------------


@dataclass
class EnfRow:
    target_name: str
    issuance_date: str
    issuing_authority: str
    enforcement_kind: str
    reason_summary: str
    related_law_ref: str
    source_url: str
    amount_yen: int | None = None
    exclusion_start: str | None = None
    exclusion_end: str | None = None
    extra: dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# HTTP / fetch
# ---------------------------------------------------------------------------


class Fetcher:
    """Single-shared session with simple per-host pacing."""

    def __init__(self) -> None:
        self.s = requests.Session()
        self.s.headers["User-Agent"] = USER_AGENT
        self.s.headers["Accept-Language"] = "ja,en;q=0.5"
        self.host_clock: dict[str, float] = {}

    def get_bytes(self, url: str) -> bytes:
        host = urlparse(url).netloc
        last = self.host_clock.get(host, 0.0)
        wait = RATE_SLEEP - (time.monotonic() - last)
        if wait > 0:
            time.sleep(wait)
        try:
            r = self.s.get(url, timeout=HTTP_TIMEOUT, allow_redirects=True)
            self.host_clock[host] = time.monotonic()
            r.raise_for_status()
            return r.content
        except requests.RequestException as exc:
            _LOG.warning("fetch fail url=%s err=%s", url, exc)
            return b""

    def get_text(self, url: str) -> str:
        body = self.get_bytes(url)
        if not body:
            return ""
        # JPX is utf-8; FSA cpaaob is utf-8 too. Trust apparent encoding.
        try:
            return body.decode("utf-8")
        except UnicodeDecodeError:
            return body.decode("utf-8", errors="replace")

    def close(self) -> None:
        self.s.close()


# ---------------------------------------------------------------------------
# Parser dispatch
# ---------------------------------------------------------------------------


def _table_rows(html: str) -> list:
    """Return all <table>..</table> tbody rows as BeautifulSoup tags."""
    soup = BeautifulSoup(html, "html.parser")
    out = []
    for tbl in soup.find_all("table"):
        for tr in tbl.find_all("tr"):
            out.append(tr)
    return out


def _row_cells_text(tr) -> list[str]:
    """Return list of cell texts (skip header rows)."""
    return [_normalize(td.get_text(" ", strip=True)) for td in tr.find_all(["td"])]


def _row_anchors(tr) -> list[tuple[str, str]]:
    """Return list of (href, text) from anchors in row."""
    return [
        (a.get("href", ""), _normalize(a.get_text(" ", strip=True)))
        for a in tr.find_all("a")
        if a.get("href")
    ]


# -- 上場廃止銘柄 --


def parse_delisted(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    for tr in _table_rows(html):
        cells = _row_cells_text(tr)
        # Schema: 上場廃止日 / 銘柄名 / コード / 市場区分 / 上場廃止理由
        if len(cells) < 5:
            continue
        date_iso = _slash_to_iso(cells[0])
        name = _clean_company(cells[1])
        code = re.sub(r"\D", "", cells[2])[:6]
        market = cells[3]
        reason = cells[4]
        if not date_iso or not name:
            continue
        summary = (f"{name} (証券コード {code}, {market}) 上場廃止 理由: {reason}")[:1500]
        rows.append(
            EnfRow(
                target_name=name,
                issuance_date=date_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=src.url,
                extra={
                    "stock_code": code,
                    "market": market,
                    "delisting_reason": reason,
                    "source_section": src.name,
                },
            )
        )
    return rows


# -- 改善報告書 --


def parse_improvement_report(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    for tr in _table_rows(html):
        cells = _row_cells_text(tr)
        # 銘柄名 / コード / 市場区分 / 縦覧期間 / 提出理由 / 報告書提出日 / 報告書 / 備考
        # Reason is a PDF link (no text); 提出日 is "改善報告書 (yyyy/mm/dd)".
        if len(cells) < 6:
            continue
        name = _clean_company(cells[0])
        code = re.sub(r"\D", "", cells[1])[:6]
        market = cells[2]
        period = cells[3]
        # 提出日 cell (index 5 typically): "改善報告書 (2025/11/11)" or
        # "改善状況報告書 (2026/04/20)"
        issuance_iso = None
        issuance_kind = ""
        for cell in cells[5:6]:
            issuance_iso = _slash_to_iso(cell)
            if "改善状況" in cell:
                issuance_kind = "改善状況報告書"
            elif "改善報告書" in cell:
                issuance_kind = "改善報告書"
        # 縦覧期間 to populate exclusion_start / exclusion_end
        period_dates = SLASH_DATE_RE.findall(period or "")
        excl_start = None
        excl_end = None
        if len(period_dates) >= 2:
            y1, m1, d1 = (int(x) for x in period_dates[0])
            y2, m2, d2 = (int(x) for x in period_dates[1])
            excl_start = f"{y1:04d}-{m1:02d}-{d1:02d}"
            excl_end = f"{y2:04d}-{m2:02d}-{d2:02d}"
        # If 提出日 missing fall back to 縦覧期間 start
        if not issuance_iso:
            issuance_iso = excl_start
        if not issuance_iso or not name:
            continue
        summary = (
            f"{name} (証券コード {code}, {market}) "
            f"改善報告書徴求 縦覧期間: {period} 提出: {issuance_kind}"
        )[:1500]
        rows.append(
            EnfRow(
                target_name=name,
                issuance_date=issuance_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=src.url,
                exclusion_start=excl_start,
                exclusion_end=excl_end,
                extra={
                    "stock_code": code,
                    "market": market,
                    "kanran_period": period,
                    "report_kind": issuance_kind,
                    "source_section": src.name,
                },
            )
        )
    return rows


# -- 特別注意銘柄(現行) --


def parse_alert(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    seen_keys: set[tuple[str, str]] = set()
    for tr in _table_rows(html):
        cells = _row_cells_text(tr)
        # 銘柄名 / コード / 市場区分 / 指定日 (or 指定継続日) / 詳細 / 審査状況 / 備考
        if len(cells) < 4:
            continue
        # Header rows are skipped because <tr> with only <th>; cells empty.
        name = _clean_company(cells[0])
        if not name:
            continue
        code = re.sub(r"\D", "", cells[1])[:6]
        market = cells[2]
        date_iso = _slash_to_iso(cells[3])
        if not date_iso:
            continue
        status = cells[5] if len(cells) > 5 else ""
        note = cells[6] if len(cells) > 6 else ""
        key = (name, date_iso)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        summary = (
            f"{name} (証券コード {code}, {market}) "
            f"特別注意市場銘柄に指定 状況: {status} 備考: {note}"
        )[:1500]
        rows.append(
            EnfRow(
                target_name=name,
                issuance_date=date_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=src.url,
                extra={
                    "stock_code": code,
                    "market": market,
                    "review_status": status,
                    "source_section": src.name,
                },
            )
        )
    return rows


# -- 特別注意銘柄(履歴) --


def parse_alert_history(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    for tr in _table_rows(html):
        cells = _row_cells_text(tr)
        # 指定日 / 銘柄名 / コード / 市場区分 / 解除日 / 備考
        if len(cells) < 5:
            continue
        date_iso = _slash_to_iso(cells[0])
        name = _clean_company(cells[1])
        if not date_iso or not name:
            continue
        code = re.sub(r"\D", "", cells[2])[:6]
        market = cells[3]
        kaisho = cells[4]
        note = cells[5] if len(cells) > 5 else ""
        excl_end = _slash_to_iso(kaisho)
        summary = (
            f"{name} (証券コード {code}, {market}) "
            f"特別注意市場銘柄指定履歴 解除: {kaisho} 備考: {note}"
        )[:1500]
        rows.append(
            EnfRow(
                target_name=name,
                issuance_date=date_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=src.url,
                exclusion_start=date_iso,
                exclusion_end=excl_end,
                extra={
                    "stock_code": code,
                    "market": market,
                    "kaisho_date": kaisho,
                    "source_section": src.name,
                },
            )
        )
    return rows


# -- 特別注意銘柄(Excel 履歴) --


def parse_alert_excel(body: bytes, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    if not body:
        return rows
    wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in ws.iter_rows(values_only=True):
            if not r or len(r) < 4:
                continue
            shitei, name_raw, code_raw, market = r[0], r[1], r[2], r[3]
            kaisho = r[4] if len(r) > 4 else None
            note = r[5] if len(r) > 5 else None
            iso = _excel_date_to_iso(shitei)
            name = _clean_company(str(name_raw)) if name_raw else ""
            if not iso or not name:
                continue
            code = re.sub(r"\D", "", str(code_raw or ""))[:6]
            market_s = _normalize(str(market or ""))
            kaisho_iso = _excel_date_to_iso(kaisho)
            note_s = _normalize(str(note or ""))
            summary = (
                f"{name} (証券コード {code}, {market_s}) "
                f"特別注意市場銘柄指定 (2019-2023履歴) 解除: {kaisho_iso or '未解除'} "
                f"備考: {note_s}"
            )[:1500]
            rows.append(
                EnfRow(
                    target_name=name,
                    issuance_date=iso,
                    issuing_authority=src.authority,
                    enforcement_kind=src.enforcement_kind,
                    reason_summary=summary,
                    related_law_ref=src.law_basis,
                    source_url=src.url,
                    exclusion_start=iso,
                    exclusion_end=kaisho_iso,
                    extra={
                        "stock_code": code,
                        "market": market_s,
                        "source_section": src.name,
                        "remarks": note_s[:300],
                    },
                )
            )
    return rows


# -- 公表措置銘柄 --


def parse_public_announce(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    for tr in _table_rows(html):
        cells = _row_cells_text(tr)
        # 公表日 / 銘柄名 / コード / 市場区分 / 詳細 / 備考
        if len(cells) < 4:
            continue
        date_iso = _slash_to_iso(cells[0])
        name = _clean_company(cells[1])
        if not date_iso or not name:
            continue
        code = re.sub(r"\D", "", cells[2])[:6]
        market = cells[3]
        note = cells[5] if len(cells) > 5 else ""
        summary = (f"{name} (証券コード {code}, {market}) 公表措置 備考: {note}")[:1500]
        rows.append(
            EnfRow(
                target_name=name,
                issuance_date=date_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=src.url,
                extra={
                    "stock_code": code,
                    "market": market,
                    "source_section": src.name,
                },
            )
        )
    return rows


# -- 上場契約違約金徴求 --


def parse_agreement_violation(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    for tr in _table_rows(html):
        cells = _row_cells_text(tr)
        # 公表日 / 銘柄名 / コード / 市場区分 / 詳細 / 備考
        if len(cells) < 4:
            continue
        date_iso = _slash_to_iso(cells[0])
        name = _clean_company(cells[1])
        if not date_iso or not name:
            continue
        code = re.sub(r"\D", "", cells[2])[:6]
        market = cells[3]
        note = cells[5] if len(cells) > 5 else ""
        amt = _parse_yen(note)
        summary = (f"{name} (証券コード {code}, {market}) 上場契約違約金徴求 備考: {note}")[:1500]
        rows.append(
            EnfRow(
                target_name=name,
                issuance_date=date_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=src.url,
                amount_yen=amt,
                extra={
                    "stock_code": code,
                    "market": market,
                    "source_section": src.name,
                },
            )
        )
    return rows


# -- 監理・整理銘柄 --


def parse_supervision(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    for tr in _table_rows(html):
        cells = _row_cells_text(tr)
        # 指定日 / 銘柄名 / コード / 市場区分 / 詳細 / 区分 / 指定理由 / 備考
        # The table header order varies; we find the date and name in
        # whichever cell looks like a slash-date / company.
        if len(cells) < 4:
            continue
        date_iso = _slash_to_iso(cells[0])
        name = _clean_company(cells[1])
        if not date_iso or not name:
            continue
        code = re.sub(r"\D", "", cells[2])[:6]
        market = cells[3]
        kbn = cells[5] if len(cells) > 5 else ""
        reason = cells[6] if len(cells) > 6 else ""
        note = cells[7] if len(cells) > 7 else ""
        summary = (
            f"{name} (証券コード {code}, {market}) "
            f"監理銘柄等指定 区分: {kbn} 理由: {reason} 備考: {note}"
        )[:1500]
        rows.append(
            EnfRow(
                target_name=name,
                issuance_date=date_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=src.url,
                extra={
                    "stock_code": code,
                    "market": market,
                    "supervision_kind": kbn,
                    "designation_reason": reason,
                    "source_section": src.name,
                },
            )
        )
    return rows


# -- 猶予期間入り銘柄 (HTML) --


def parse_grace_period(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    for tr in _table_rows(html):
        cells = _row_cells_text(tr)
        # 銘柄名 / コード / 市場区分 / 詳細 / 猶予期間 / 備考
        # Note: index 0 is name (no public date), so use 猶予期間 as date.
        if len(cells) < 4:
            continue
        name = _clean_company(cells[0])
        if not name:
            continue
        code = re.sub(r"\D", "", cells[1])[:6]
        market = cells[2]
        period = cells[4] if len(cells) > 4 else ""
        note = cells[5] if len(cells) > 5 else ""
        period_dates = SLASH_DATE_RE.findall(period or "")
        excl_start = excl_end = None
        if period_dates:
            y, m, d = (int(x) for x in period_dates[0])
            excl_start = f"{y:04d}-{m:02d}-{d:02d}"
        if len(period_dates) >= 2:
            y, m, d = (int(x) for x in period_dates[1])
            excl_end = f"{y:04d}-{m:02d}-{d:02d}"
        date_iso = excl_start
        if not date_iso:
            continue
        summary = (
            f"{name} (証券コード {code}, {market}) "
            f"上場廃止基準猶予期間入り 期間: {period} 備考: {note}"
        )[:1500]
        rows.append(
            EnfRow(
                target_name=name,
                issuance_date=date_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=src.url,
                exclusion_start=excl_start,
                exclusion_end=excl_end,
                extra={
                    "stock_code": code,
                    "market": market,
                    "grace_period": period,
                    "source_section": src.name,
                },
            )
        )
    return rows


# -- 猶予期間 Excel 履歴 --


def parse_grace_excel(body: bytes, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    if not body:
        return rows
    wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        # Header: 公表日 / 銘柄名 / コード / 市場区分 / 猶予期間入り等理由 /
        #         解除公表日 / 備考
        for r in ws.iter_rows(values_only=True):
            if not r or len(r) < 5:
                continue
            kouhyou, name_raw, code_raw, market, reason = r[:5]
            kaisho = r[5] if len(r) > 5 else None
            note = r[6] if len(r) > 6 else None
            iso = _excel_date_to_iso(kouhyou)
            name = _clean_company(str(name_raw)) if name_raw else ""
            if not iso or not name:
                continue
            code = re.sub(r"\D", "", str(code_raw or ""))[:6]
            market_s = _normalize(str(market or ""))
            reason_s = _normalize(str(reason or ""))
            kaisho_iso = _excel_date_to_iso(kaisho)
            note_s = _normalize(str(note or ""))
            summary = (
                f"{name} (証券コード {code}, {market_s}) "
                f"上場廃止基準猶予期間入り(履歴) 理由: {reason_s} "
                f"解除: {kaisho_iso or '未解除'} 備考: {note_s}"
            )[:1500]
            rows.append(
                EnfRow(
                    target_name=name,
                    issuance_date=iso,
                    issuing_authority=src.authority,
                    enforcement_kind=src.enforcement_kind,
                    reason_summary=summary,
                    related_law_ref=src.law_basis,
                    source_url=src.url,
                    exclusion_start=iso,
                    exclusion_end=kaisho_iso,
                    extra={
                        "stock_code": code,
                        "market": market_s,
                        "grace_reason": reason_s,
                        "source_section": src.name,
                        "remarks": note_s[:300],
                    },
                )
            )
    return rows


# -- 改善期間該当銘柄 (HTML) --


def parse_improvement_period(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    for tr in _table_rows(html):
        cells = _row_cells_text(tr)
        if len(cells) < 4:
            continue
        # Schema is similar to grace-period but lighter
        name = _clean_company(cells[0])
        if not name:
            continue
        code = re.sub(r"\D", "", cells[1])[:6]
        market = cells[2]
        # Try to find a date in any later cell
        date_iso = None
        for c in cells[3:]:
            date_iso = _slash_to_iso(c)
            if date_iso:
                break
        if not date_iso:
            continue
        note = cells[-1] if cells else ""
        summary = (f"{name} (証券コード {code}, {market}) 改善期間該当 備考: {note}")[:1500]
        rows.append(
            EnfRow(
                target_name=name,
                issuance_date=date_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=src.url,
                extra={
                    "stock_code": code,
                    "market": market,
                    "source_section": src.name,
                },
            )
        )
    return rows


# -- 改善期間 Excel --


def parse_improvement_period_excel(body: bytes, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    if not body:
        return rows
    wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        # Layout (variable). Iterate rows; pick any row whose row contains
        # both a 銘柄名 and at least one 公表日/期間 date.
        for r in ws.iter_rows(values_only=True):
            if not r:
                continue
            # find first date cell + first 銘柄-shaped cell
            iso = None
            name = None
            code = ""
            market = ""
            for c in r:
                if iso is None:
                    iso = _excel_date_to_iso(c)
                if (
                    name is None
                    and isinstance(c, str)
                    and ("（株）" in c or "株式会社" in c or "(株)" in c)
                ):
                    name = _clean_company(c)
            if not iso or not name:
                continue
            # try to find code (4-digit number) and market name
            for c in r:
                if isinstance(c, (int, float)):
                    cc = str(int(c))
                    if len(cc) == 4 or len(cc) == 5:
                        code = cc
                        break
                elif isinstance(c, str):
                    cs = c.strip()
                    if re.fullmatch(r"\d{4,5}", cs):
                        code = cs
                        break
            for c in r:
                if isinstance(c, str) and c.strip() in (
                    "プライム",
                    "スタンダード",
                    "グロース",
                    "JASDAQスタンダード",
                    "JASDAQグロース",
                    "東証一部",
                    "東証二部",
                ):
                    market = c.strip()
                    break
            note_chunks = [_normalize(str(c)) for c in r if isinstance(c, str) and c.strip()]
            note = " | ".join(note_chunks)[:600]
            summary = (f"{name} (証券コード {code}, {market}) 改善期間該当銘柄 詳細: {note}")[:1500]
            rows.append(
                EnfRow(
                    target_name=name,
                    issuance_date=iso,
                    issuing_authority=src.authority,
                    enforcement_kind=src.enforcement_kind,
                    reason_summary=summary,
                    related_law_ref=src.law_basis,
                    source_url=src.url,
                    extra={
                        "stock_code": code,
                        "market": market,
                        "source_section": src.name,
                    },
                )
            )
    return rows


# -- CPAAOB 監査法人勧告 --

CPAAOB_PDF_RE = re.compile(r'/cpaaob/sonota/houdou/[^"]+\.pdf')
CPAAOB_TITLE_RE = re.compile(
    r"(.+?)に対する検査結果に基づく勧告について(?:\s*[（(]([^)）]+)[)）])?"
)


def parse_cpaaob_kankoku(html: str, src: JpxSource) -> list[EnfRow]:
    rows: list[EnfRow] = []
    soup = BeautifulSoup(html, "html.parser")
    seen_keys: set[tuple[str, str]] = set()
    for a in soup.find_all("a"):
        href = a.get("href", "") or ""
        if not href.endswith(".pdf"):
            continue
        if "/cpaaob/sonota/houdou/" not in href:
            continue
        text = _normalize(a.get_text(" ", strip=True))
        m = CPAAOB_TITLE_RE.search(text)
        if not m:
            continue
        target = _clean_company(m.group(1))
        date_blob = m.group(2) or ""
        date_iso = _wareki_to_iso(date_blob) or _wareki_to_iso(text)
        if not date_iso:
            continue
        if not target:
            continue
        key = (target, date_iso)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        url = urljoin(src.url, href)
        summary = (
            f"{target} に対する検査結果に基づく勧告 ({date_blob}) "
            f"金融庁長官に対し公認会計士法第41条の2に基づき勧告。"
        )[:1500]
        rows.append(
            EnfRow(
                target_name=target,
                issuance_date=date_iso,
                issuing_authority=src.authority,
                enforcement_kind=src.enforcement_kind,
                reason_summary=summary,
                related_law_ref=src.law_basis,
                source_url=url,
                extra={
                    "kankoku_date": date_blob,
                    "source_section": src.name,
                    "title_raw": text[:300],
                },
            )
        )
    return rows


# ---------------------------------------------------------------------------
# Source dispatch
# ---------------------------------------------------------------------------


PARSER_DISPATCH = {
    "delisted": ("html", parse_delisted),
    "improvement_report": ("html", parse_improvement_report),
    "alert": ("html", parse_alert),
    "alert_history": ("html", parse_alert_history),
    "alert_excel": ("bytes", parse_alert_excel),
    "public_announce": ("html", parse_public_announce),
    "agreement_violation": ("html", parse_agreement_violation),
    "supervision": ("html", parse_supervision),
    "grace_period": ("html", parse_grace_period),
    "grace_excel": ("bytes", parse_grace_excel),
    "improvement_period": ("html", parse_improvement_period),
    "improvement_period_excel": ("bytes", parse_improvement_period_excel),
    "cpaaob_kankoku": ("html", parse_cpaaob_kankoku),
}


def collect_rows(fetcher: Fetcher) -> list[EnfRow]:
    out: list[EnfRow] = []
    for src in SOURCES:
        kind, fn = PARSER_DISPATCH[src.parser]
        if kind == "html":
            html = fetcher.get_text(src.url)
            if not html:
                _LOG.warning("[%s] empty fetch %s", src.name, src.url)
                continue
            try:
                got = fn(html, src)
            except Exception as exc:  # parsing edge guard
                _LOG.error("[%s] parse error: %s", src.name, exc)
                continue
        else:
            body = fetcher.get_bytes(src.url)
            if not body:
                _LOG.warning("[%s] empty fetch %s", src.name, src.url)
                continue
            try:
                got = fn(body, src)
            except Exception as exc:
                _LOG.error("[%s] parse error: %s", src.name, exc)
                continue
        _LOG.info("[%s] parsed=%d", src.name, len(got))
        out.extend(got)
    return out


# ---------------------------------------------------------------------------
# DB layer
# ---------------------------------------------------------------------------


def ensure_tables(conn: sqlite3.Connection) -> None:
    for tbl in ("am_entities", "am_enforcement_detail"):
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (tbl,),
        ).fetchone()
        if not row:
            raise SystemExit(f"missing table '{tbl}' — apply migrations first")


def existing_dedup_keys(conn: sqlite3.Connection) -> set[tuple[str, str, str]]:
    """Used both for our own JPX/CPAAOB dedup and to skip FSA #21 overlaps for
    the fine kind."""
    out: set[tuple[str, str, str]] = set()
    cur = conn.execute(
        "SELECT issuing_authority, issuance_date, target_name FROM am_enforcement_detail"
    )
    for a, d, n in cur.fetchall():
        if a and d and n:
            out.add((a, d, n))
    return out


def existing_fine_keys(conn: sqlite3.Connection) -> set[tuple[str, str, int | None]]:
    """For dedup of `fine` rows against #21 FSA: (date, target_name, amount)."""
    out: set[tuple[str, str, int | None]] = set()
    cur = conn.execute(
        "SELECT issuance_date, target_name, amount_yen "
        "FROM am_enforcement_detail "
        "WHERE enforcement_kind='fine'"
    )
    for d, n, a in cur.fetchall():
        if d and n:
            out.add((d, n, a))
    return out


def upsert_entity_and_enforcement(
    conn: sqlite3.Connection,
    row: EnfRow,
    seq: int,
    now_iso: str,
) -> None:
    slug = _slug8(row.issuing_authority, row.issuance_date, row.target_name)
    canonical_id = f"AM-ENF-JPX-{slug}{seq:04d}"
    domain = urlparse(row.source_url).netloc or None
    primary_name = (
        f"{row.target_name} ({row.issuance_date}) - {row.issuing_authority} {row.enforcement_kind}"
    )[:500]
    raw_json = json.dumps(
        {
            "target_name": row.target_name,
            "issuance_date": row.issuance_date,
            "issuing_authority": row.issuing_authority,
            "enforcement_kind": row.enforcement_kind,
            "related_law_ref": row.related_law_ref,
            "amount_yen": row.amount_yen,
            "exclusion_start": row.exclusion_start,
            "exclusion_end": row.exclusion_end,
            "reason_summary": row.reason_summary,
            "source_url": row.source_url,
            "extra": row.extra,
            "source_attribution": (
                "日本取引所グループ ウェブサイト"
                if "jpx.co.jp" in (domain or "")
                else "金融庁 公認会計士・監査審査会 ウェブサイト"
            ),
            "license": "政府機関の著作物（出典明記で転載引用可）",
        },
        ensure_ascii=False,
    )
    conn.execute(
        """
        INSERT INTO am_entities (
            canonical_id, record_kind, source_topic, source_record_index,
            primary_name, authority_canonical, confidence,
            source_url, source_url_domain, fetched_at, raw_json,
            canonical_status, citation_status
        ) VALUES (?, 'enforcement', 'jpx_listing_action', NULL,
                  ?, NULL, 0.9, ?, ?, ?, ?, 'active', 'ok')
        ON CONFLICT(canonical_id) DO UPDATE SET
            primary_name      = excluded.primary_name,
            source_url        = excluded.source_url,
            source_url_domain = excluded.source_url_domain,
            fetched_at        = excluded.fetched_at,
            raw_json          = excluded.raw_json,
            updated_at        = datetime('now')
        """,
        (
            canonical_id,
            primary_name,
            row.source_url,
            domain,
            now_iso,
            raw_json,
        ),
    )
    conn.execute(
        """
        INSERT INTO am_enforcement_detail (
            entity_id, houjin_bangou, target_name, enforcement_kind,
            issuing_authority, issuance_date, exclusion_start, exclusion_end,
            reason_summary, related_law_ref, amount_yen,
            source_url, source_fetched_at
        ) VALUES (?, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            canonical_id,
            row.target_name[:500],
            row.enforcement_kind,
            row.issuing_authority,
            row.issuance_date,
            row.exclusion_start,
            row.exclusion_end,
            row.reason_summary[:4000],
            row.related_law_ref[:1000],
            row.amount_yen,
            row.source_url,
            now_iso,
        ),
    )


def write_rows(
    conn: sqlite3.Connection,
    rows: list[EnfRow],
    *,
    now_iso: str,
    max_inserts: int,
) -> tuple[int, int, int, int]:
    if not rows:
        return 0, 0, 0, 0
    db_keys = existing_dedup_keys(conn)
    fine_keys = existing_fine_keys(conn)
    batch_keys: set[tuple[str, str, str]] = set()
    inserted = dup_db = dup_batch = dup_fine = 0
    try:
        conn.execute("BEGIN IMMEDIATE")
        for r in rows:
            if inserted >= max_inserts:
                break
            key = (r.issuing_authority, r.issuance_date, r.target_name)
            if key in db_keys:
                dup_db += 1
                continue
            if key in batch_keys:
                dup_batch += 1
                continue
            # Cross-source guard: if a 'fine' with same target+date already in
            # FSA #21, skip JPX 上場契約違約金 record. (Defensive — JPX violation
            # money is paid to TSE, FSA fine to government, but same listed
            # company name would be the same enforcement event window.)
            if r.enforcement_kind == "fine":
                if (r.issuance_date, r.target_name, r.amount_yen) in fine_keys:
                    dup_fine += 1
                    continue
            batch_keys.add(key)
            try:
                upsert_entity_and_enforcement(conn, r, inserted, now_iso)
                inserted += 1
            except sqlite3.Error as exc:
                _LOG.error(
                    "insert error name=%r date=%s err=%s",
                    r.target_name,
                    r.issuance_date,
                    exc,
                )
                continue
        conn.commit()
    except sqlite3.Error as exc:
        _LOG.error("BEGIN/commit failed: %s", exc)
        with contextlib.suppress(sqlite3.Error):
            conn.rollback()
    return inserted, dup_db, dup_batch, dup_fine


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--db", type=Path, default=DEFAULT_DB)
    ap.add_argument("--max-rows", type=int, default=600)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--verbose", "-v", action="store_true")
    return ap.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )
    fetcher = Fetcher()
    now_iso = datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")

    rows = collect_rows(fetcher)
    fetcher.close()
    _LOG.info("total parsed=%d", len(rows))

    if args.dry_run:
        # breakdown by source_section
        by_section: dict[str, int] = {}
        by_authority: dict[str, int] = {}
        by_kind: dict[str, int] = {}
        for r in rows:
            sec = r.extra.get("source_section", "?")
            by_section[sec] = by_section.get(sec, 0) + 1
            by_authority[r.issuing_authority] = by_authority.get(r.issuing_authority, 0) + 1
            by_kind[r.enforcement_kind] = by_kind.get(r.enforcement_kind, 0) + 1
        _LOG.info("--- by section ---")
        for k, v in sorted(by_section.items(), key=lambda x: -x[1]):
            _LOG.info("  %-32s %d", k, v)
        _LOG.info("--- by authority ---")
        for k, v in sorted(by_authority.items(), key=lambda x: -x[1]):
            _LOG.info("  %-32s %d", k, v)
        _LOG.info("--- by kind ---")
        for k, v in sorted(by_kind.items(), key=lambda x: -x[1]):
            _LOG.info("  %-32s %d", k, v)
        for r in rows[:5]:
            _LOG.info(
                "sample: name=%s date=%s auth=%s kind=%s law=%s",
                r.target_name,
                r.issuance_date,
                r.issuing_authority,
                r.enforcement_kind,
                r.related_law_ref,
            )
        return 0

    if not args.db.exists():
        _LOG.error("autonomath.db missing: %s", args.db)
        return 2
    conn = sqlite3.connect(str(args.db))
    conn.execute("PRAGMA busy_timeout=300000")
    conn.execute("PRAGMA foreign_keys=ON")
    ensure_tables(conn)

    inserted, dup_db, dup_batch, dup_fine = write_rows(
        conn,
        rows,
        now_iso=now_iso,
        max_inserts=args.max_rows,
    )

    # Authority + law breakdown of inserted rows (post-write)
    cur = conn.execute(
        "SELECT issuing_authority, related_law_ref, enforcement_kind, COUNT(*) "
        "FROM am_enforcement_detail "
        "WHERE source_fetched_at = ? "
        "GROUP BY issuing_authority, related_law_ref, enforcement_kind "
        "ORDER BY 4 DESC",
        (now_iso,),
    )
    breakdown = cur.fetchall()
    with contextlib.suppress(sqlite3.Error):
        conn.close()

    print(
        f"JPX enforcement ingest: parsed={len(rows)} inserted={inserted} "
        f"dup_db={dup_db} dup_batch={dup_batch} dup_fine={dup_fine}"
    )
    print("--- breakdown by authority/law/kind ---")
    for auth, law, kind, cnt in breakdown:
        print(f"  {cnt:>4}  {auth} | {kind} | {law}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
